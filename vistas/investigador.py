import logging
import os
import subprocess
import uuid
from pathlib import Path
from datetime import datetime
import json

from ia.constants import MIN_MICROGLIA_SIZE

from bd.database import conectar

from PyQt6.QtCore import pyqtSignal, QRect, Qt, QSize, QEvent, QPoint
from PyQt6.QtGui import QColor, QImage, QPainter, QPen, QPixmap, QIcon, QIntValidator
from PyQt6.QtWidgets import (
    QApplication,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QPushButton,
    QVBoxLayout,
    QWidget,
    QSlider,
    QCheckBox,
    QLineEdit,
    QGridLayout,
    QMessageBox,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QAbstractItemView,
    QToolTip,
    QTreeWidget,
    QTreeWidgetItem,
    QScrollArea,
)
from PyQt6.QtCore import QTimer

# os.environ["QT_QPA_PLATFORM"] = "xcb"


# MorphologyAnalyzer will be imported inside methods that use it

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
# Silence chatty libraries
logging.getLogger("PIL").setLevel(logging.WARNING)
logging.getLogger("fpdf").setLevel(logging.WARNING)

# =========================================================
# BOTÓN PERSONALIZADO PARA TOOLTIPS SEGUROS (ANTI-CORTES)
# =========================================================
class SafeToolTipButton(QPushButton):
    """
    Un botón que fuerza al ToolTip a mostrarse desplazado hacia la izquierda.
    Garantiza que el texto nunca se corte en el borde derecho de la pantalla (WSL/Pantalla Completa).
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.custom_tooltip = ""

    def setCustomToolTip(self, text):
        self.custom_tooltip = text
        self.setToolTip(text)

    def event(self, e):
        if e.type() == QEvent.Type.ToolTip:
            if self.custom_tooltip:
                global_pos = self.mapToGlobal(QPoint(0, self.height() + 2))
                global_pos.setX(global_pos.x() - 80)
                QToolTip.showText(global_pos, self.custom_tooltip, self)
            return True
        return super().event(e)


# =========================================================
# ZONA DE ARRASTRAR Y SOLTAR (DRAG & DROP)
# =========================================================
class DropZone(QFrame):
    file_selected = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        
        self.setObjectName("DropZoneObj")
        
        self.layout = QVBoxLayout(self)
        self.layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.layout.setContentsMargins(10, 10, 10, 15) 
        
        self.lbl_icono = QLabel()
        self.lbl_icono.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_icono.setStyleSheet("border: none; background: transparent;")
        self.lbl_icono.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        
        self.lbl_texto = QLabel("Haz clic para cargar una imagen o arrástrala aquí")
        self.lbl_texto.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_texto.setWordWrap(True)
        self.lbl_texto.setStyleSheet("color: #888888; font-size: 12px; font-weight: normal; border: none; background: transparent;")
        self.lbl_texto.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        
        self.layout.addWidget(self.lbl_icono, stretch=1)
        self.layout.addWidget(self.lbl_texto)
        
        self.pixmap_original = None
        self.is_preview = False
        
        self.mostrar_placeholder()

    def mostrar_placeholder(self):
        self.is_preview = False
        self.pixmap_original = QPixmap("assets/cargar.png")
        
        self.lbl_texto.setText("Haz clic para cargar una imagen o arrástrala aquí")
        self.lbl_texto.setStyleSheet("color: #888888; font-size: 12px; font-weight: normal; border: none; background: transparent;")
        
        self.setStyleSheet("""
            #DropZoneObj {
                border: 2px dashed #0969da;
                border-radius: 8px;
                background-color: #ffffff;
            }
            #DropZoneObj:hover {
                background-color: #f6f8fa;
                border: 2px dashed #0550ae;
            }
        """)
        self.actualizar_imagen()

    def mostrar_imagen(self, pixmap, nombre_archivo):
        self.is_preview = True
        self.pixmap_original = pixmap
        
        self.lbl_texto.setText(f"Archivo seleccionado: <b>{nombre_archivo}</b>")
        self.lbl_texto.setStyleSheet("color: #333333; font-size: 12px; border: none; background: transparent;")
        
        self.setStyleSheet("""
            #DropZoneObj {
                border: 2px solid #28a745;
                border-radius: 8px;
                background-color: #ffffff;
            }
        """)
        self.actualizar_imagen()

    def actualizar_imagen(self):
        if self.pixmap_original and not self.pixmap_original.isNull():
            w = self.width() - 20
            h = self.height() - 60 
            
            if w > 0 and h > 0:
                self.lbl_icono.setPixmap(self.pixmap_original.scaled(
                    w, h, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation
                ))

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.actualizar_imagen()

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls(): event.accept()
        else: event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls(): event.accept()
        else: event.ignore()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            event.setDropAction(Qt.DropAction.CopyAction)
            event.accept()
            urls = event.mimeData().urls()
            if urls and urls[0].isLocalFile():
                file_path = urls[0].toLocalFile()
                ext = Path(file_path).suffix.lower()
                if ext in [".tif", ".tiff", ".png", ".jpg", ".jpeg"]:
                    self.file_selected.emit(file_path)
        else:
            event.ignore()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            ruta_archivo, _ = QFileDialog.getOpenFileName(
                self, "Seleccionar imagen", "", "Imágenes TIFF (*.tiff *.tif);;Todas las imágenes (*.png *.jpg *.jpeg)"
            )
            if ruta_archivo:
                self.file_selected.emit(ruta_archivo)


# =========================================================
# DIÁLOGO INTERMEDIO PARA CARGA DE DATOS
# =========================================================
class DialogoCargarImagen(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Detalles del Campo")
        self.setFixedSize(580, 520) 
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True)
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        
        self.ruta_seleccionada = None
        self.campo_val = ""
        self.tiempo_val = ""

        main_layout = QVBoxLayout(self)
        frame = QFrame(self)
        frame.setStyleSheet("QFrame { background-color: #FFFFFF; border-radius: 12px; border: 1px solid #d0d7de; } QLabel { border: none; }")
        layout = QVBoxLayout(frame)
        
        header_layout = QHBoxLayout()
        lbl_titulo = QLabel("Registro del Campo")
        lbl_titulo.setStyleSheet("font-size: 18px; font-weight: bold; color: #0969da; border: none;")
        lbl_titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        header_layout.setContentsMargins(10, 10, 10, 0)
        self.btn_cerrar_x = QPushButton()
        self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))
        self.btn_cerrar_x.setIconSize(QSize(20, 20))
        self.btn_cerrar_x.setFixedSize(35, 35)
        self.btn_cerrar_x.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cerrar_x.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #f6f8fa; border-radius: 17px; }
        """)
        self.btn_cerrar_x.enterEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xroja.png"))
        self.btn_cerrar_x.leaveEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))
        self.btn_cerrar_x.clicked.connect(self.reject)
        
        header_layout.addStretch()
        header_layout.addWidget(lbl_titulo)
        header_layout.addStretch()
        header_layout.addWidget(self.btn_cerrar_x)
        layout.addLayout(header_layout)
        layout.addSpacing(10)

        layout_campos = QHBoxLayout()
        
        layout_campo = QVBoxLayout()
        lbl_campo = QLabel("Campo:")
        lbl_campo.setStyleSheet("font-weight: bold; font-size: 13px;")
        self.input_campo = QLineEdit()
        self.input_campo.setPlaceholderText("Ej. Campo A, 1...")
        layout_campo.addWidget(lbl_campo)
        layout_campo.addWidget(self.input_campo)

        layout_tiempo = QVBoxLayout()
        lbl_tiempo = QLabel("Tiempo:")
        lbl_tiempo.setStyleSheet("font-weight: bold; font-size: 12px; color: #57606a;")
        self.input_tiempo = QLineEdit()
        self.input_tiempo.setPlaceholderText("Ej. 1hr, 2hrs...")
        layout_tiempo.addWidget(lbl_tiempo)
        layout_tiempo.addWidget(self.input_tiempo)

        layout_campos.addLayout(layout_campo)
        layout_campos.addSpacing(15)
        layout_campos.addLayout(layout_tiempo)
        
        layout.addLayout(layout_campos)
        layout.addSpacing(15)

        self.drop_zone = DropZone(self)
        self.drop_zone.file_selected.connect(self.procesar_archivo)
        layout.addWidget(self.drop_zone, stretch=1)
        layout.addSpacing(10)

        layout_botones = QHBoxLayout()
        
        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_cancelar.clicked.connect(self.reject)
        
        btn_continuar = QPushButton("Cargar Imagen")
        btn_continuar.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_continuar.setStyleSheet("background-color: #2da44e; color: white; border: 1px solid #1a7f37;")
        btn_continuar.clicked.connect(self.validar_y_continuar)
        
        layout_botones.addStretch()
        layout_botones.addWidget(btn_continuar)
        layout_botones.addStretch()
        
        layout.addLayout(layout_botones)
        main_layout.addWidget(frame)

    def showEvent(self, event):
        super().showEvent(event)
        if self.parent():
            p_geom = self.parent().geometry()
            self.move(p_geom.x() + (p_geom.width() - self.width()) // 2, p_geom.y() + (p_geom.height() - self.height()) // 2)
        else:
            screen = QApplication.primaryScreen().geometry()
            self.move((screen.width() - self.width()) // 2, (screen.height() - self.height()) // 2)

    def procesar_archivo(self, ruta):
        self.ruta_seleccionada = ruta
        nombre_archivo = os.path.basename(ruta)
        
        try:
            import cv2
            import numpy as np
            cv_img = cv2.imread(ruta, cv2.IMREAD_UNCHANGED)
            if cv_img is not None:
                if cv_img.dtype == np.uint16:
                    cv_img = ((cv_img - cv_img.min()) / (cv_img.max() - cv_img.min()) * 255).astype(np.uint8)
                if len(cv_img.shape) == 2:
                    h, w = cv_img.shape
                    qimg = QImage(cv_img.data, w, h, w, QImage.Format.Format_Grayscale8)
                else:
                    cv_img = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
                    h, w, ch = cv_img.shape
                    qimg = QImage(cv_img.data, w, h, ch * w, QImage.Format.Format_RGB888)
                pixmap_preview = QPixmap.fromImage(qimg)
            else:
                pixmap_preview = QPixmap(ruta)
        except Exception as e:
            logging.error(f"Error al previsualizar: {e}")
            pixmap_preview = QPixmap(ruta)

        if not pixmap_preview.isNull():
            self.drop_zone.mostrar_imagen(pixmap_preview, nombre_archivo)

    def validar_y_continuar(self):
        if not self.input_campo.text().strip() or not self.input_tiempo.text().strip() or not self.ruta_seleccionada:
            from vistas.utilidades import DialogoNotificacion
            DialogoNotificacion("Error", "Campos incompletos", "warning", self).exec()
            return
        
        self.campo_val = self.input_campo.text().strip()
        self.tiempo_val = self.input_tiempo.text().strip()
        self.accept()


# =========================================================
# CLASES AUXILIARES Y VISOR INTERACTIVO
# =========================================================
class DialogoCarga(QDialog):
    def __init__(self, mensaje="Procesando...", parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True)
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        
        layout = QVBoxLayout(self)
        frame = QFrame(self)
        frame.setStyleSheet("""
            QFrame { background-color: #FFFFFF; border-radius: 12px; border: 2px solid #003366; }
            QLabel { color: #003366; font-size: 16px; font-weight: bold; padding: 20px; }
        """)
        flayout = QVBoxLayout(frame)
        label = QLabel(mensaje)
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        flayout.addWidget(label)
        layout.addWidget(frame)
        self.setLayout(layout)

    def showEvent(self, event):
        super().showEvent(event)
        if self.parent():
            p_geom = self.parent().geometry()
            self.move(p_geom.x() + (p_geom.width() - self.width()) // 2, p_geom.y() + (p_geom.height() - self.height()) // 2)
        else:
            screen = QApplication.primaryScreen().geometry()
            self.move((screen.width() - self.width()) // 2, (screen.height() - self.height()) // 2)


class DialogoComparativo(QDialog):
    """
    Muestra las 3 fases del proceso lado a lado para una microglía,
    o una versión sobrepuesta (esqueleto sobre original).
    """
    def __init__(self, fases, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Proceso de la Microglía")
        self.setFixedSize(950, 500) 
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        
        self.fases = fases
        
        main_layout = QVBoxLayout(self)
        self.frame = QFrame(self)
        self.frame.setStyleSheet("QFrame { background-color: #FFFFFF; border-radius: 12px; border: 2px solid #003366; } QLabel { border: none; }")
        self.layout_principal = QVBoxLayout(self.frame)
        
        header_layout = QHBoxLayout()
        self.lbl_titulo = QLabel("<b>Comparativa del Proceso</b>")
        self.lbl_titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_titulo.setStyleSheet("font-size: 18px; color: #003366; border: none;")
        
        header_layout.setContentsMargins(10, 10, 10, 0)
        self.btn_cerrar_x = QPushButton()
        self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))
        self.btn_cerrar_x.setIconSize(QSize(20, 20))
        self.btn_cerrar_x.setFixedSize(35, 35)
        self.btn_cerrar_x.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cerrar_x.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #f6f8fa; border-radius: 17px; }
        """)
        self.btn_cerrar_x.enterEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xroja.png"))
        self.btn_cerrar_x.leaveEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))

        self.btn_cerrar_x.clicked.connect(self.accept)
        
        header_layout.addStretch()
        header_layout.addWidget(self.lbl_titulo)
        header_layout.addStretch()
        header_layout.addWidget(self.btn_cerrar_x)
        self.layout_principal.addLayout(header_layout)
        self.layout_principal.addSpacing(10)

        # Widget para vista lado a lado
        self.widget_lado_lado = QWidget()
        self.layout_lado_lado = QHBoxLayout(self.widget_lado_lado)
        for fase in self.fases:
            v_layout = QVBoxLayout()
            lbl_nombre_fase = QLabel(fase["nombre"])
            lbl_nombre_fase.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl_nombre_fase.setStyleSheet("font-weight: bold; color: #555; margin-bottom: 5px;")
            
            lbl_img = QLabel()
            lbl_img.setAlignment(Qt.AlignmentFlag.AlignCenter)
            pixmap = fase["pixmap"] if fase["pixmap"] else QPixmap(fase["path"])
            if pixmap and not pixmap.isNull():
                lbl_img.setPixmap(pixmap.scaled(280, 280, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
            else:
                lbl_img.setText("N/D")
            v_layout.addWidget(lbl_nombre_fase)
            v_layout.addWidget(lbl_img)
            self.layout_lado_lado.addLayout(v_layout)
        
        self.layout_principal.addWidget(self.widget_lado_lado)
        
        self.layout_principal.addSpacing(15)
        
        main_layout.addWidget(self.frame)
        self.setLayout(main_layout)

    def ajustar_posicion(self):
        if self.parent():
            p_geom = self.parent().geometry()
            self.move(p_geom.x() + (p_geom.width() - self.width()) // 2, p_geom.y() + (p_geom.height() - self.height()) // 2)

    def showEvent(self, event):
        super().showEvent(event)
        self.ajustar_posicion()


class InteractiveLabelDetail(QLabel):
    def __init__(self, parent_dialog):
        super().__init__()
        self.parent_dialog = parent_dialog
        
        # Crear cursor de lápiz
        from PyQt6.QtGui import QPixmap, QPainter, QPen, QCursor
        from PyQt6.QtCore import Qt
        pixmap = QPixmap(16, 16)
        pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(pixmap)
        painter.setPen(QPen(Qt.GlobalColor.white, 1.5))
        # Dibujar un lapiz simple blanco
        painter.drawLine(0, 15, 5, 10)
        painter.drawLine(5, 10, 15, 0)
        painter.drawLine(15, 0, 16, 1)
        painter.drawLine(16, 1, 6, 11)
        painter.drawLine(6, 11, 0, 15)
        painter.end()
        self.pencil_cursor = QCursor(pixmap, 0, 15)
        self.is_drawing = False
        self.start_pos = None
        self.current_pos = None
        self.mode = "pointer" 
        self.setMouseTracking(True)
        self.pencil_path = [] # Lista de puntos para unir ramas
        self.is_dragging_point = False
        self.dragging_point = None

    def get_point_in_label_coords(self, ox, oy):
        from PyQt6.QtCore import QPoint
        pix = self.pixmap()
        if not pix or pix.isNull(): return None
        lbl_w, lbl_h = self.width(), self.height()
        pix_w, pix_h = pix.width(), pix.height()
        dx = (lbl_w - pix_w) // 2; dy = (lbl_h - pix_h) // 2
        orig_w, orig_h = self.parent_dialog.get_original_crop_size()
        if orig_w == 0 or orig_h == 0: return None
        scale_x = pix_w / orig_w; scale_y = pix_h / orig_h
        return QPoint(int(ox * scale_x) + dx, int(oy * scale_y) + dy)
        
    def point_to_segment_dist_squared(self, p, a, b):
        x0, y0 = p.x(), p.y()
        x1, y1 = a.x(), a.y()
        x2, y2 = b.x(), b.y()
        dx = x2 - x1
        dy = y2 - y1
        if dx == 0 and dy == 0:
            return (x0 - x1)**2 + (y0 - y1)**2
        t = ((x0 - x1) * dx + (y0 - y1) * dy) / float(dx*dx + dy*dy)
        t = max(0.0, min(1.0, t))
        px = x1 + t * dx
        py = y1 + t * dy
        return (x0 - px)**2 + (y0 - py)**2

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            event.accept()
            if self.mode == "eraser":
                # Primero verificar si el clic es para eliminar un área existente
                coords = self.get_rect_in_original_coords(event.pos(), event.pos())
                if coords:
                    ox, oy = coords["x"], coords["y"]
                    areas = self.parent_dialog.box.get("removal_areas", [])
                    for i, area in enumerate(areas):
                        if area["x"] <= ox <= area["x"]+area["w"] and area["y"] <= oy <= area["y"]+area["h"]:
                            areas.pop(i)
                            self.parent_dialog.actualizar_visibilidad_boton_limpieza()
                            self.update()
                            return
                
                self.is_drawing = True
                self.start_pos = event.pos()
                self.current_pos = event.pos()
                self.update()
            elif self.mode == "pencil_line":
                click_pos = event.pos()
                connections = self.parent_dialog.box.get("manual_connections", [])
                
                for p_idx, path in enumerate(connections):
                    for pt_idx, pt in enumerate(path):
                        lbl_pt = self.get_point_in_label_coords(pt[0], pt[1])
                        if lbl_pt:
                            dist_sq = (click_pos.x() - lbl_pt.x())**2 + (click_pos.y() - lbl_pt.y())**2
                            if dist_sq <= 36:
                                self.is_dragging_point = True
                                self.dragging_point = (p_idx, pt_idx)
                                return
                                
                for p_idx, path in enumerate(connections):
                    for pt_idx in range(len(path) - 1):
                        pA = self.get_point_in_label_coords(path[pt_idx][0], path[pt_idx][1])
                        pB = self.get_point_in_label_coords(path[pt_idx+1][0], path[pt_idx+1][1])
                        if pA and pB:
                            dist_sq = self.point_to_segment_dist_squared(click_pos, pA, pB)
                            if dist_sq <= 25:
                                c = self.get_rect_in_original_coords(click_pos, click_pos)
                                if c:
                                    path.insert(pt_idx + 1, (c["x"], c["y"]))
                                    self.is_dragging_point = True
                                    self.dragging_point = (p_idx, pt_idx + 1)
                                    self.update()
                                    return
                                    
                if not getattr(self, "is_waiting_second_click", False):
                    self.is_waiting_second_click = True
                    self.line_start_point = event.pos()
                    self.current_pos = event.pos()
                else:
                    self.is_waiting_second_click = False
                    self.current_pos = event.pos()
                    p1 = self.get_rect_in_original_coords(self.line_start_point, self.line_start_point)
                    p2 = self.get_rect_in_original_coords(self.current_pos, self.current_pos)
                    if p1 and p2:
                        self.parent_dialog._push_undo_state()
                        orig_points = [(p1["x"], p1["y"]), (p2["x"], p2["y"])]
                        if "manual_connections" not in self.parent_dialog.box: self.parent_dialog.box["manual_connections"] = []
                        self.parent_dialog.box["manual_connections"].append(orig_points)
                        self.parent_dialog.actualizar_visibilidad_boton_limpieza()
                    self.line_start_point = None
                    self.current_pos = None
                self.update()
                
            elif self.mode == "pencil_freehand":
                self.is_drawing_freehand = True
                self.pencil_path = [event.pos()]
                self.update()
                
            elif self.mode == "eraser_union":
                self.is_drawing_eraser = True
                self.current_pos = event.pos()
                self.parent_dialog._push_undo_state()
                self.erase_at_position(event.pos())
                self.update()


    def mouseMoveEvent(self, event):
        event.accept()
        if getattr(self, "is_dragging_point", False) and getattr(self, "dragging_point", None):
            c = self.get_rect_in_original_coords(event.pos(), event.pos())
            if c:
                p_idx, pt_idx = self.dragging_point
                self.parent_dialog.box["manual_connections"][p_idx][pt_idx] = (c["x"], c["y"])
                self.update()
        elif self.mode == "pencil_line" and getattr(self, "is_waiting_second_click", False):
            self.current_pos = event.pos()
            self.update()
        elif getattr(self, "is_drawing_freehand", False):
            self.pencil_path.append(event.pos())
            self.update()
        elif getattr(self, "is_drawing_eraser", False) or self.is_drawing:
            self.current_pos = event.pos()
            if self.mode == "eraser_union" and getattr(self, "is_drawing_eraser", False):
                self.erase_at_position(event.pos())
            self.update()
        elif self.mode == "eraser_union":
            self.current_pos = event.pos()
            self.update()

    def leaveEvent(self, event):
        self.current_pos = None
        self.update()

    def erase_at_position(self, pos):
        """Borra píxeles en la imagen de esqueleto EN MEMORIA (no en disco)."""
        import numpy as np
        skeleton_img = getattr(self.parent_dialog, 'skeleton_working', None)
        if skeleton_img is None:
            return
        from PyQt6.QtCore import QPoint
        # Map the 16x16 label square to original crop coordinates
        p1 = QPoint(pos.x() - 8, pos.y() - 8)
        p2 = QPoint(pos.x() + 8, pos.y() + 8)
        c = self.get_rect_in_original_coords(p1, p2)
        if c:
            x, y, w, h = c["x"], c["y"], c["w"], c["h"]
            # Poner a 0 (negro) los píxeles dentro del cuadro de la goma
            self.parent_dialog.skeleton_working[y:y+h, x:x+w] = 0
            self.parent_dialog.skeleton_has_changes = True
            
            # Filtrar puntos de manual_connections que estén dentro del cuadro de la goma
            if "manual_connections" in self.parent_dialog.box:
                new_connections = []
                for path in self.parent_dialog.box["manual_connections"]:
                    new_path = []
                    for pt in path:
                        if not (x <= pt[0] <= x + w and y <= pt[1] <= y + h):
                            new_path.append(pt)
                    if len(new_path) > 1:
                        new_connections.append(new_path)
                self.parent_dialog.box["manual_connections"] = new_connections
            
            # Actualizar el pixmap en memoria y refrescar vista
            self.parent_dialog._actualizar_pixmap_esqueleto_memoria()
            self.parent_dialog.actualizar_vista()
            self.parent_dialog.actualizar_visibilidad_boton_limpieza()

    def mouseReleaseEvent(self, event):
        event.accept()
        if self.is_drawing:
            self.is_drawing = False
            if self.mode == "eraser":
                area = self.get_rect_in_original_coords(self.start_pos, self.current_pos)
                if area:
                    if "removal_areas" not in self.parent_dialog.box: self.parent_dialog.box["removal_areas"] = []
                    self.parent_dialog.box["removal_areas"].append(area)
                    self.parent_dialog.actualizar_visibilidad_boton_limpieza()
                    self.parent_dialog.actualizar_offsets()
        
        if self.mode == "pencil_line":
            if getattr(self, "is_dragging_point", False):
                self.is_dragging_point = False
                self.dragging_point = None
        elif self.mode == "pencil_freehand":
            if getattr(self, "is_drawing_freehand", False):
                self.is_drawing_freehand = False
                orig_points = []
                for p in getattr(self, "pencil_path", []):
                    c = self.get_rect_in_original_coords(p, p)
                    if c: orig_points.append((c["x"], c["y"]))
                if len(orig_points) > 1:
                    self.parent_dialog._push_undo_state()
                    if "manual_connections" not in self.parent_dialog.box: self.parent_dialog.box["manual_connections"] = []
                    self.parent_dialog.box["manual_connections"].append(orig_points)
                    self.parent_dialog.actualizar_visibilidad_boton_limpieza()

        elif self.mode == "eraser_union":
            if getattr(self, "is_drawing_eraser", False):
                self.is_drawing_eraser = False
                    
        self.start_pos = None; self.current_pos = None
        self.pencil_path = []
        self.update()

    def get_rect_in_original_coords(self, p1, p2):
        pix = self.pixmap()
        if not pix or pix.isNull(): return None
        lbl_w, lbl_h = self.width(), self.height()
        pix_w, pix_h = pix.width(), pix.height()
        dx = (lbl_w - pix_w) // 2; dy = (lbl_h - pix_h) // 2
        x1 = p1.x() - dx; y1 = p1.y() - dy
        x2 = p2.x() - dx; y2 = p2.y() - dy
        orig_w, orig_h = self.parent_dialog.get_original_crop_size()
        if orig_w == 0 or orig_h == 0: return None
        scale_x = orig_w / pix_w; scale_y = orig_h / pix_h
        rx = min(x1, x2); ry = min(y1, y2); rw = abs(x2 - x1); rh = abs(y2 - y1)
        return {"x": int(rx * scale_x), "y": int(ry * scale_y), "w": int(rw * scale_x), "h": int(rh * scale_y)}

    def paintEvent(self, event):
        super().paintEvent(event)
        pix = self.pixmap()
        if not pix or pix.isNull(): return
        painter = QPainter(self)
        lbl_w, lbl_h = self.width(), self.height()
        pix_w, pix_h = pix.width(), pix.height()
        dx = (lbl_w - pix_w) // 2; dy = (lbl_h - pix_h) // 2
        orig_w, orig_h = self.parent_dialog.get_original_crop_size()
        if orig_w > 0:
            scale_x = pix_w / orig_w; scale_y = pix_h / orig_h
            painter.setPen(QPen(QColor(220, 53, 69), 2))
            for area in self.parent_dialog.box.get("removal_areas", []):
                rx = int(area["x"] * scale_x) + dx; ry = int(area["y"] * scale_y) + dy
                rw = int(area["w"] * scale_x); rh = int(area["h"] * scale_y)
                painter.drawRect(rx, ry, rw, rh)
                painter.drawLine(rx, ry, rx + rw, ry + rh); painter.drawLine(rx + rw, ry, rx, ry + rh)
        if self.mode == "eraser" and self.start_pos and self.current_pos:
            # print("Drawing eraser rect")
            painter.setPen(QPen(QColor(255, 0, 0, 100), 1, Qt.PenStyle.DashLine))
            painter.setBrush(QColor(255, 0, 0, 50))
            rx = min(self.start_pos.x(), self.current_pos.x())
            ry = min(self.start_pos.y(), self.current_pos.y())
            rw = abs(self.current_pos.x() - self.start_pos.x())
            rh = abs(self.current_pos.y() - self.start_pos.y())
            painter.drawRect(rx, ry, rw, rh)
            
        # Dibujar cuadrito azul de la goma pixel-por-pixel
        if self.mode == "eraser_union" and getattr(self, "current_pos", None):
            painter.setPen(QPen(QColor(0, 120, 255), 1.5))
            painter.setBrush(QColor(0, 120, 255, 80)) # Azul semi-transparente
            painter.drawRect(self.current_pos.x() - 8, self.current_pos.y() - 8, 16, 16)
        
        # Dibujar pincel libre
        if getattr(self, "is_drawing_freehand", False) and getattr(self, "pencil_path", None) and len(self.pencil_path) > 1:
            painter.setPen(QPen(QColor(255, 255, 0), 2, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap, Qt.PenJoinStyle.RoundJoin))
            for i in range(len(self.pencil_path) - 1):
                painter.drawLine(self.pencil_path[i], self.pencil_path[i+1])
                
        # Dibujar la línea recta del lápiz mientras se espera el segundo clic
        if self.mode == "pencil_line" and getattr(self, "is_waiting_second_click", False) and getattr(self, "line_start_point", None) and self.current_pos:
            painter.setPen(QPen(QColor(255, 255, 0), 2, Qt.PenStyle.DashLine, Qt.PenCapStyle.RoundCap, Qt.PenJoinStyle.RoundJoin))
            painter.drawLine(self.line_start_point, self.current_pos)

        # Dibujar conexiones ya guardadas (solo si estamos en modo esqueleto)
        if self.parent_dialog.fases_disponibles[self.parent_dialog.indice_fase]["nombre"] == "ESQUELETIZADO":
            connections = self.parent_dialog.box.get("manual_connections", [])
            if connections:
                painter.setPen(QPen(Qt.GlobalColor.white, 2))
                # Necesitamos mapear de original a label
                pix = self.pixmap()
                lbl_w, lbl_h = self.width(), self.height()
                pix_w, pix_h = pix.width(), pix.height()
                dx = (lbl_w - pix_w) // 2; dy = (lbl_h - pix_h) // 2
                orig_w, orig_h = self.parent_dialog.get_original_crop_size()
                if orig_w > 0 and orig_h > 0:
                    scale_x = pix_w / orig_w; scale_y = pix_h / orig_h
                    for path in connections:
                        for i in range(len(path) - 1):
                            p1 = QPoint(int(path[i][0] * scale_x) + dx, int(path[i][1] * scale_y) + dy)
                            p2 = QPoint(int(path[i+1][0] * scale_x) + dx, int(path[i+1][1] * scale_y) + dy)
                            painter.drawLine(p1, p2)
        painter.end()

class DialogoVistaCelular(QDialog):

    """
    Ventana detallada que permite navegar entre las fases de procesamiento
    de una microglía específica (Original < Filtrado > Esqueletizado).
    """
    def __init__(self, box, pixmap_mem=None, modo_inicial="Original", parent=None):
        super().__init__(parent)
        self.setWindowTitle("Vista Detallada de la Célula")
        self.resize(500, 750) 
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        
        self.box = box
        self.crop_path = box["crop_path"].replace("\\", "/")
        self.pixmap_mem_filtrado = pixmap_mem
        self.drag_position = None

        self.fases_disponibles = []
        self.preparar_fases()
        
        # Establecer índice inicial basado en el modo de la ventana principal
        self.indice_fase = 0
        mapping = {"Original": "ORIGINAL", "Filtrada": "FILTRADO", "Esqueleto": "ESQUELETIZADO", "Previsualización": "FILTRADO"}
        nombre_buscado = mapping.get(modo_inicial, "ORIGINAL")
        for i, f in enumerate(self.fases_disponibles):
            if f["nombre"] == nombre_buscado:
                self.indice_fase = i
                break
        
        main_layout = QVBoxLayout(self)
        frame = QFrame(self)
        frame.setStyleSheet("QFrame { background-color: #FFFFFF; border-radius: 12px; border: 2px solid #003366; } QLabel { border: none; }")
        
        def start_drag(event):
            from PyQt6.QtCore import Qt
            if event.button() == Qt.MouseButton.LeftButton:
                if self.window().windowHandle():
                    self.window().windowHandle().startSystemMove()
                event.accept()

        frame.mousePressEvent = start_drag
        layout = QVBoxLayout(frame)
        
        nombre_archivo = os.path.basename(self.crop_path)
        header_layout = QHBoxLayout()
        lbl_nombre = QLabel(f"Identificador: <b>{nombre_archivo}</b>")
        lbl_nombre.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_nombre.setStyleSheet("font-size: 15px; color: #003366; border: none;")
        lbl_nombre.mousePressEvent = start_drag
        
        header_layout.setContentsMargins(10, 10, 10, 0)
        self.btn_cerrar_x = QPushButton()
        self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))
        self.btn_cerrar_x.setIconSize(QSize(20, 20))
        self.btn_cerrar_x.setFixedSize(35, 35)
        self.btn_cerrar_x.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cerrar_x.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #f6f8fa; border-radius: 17px; }
        """)
        self.btn_cerrar_x.enterEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xroja.png"))
        self.btn_cerrar_x.leaveEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))

        self.btn_cerrar_x.clicked.connect(self.accept)
        
        header_layout.addStretch()
        header_layout.addWidget(lbl_nombre)
        header_layout.addStretch()
        header_layout.addWidget(self.btn_cerrar_x)
        layout.addLayout(header_layout)
        layout.addSpacing(10)

        # Nombre de la fase actual
        self.lbl_fase = QLabel("FASE: ORIGINAL")
        self.lbl_fase.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_fase.setStyleSheet("font-size: 13px; font-weight: bold; color: #555; margin-bottom: 5px;")
        self.lbl_fase.mousePressEvent = start_drag
        layout.addWidget(self.lbl_fase)

        # Contenedor de imagen con botones de navegación lateral
        layout_imagen_nav = QHBoxLayout()
        
        self.btn_ant = QPushButton()
        self.btn_ant.setIcon(QIcon("assets/buttons/izq.png"))
        self.btn_ant.setIconSize(QSize(12, 12))
        self.btn_ant.setFixedSize(22, 22)
        self.btn_ant.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_ant.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #f0f0f0; border-radius: 11px; }
            QPushButton:disabled { opacity: 0.3; }
        """)
        self.btn_ant.clicked.connect(self.mostrar_anterior)
        
        self.label_imagen = InteractiveLabelDetail(self)
        self.label_imagen.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label_imagen.setFixedSize(380, 380)

        
        self.btn_sig = QPushButton()
        self.btn_sig.setIcon(QIcon("assets/buttons/der.png"))
        self.btn_sig.setIconSize(QSize(12, 12))
        self.btn_sig.setFixedSize(22, 22)
        self.btn_sig.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_sig.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #f0f0f0; border-radius: 11px; }
            QPushButton:disabled { opacity: 0.3; }
        """)
        self.btn_sig.clicked.connect(self.mostrar_siguiente)
        
        layout_imagen_nav.addWidget(self.btn_ant)
        layout_imagen_nav.addWidget(self.label_imagen, stretch=1)
        layout_imagen_nav.addWidget(self.btn_sig)
        
        layout.addLayout(layout_imagen_nav)
        
        # Herramientas de limpieza
        self.frame_tools_limpieza = QFrame()
        self.frame_tools_limpieza.setStyleSheet("QFrame { border: none; background: transparent; }")
        layout_tools_limpieza = QHBoxLayout(self.frame_tools_limpieza)
        layout_tools_limpieza.setContentsMargins(0,0,0,0)
        
        self.btn_tool_limpieza = QPushButton("Off")
        self.btn_tool_limpieza.setCheckable(True)
        self.btn_tool_limpieza.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_tool_limpieza.setStyleSheet("QPushButton { padding: 4px 16px; background-color: #e1e4e8; border: 2px solid #d1d5da; border-radius: 12px; font-weight: bold; color: #586069; font-size: 11px; } QPushButton:checked { background-color: #2da44e; border-color: #2da44e; color: white; }")
        self.btn_tool_limpieza.clicked.connect(self.toggle_modo_limpieza)
        
        self.btn_aplicar_limpieza = QPushButton("Eliminar áreas")
        self.btn_aplicar_limpieza.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_aplicar_limpieza.setStyleSheet("QPushButton { padding: 5px 15px; background-color: #007bff; color: white; border-radius: 4px; font-weight: bold; font-size: 11px; }")
        self.btn_aplicar_limpieza.hide()
        self.btn_aplicar_limpieza.clicked.connect(self.aplicar_limpieza)
        
        self.btn_limpiar_todo = QPushButton("Deshacer limpieza")
        self.btn_limpiar_todo.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_limpiar_todo.setStyleSheet("QPushButton { padding: 5px; background-color: #6c757d; color: white; border-radius: 4px; font-size: 11px; }")
        self.btn_limpiar_todo.hide()
        self.btn_limpiar_todo.clicked.connect(self.deshacer_limpieza)
        
        # Herramientas de unión de ramas (Esqueleto)
        self.frame_tools_esqueleto = QFrame()
        self.frame_tools_esqueleto.setStyleSheet("QFrame { border: none; background: transparent; }")
        layout_tools_esqueleto = QHBoxLayout(self.frame_tools_esqueleto)
        layout_tools_esqueleto.setContentsMargins(0,0,0,0)
        
        self.frame_subtools_union = QFrame()
        self.frame_subtools_union.setStyleSheet("QFrame { background-color: #f1f3f5; border-radius: 6px; margin-left: 10px; }")
        layout_sub = QHBoxLayout(self.frame_subtools_union)
        layout_sub.setContentsMargins(5,2,5,2)
        
        self.btn_sub_pincel = QPushButton()
        self.btn_sub_pincel.setIcon(QIcon("assets/buttons/editar.png"))
        self.btn_sub_pincel.setIconSize(QSize(20, 20))
        self.btn_sub_pincel.setToolTip("Pincel")
        
        self.btn_sub_linea = QPushButton()
        self.btn_sub_linea.setIcon(QIcon("assets/buttons/recta.png"))
        self.btn_sub_linea.setIconSize(QSize(20, 20))
        self.btn_sub_linea.setToolTip("Línea Recta")
        
        self.btn_sub_goma = QPushButton()
        self.btn_sub_goma.setIcon(QIcon("assets/buttons/goma.png"))
        self.btn_sub_goma.setIconSize(QSize(20, 20))
        self.btn_sub_goma.setToolTip("Goma")
        
        estilo_sub = "QPushButton { background-color: transparent; border: none; padding: 2px; } QPushButton:hover { background-color: #ddf4ff; border-radius: 17px; } QPushButton:checked { background-color: #cce5ff; border: 1px solid #007bff; border-radius: 17px; }"
        for btn in [self.btn_sub_pincel, self.btn_sub_linea, self.btn_sub_goma]:
            btn.setCheckable(True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet(estilo_sub)
            btn.setFixedSize(34, 34)
            btn.clicked.connect(self.cambiar_subherramienta)
            layout_sub.addWidget(btn)
        
        # Botón de deshacer paso a paso (dentro del sub-frame de herramientas)
        self.btn_deshacer_paso = QPushButton()
        self.btn_deshacer_paso.setIcon(QIcon("assets/buttons/deshacer.png"))
        self.btn_deshacer_paso.setIconSize(QSize(20, 20))
        self.btn_deshacer_paso.setToolTip("Deshacer último paso")
        self.btn_deshacer_paso.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_deshacer_paso.setStyleSheet("QPushButton { background-color: transparent; border: none; padding: 2px; } QPushButton:hover { background-color: #ddf4ff; border-radius: 17px; } QPushButton:disabled { opacity: 0.4; }")
        self.btn_deshacer_paso.setFixedSize(34, 34)
        self.btn_deshacer_paso.clicked.connect(self.deshacer_paso)
        self.btn_deshacer_paso.setEnabled(False)
        layout_sub.addWidget(self.btn_deshacer_paso)
        
        self.btn_sub_linea.setChecked(False)
        self.frame_subtools_union.hide()
        
        lbl_unir = QLabel("Unir Ramas:")
        lbl_unir.setStyleSheet("font-size: 11px; font-weight: bold; color: #555;")
        
        self.btn_tool_unir = QPushButton("Off")
        self.btn_tool_unir.setCheckable(True)
        self.btn_tool_unir.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_tool_unir.setStyleSheet("QPushButton { padding: 4px 16px; background-color: #e1e4e8; border: 2px solid #d1d5da; border-radius: 12px; font-weight: bold; color: #586069; font-size: 11px; } QPushButton:checked { background-color: #2da44e; border-color: #2da44e; color: white; }")
        self.btn_tool_unir.clicked.connect(self.toggle_modo_union)
        
        self.btn_ver_sobrepuesta = QPushButton("Ver Original")
        self.btn_ver_sobrepuesta.setCheckable(True)
        self.btn_ver_sobrepuesta.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_ver_sobrepuesta.setStyleSheet("QPushButton { padding: 4px 12px; background-color: #e1e4e8; border: 2px solid #d1d5da; border-radius: 12px; font-weight: bold; color: #586069; font-size: 11px; } QPushButton:checked { background-color: #0969da; border-color: #0969da; color: white; }")
        self.btn_ver_sobrepuesta.clicked.connect(self.actualizar_vista)
        
        self.btn_aplicar_union = QPushButton("Guardar")
        self.btn_aplicar_union.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_aplicar_union.setStyleSheet("QPushButton { padding: 5px 15px; background-color: #007bff; color: white; border-radius: 4px; font-weight: bold; font-size: 11px; } QPushButton:disabled { background-color: #cccccc; color: #666666; }")
        self.btn_aplicar_union.hide()
        self.btn_aplicar_union.setEnabled(False)
        self.btn_aplicar_union.clicked.connect(self.aplicar_union_esqueleto)
        
        self.btn_reset = QPushButton()
        self.btn_reset.setIcon(QIcon("assets/buttons/reset.png"))
        self.btn_reset.setIconSize(QSize(20, 20))
        self.btn_reset.setToolTip("Restablecer")
        self.btn_reset.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_reset.setStyleSheet("QPushButton { background-color: transparent; border: none; padding: 2px; } QPushButton:hover { background-color: #ffdddd; border-radius: 17px; }")
        self.btn_reset.setFixedSize(34, 34)
        self.btn_reset.hide()
        self.btn_reset.clicked.connect(self.restablecer_esqueleto)
        
        layout_tools_esqueleto.addWidget(lbl_unir)
        layout_tools_esqueleto.addWidget(self.btn_tool_unir)
        layout_tools_esqueleto.addWidget(self.btn_ver_sobrepuesta)
        layout_tools_esqueleto.addWidget(self.frame_subtools_union)
        layout_tools_esqueleto.addWidget(self.btn_aplicar_union)
        layout_tools_esqueleto.addWidget(self.btn_reset)
        layout_tools_esqueleto.addStretch()
        
        layout_tools_limpieza.addWidget(QLabel("Limpieza:"))
        layout_tools_limpieza.addWidget(self.btn_tool_limpieza)
        layout_tools_limpieza.addStretch()
        layout_tools_limpieza.addWidget(self.btn_limpiar_todo)
        layout_tools_limpieza.addWidget(self.btn_aplicar_limpieza)
        layout.addWidget(self.frame_tools_limpieza)
        layout.addWidget(self.frame_tools_esqueleto)


        
        # Filtros Individuales (Offsets) - Solo si estamos en modo investigador y no es modo lectura
        self.frame_offsets = QFrame()
        self.frame_offsets.setStyleSheet("QFrame { background-color: #f9f9f9; border: 1px solid #ddd; border-radius: 8px; margin: 5px; } QLabel { color: #333; font-weight: bold; font-size: 11px; }")
        layout_offsets = QVBoxLayout(self.frame_offsets)
        
        lbl_off_tit = QLabel("Ajuste de Filtros Individuales (Offsets)")
        lbl_off_tit.setStyleSheet("color: #003366; font-size: 12px; margin-bottom: 5px;")
        layout_offsets.addWidget(lbl_off_tit)
        
        grid_offsets = QGridLayout()
        
        # CLAHE Offset
        grid_offsets.addWidget(QLabel("Contraste:"), 0, 0)
        self.sld_o_clahe = QSlider(Qt.Orientation.Horizontal); self.sld_o_clahe.setRange(-5, 5)
        self.sld_o_clahe.setValue(self.box["offsets"]["clahe"])
        grid_offsets.addWidget(self.sld_o_clahe, 0, 1)
        
        # Gauss Offset
        grid_offsets.addWidget(QLabel("Suavizado:"), 1, 0)
        self.sld_o_gauss = QSlider(Qt.Orientation.Horizontal); self.sld_o_gauss.setRange(-6, 6)
        self.sld_o_gauss.setValue(self.box["offsets"]["gauss"])
        grid_offsets.addWidget(self.sld_o_gauss, 1, 1)
        
        # Otsu Offset
        grid_offsets.addWidget(QLabel("Umbral:"), 2, 0)
        self.sld_o_otsu = QSlider(Qt.Orientation.Horizontal); self.sld_o_otsu.setRange(-50, 50)
        self.sld_o_otsu.setValue(self.box["offsets"]["otsu"])
        grid_offsets.addWidget(self.sld_o_otsu, 2, 1)
        
        layout_offsets.addLayout(grid_offsets)
        
        # Estilo para sliders de offset
        estilo_off = "QSlider::groove:horizontal { height: 4px; background: #ddd; } QSlider::handle:horizontal { background: #3a61a0; width: 12px; height: 12px; margin: -4px 0; border-radius: 6px; }"
        for s in [self.sld_o_clahe, self.sld_o_gauss, self.sld_o_otsu]:
            s.setStyleSheet(estilo_off)
            s.valueChanged.connect(self.actualizar_offsets)
            
        layout.addWidget(self.frame_offsets)
        
        # La visibilidad se controla en actualizar_vista()


        layout.addSpacing(10)


        
        # Botones inferiores
        layout_inferior = QHBoxLayout()
        
        self.btn_comparativa = QPushButton("Ver Proceso Completo")
        self.btn_comparativa.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_comparativa.setStyleSheet("""
            QPushButton { padding: 8px 15px; background-color: #28a745; border-radius: 6px; font-weight: bold; color: white; font-size: 13px;}
            QPushButton:hover { background-color: #218838; }
            QPushButton:disabled { background-color: #ccc; }
        """)
        self.btn_comparativa.clicked.connect(self.mostrar_comparativa)
        # Solo habilitar si el proceso está terminado (las 3 fases existen)
        self.btn_comparativa.setEnabled(len(self.fases_disponibles) == 3)
        
        self.actualizar_visibilidad_boton_limpieza()

        
        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_cerrar.setStyleSheet("""
            QPushButton { padding: 8px 20px; background-color: #dc3545; border-radius: 6px; font-weight: bold; color: white; font-size: 14px;}
            QPushButton:hover { background-color: #c82333; }
        """)
        btn_cerrar.clicked.connect(self.accept)
        
        layout_inferior.addStretch()
        layout_inferior.addWidget(self.btn_comparativa)
        layout_inferior.addStretch()
        
        layout.addLayout(layout_inferior)
        main_layout.addWidget(frame)
        self.setLayout(main_layout)
        
        self.actualizar_vista()

    def preparar_fases(self):
        self.undo_stack = []  # Pila de estados para deshacer paso a paso
        # Fase 0: Original
        self.fases_disponibles.append({"nombre": "ORIGINAL", "path": self.crop_path, "pixmap": None})
        
        # Fase 1: Filtrado
        path_filtrado = self.crop_path.replace("\\", "/").replace("/crops/", "/filtradas/")

        if self.pixmap_mem_filtrado:
            self.fases_disponibles.append({"nombre": "FILTRADO", "path": path_filtrado, "pixmap": self.pixmap_mem_filtrado})
        elif os.path.exists(path_filtrado):
            self.fases_disponibles.append({"nombre": "FILTRADO", "path": path_filtrado, "pixmap": None})
            
        # Fase 2: Esqueletizado
        path_esqueleto = self.crop_path.replace("\\", "/").replace("/crops/", "/esqueletos/")

        if os.path.exists(path_esqueleto):
            import cv2; import numpy as np
            # Cargar en memoria: backup (intocable) + working (editable)
            self.skeleton_backup = cv2.imread(path_esqueleto, cv2.IMREAD_GRAYSCALE)
            self.skeleton_working = self.skeleton_backup.copy() if self.skeleton_backup is not None else None
            self.skeleton_has_changes = False
            self.fases_disponibles.append({"nombre": "ESQUELETIZADO", "path": path_esqueleto, "pixmap": None})

    def actualizar_vista(self):
        from PyQt6.QtCore import Qt
        from PyQt6.QtGui import QPainter
        
        fase = self.fases_disponibles[self.indice_fase]
        self.lbl_fase.setText(f"FASE: {fase['nombre']}")
        
        pixmap = fase["pixmap"]
        if not pixmap:
            pixmap = QPixmap(fase["path"])
            
        # Si estamos en ESQUELETIZADO, usar la imagen en memoria (skeleton_working)
        if fase["nombre"] == "ESQUELETIZADO" and getattr(self, 'skeleton_working', None) is not None:
            from PyQt6.QtGui import QImage
            h_img, w_img = self.skeleton_working.shape
            qimg = QImage(self.skeleton_working.data, w_img, h_img, w_img, QImage.Format.Format_Grayscale8)
            pixmap = QPixmap.fromImage(qimg)
        
        if pixmap and not pixmap.isNull():
            # Si estamos en ESQUELETIZADO y el botón de ver sobrepuesta está activo, combinar con la original
            if fase["nombre"] == "ESQUELETIZADO" and getattr(self, "btn_ver_sobrepuesta", None) and self.btn_ver_sobrepuesta.isChecked():
                pix_orig = QPixmap(self.crop_path)
                if pix_orig and not pix_orig.isNull():
                    combined = QPixmap(pix_orig.size())
                    combined.fill(Qt.GlobalColor.black)
                    
                    painter = QPainter(combined)
                    painter.drawPixmap(0, 0, pix_orig)
                    painter.setCompositionMode(QPainter.CompositionMode.CompositionMode_Screen)
                    painter.setOpacity(0.8)
                    painter.drawPixmap(0, 0, pixmap)
                    painter.end()
                    
                    self.label_imagen.setPixmap(combined.scaled(380, 380, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
                else:
                    self.label_imagen.setPixmap(pixmap.scaled(380, 380, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
            else:
                self.label_imagen.setPixmap(pixmap.scaled(380, 380, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
        else:
            self.label_imagen.setText(f"No se pudo cargar la imagen de {fase['nombre']}.")

        self.btn_ant.setVisible(self.indice_fase > 0)
        self.btn_sig.setVisible(self.indice_fase < len(self.fases_disponibles) - 1)
        self.btn_comparativa.setVisible(len(self.fases_disponibles) == 3)

        # Actualizar visibilidad de herramientas según la fase
        if fase["nombre"] == "FILTRADO":
            filtros_activos = True
            if self.parent() and hasattr(self.parent(), "combo_vista"):
                if self.parent().combo_vista.currentText() != "Previsualización":
                    filtros_activos = False
            elif len(self.fases_disponibles) >= 3:
                filtros_activos = False

            if not filtros_activos:
                self.frame_offsets.hide()
                self.frame_tools_limpieza.hide()
                self.resize(450, 520)
            else:
                self.frame_offsets.show()
                self.frame_tools_limpieza.show()
                self.resize(500, 780)
            self.frame_tools_esqueleto.hide()
        elif fase["nombre"] == "ESQUELETIZADO":
            self.frame_offsets.hide()
            self.frame_tools_limpieza.hide()
            self.frame_tools_esqueleto.show()
            self.resize(500, 600)
        else:
            self.frame_offsets.hide()
            self.frame_tools_limpieza.hide()
            self.frame_tools_esqueleto.hide()
            self.resize(450, 520)


    def actualizar_offsets(self):
        # 1. Actualizar valores en el box
        self.box["offsets"]["clahe"] = self.sld_o_clahe.value()
        self.box["offsets"]["gauss"] = self.sld_o_gauss.value()
        self.box["offsets"]["otsu"] = self.sld_o_otsu.value()
        
        # 2. Notificar al padre para que reprocese globalmente
        if hasattr(self.parent(), "previsualizar_filtrado"):
            self.parent().previsualizar_filtrado()
            
            # 3. Obtener el nuevo pixmap filtrado para este crop
            nombre = os.path.basename(self.crop_path)
            if nombre in self.parent().crops_filtrados_temp:
                bin_img = self.parent().crops_filtrados_temp[nombre]
                import cv2; from PyQt6.QtGui import QImage
                h, w = bin_img.shape
                qimg = QImage(bin_img.data, w, h, w, QImage.Format.Format_Grayscale8)
                nuevo_pixmap = QPixmap.fromImage(qimg)
                
                # Actualizar en fases_disponibles
                for f in self.fases_disponibles:
                    if f["nombre"] == "FILTRADO":
                        f["pixmap"] = nuevo_pixmap
                        break
                
                # Si estamos viendo la fase filtrado, refrescar
                if self.fases_disponibles[self.indice_fase]["nombre"] == "FILTRADO":
                    self.actualizar_vista()

    def get_original_crop_size(self):
        pix = QPixmap(self.crop_path)
        if not pix.isNull(): return pix.width(), pix.height()
        return 0, 0

    def toggle_modo_limpieza(self, checked):
        if checked:
            self.label_imagen.mode = "eraser"
            self.label_imagen.setCursor(Qt.CursorShape.CrossCursor)
            self.btn_tool_limpieza.setText("On")
            self.btn_tool_unir.blockSignals(True)
            self.btn_tool_unir.setChecked(False)
            self.btn_tool_unir.setText("Off")
            self.btn_tool_unir.blockSignals(False)
        else:
            self.label_imagen.mode = "pointer"
            self.label_imagen.setCursor(Qt.CursorShape.ArrowCursor)
            self.btn_tool_limpieza.setText("Off")

    def cambiar_subherramienta(self):
        sender = self.sender()
        if not sender:
            # Desmarcar todo si no se llamó desde un botón específico
            self.btn_sub_pincel.setChecked(False)
            self.btn_sub_linea.setChecked(False)
            self.btn_sub_goma.setChecked(False)
            self.label_imagen.mode = "pointer"
            self.label_imagen.setCursor(Qt.CursorShape.ArrowCursor)
            return

        self.btn_sub_pincel.setChecked(sender == self.btn_sub_pincel)
        self.btn_sub_linea.setChecked(sender == self.btn_sub_linea)
        self.btn_sub_goma.setChecked(sender == self.btn_sub_goma)
        
        if sender == self.btn_sub_goma:
            self.label_imagen.mode = "eraser_union"
            self.label_imagen.setCursor(Qt.CursorShape.BlankCursor)
        elif sender == self.btn_sub_pincel:
            self.label_imagen.mode = "pencil_freehand"
            self.label_imagen.setCursor(self.label_imagen.pencil_cursor)
        elif sender == self.btn_sub_linea:
            self.label_imagen.mode = "pencil_line"
            self.label_imagen.setCursor(self.label_imagen.pencil_cursor)

    def toggle_modo_union(self, checked):
        if checked:
            self.frame_subtools_union.show()
            # En lugar de seleccionar un modo por defecto, desmarcamos todos para obligar al usuario a elegir
            self.btn_sub_pincel.setChecked(False)
            self.btn_sub_linea.setChecked(False)
            self.btn_sub_goma.setChecked(False)
            self.label_imagen.mode = "pointer"
            self.label_imagen.setCursor(Qt.CursorShape.ArrowCursor)
            
            self.btn_tool_unir.setText("On")
            self.actualizar_visibilidad_boton_limpieza()
            self.btn_tool_limpieza.blockSignals(True)
            self.btn_tool_limpieza.setChecked(False)
            self.btn_tool_limpieza.setText("Off")
            self.btn_tool_limpieza.blockSignals(False)
        else:
            self.frame_subtools_union.hide()
            self.label_imagen.mode = "pointer"
            self.label_imagen.setCursor(Qt.CursorShape.ArrowCursor)
            self.btn_tool_unir.setText("Off")
            self.actualizar_visibilidad_boton_limpieza()

    def _actualizar_pixmap_esqueleto_memoria(self):
        """Genera un QPixmap a partir de skeleton_working y lo asigna a la fase ESQUELETIZADO."""
        if self.skeleton_working is None:
            return
        from PyQt6.QtGui import QImage
        h_img, w_img = self.skeleton_working.shape
        qimg = QImage(self.skeleton_working.data, w_img, h_img, w_img, QImage.Format.Format_Grayscale8)
        pix = QPixmap.fromImage(qimg)
        for f in self.fases_disponibles:
            if f["nombre"] == "ESQUELETIZADO":
                f["pixmap"] = pix
                break

    def aplicar_union_esqueleto(self):
        """Guarda todos los cambios (goma + uniones) al esqueleto en disco."""
        import cv2; import numpy as np
        from skimage.morphology import skeletonize

        if self.skeleton_working is None:
            return

        # Generar la imagen final con las uniones de ramas aplicadas sobre el working actual
        img_final = self.skeleton_working.copy()
        for conn_path in self.box.get("manual_connections", []):
            for i in range(len(conn_path) - 1):
                p1 = (conn_path[i][0], conn_path[i][1])
                p2 = (conn_path[i+1][0], conn_path[i+1][1])
                cv2.line(img_final, p1, p2, 255, 2)

        # Skeletonize para limpiar uniones a 1px
        if len(self.box.get("manual_connections", [])) > 0:
            img_bool = img_final > 0
            skeleton = skeletonize(img_bool)
            img_final = (skeleton * 255).astype(np.uint8)

        # Mostrar previsualización del resultado final
        self.skeleton_working = img_final
        self._actualizar_pixmap_esqueleto_memoria()
        self.actualizar_vista()
        self.label_imagen.update()

        # Preguntar si desea guardar permanentemente
        from vistas.utilidades import DialogoConfirmacion
        diag = DialogoConfirmacion("Guardar Cambios", "¿Confirmas que deseas guardar estos cambios?")
        if diag.exec():
            # Guardar permanentemente en disco
            path_esqueleto = self.crop_path.replace("\\", "/").replace("/crops/", "/esqueletos/")
            cv2.imwrite(path_esqueleto, img_final)
            self.box["esqueleto_modificado"] = True
            # Limpiar conexiones ya aplicadas
            self.box["manual_connections"] = []
            # Actualizar backup a la versión guardada
            self.skeleton_backup = img_final.copy()
            self.skeleton_has_changes = False
            self.undo_stack = []
            self.actualizar_visibilidad_boton_limpieza()
            self.actualizar_vista()
            self.label_imagen.update()

            # Desactivar modo unión tras guardar cambios
            self.btn_tool_unir.setChecked(False)
            self.toggle_modo_union(False)

            # Refrescar imagen global
            self._refrescar_imagen_global_esqueleto()

            self.mostrar_notificacion("Éxito", "Los cambios se han guardado correctamente.", "info")
        else:
            # Revertir: recalcular working desde el estado actual sin las uniones
            # (las uniones no se habían aplicado al working antes, solo temporalmente)
            # Restaurar skeleton_working al estado previo sin uniones aplicadas
            # Re-leer desde la versión que teníamos antes de aplicar uniones
            self.skeleton_working = self.skeleton_backup.copy()
            self.skeleton_has_changes = False
            self.box["manual_connections"] = []
            self.undo_stack = []
            self._actualizar_pixmap_esqueleto_memoria()
            self.actualizar_vista()
            self.label_imagen.update()
            self.actualizar_visibilidad_boton_limpieza()

    def _refrescar_imagen_global_esqueleto(self):
        """Método auxiliar para actualizar la imagen global de esqueleto en la ventana principal."""
        parent_win = self.parent()
        if parent_win and hasattr(parent_win, "pixmaps_globales") and hasattr(parent_win, "construir_imagen_global"):
            pixmap_esqueleto = parent_win.construir_imagen_global("esqueletos")
            parent_win.pixmaps_globales["Esqueleto"] = pixmap_esqueleto
            if hasattr(parent_win, "combo_vista") and parent_win.combo_vista.currentText() == "Esqueleto":
                parent_win.visor_imagen.set_view_mode("Esqueleto", pixmap_esqueleto)

    def actualizar_visibilidad_boton_limpieza(self):
        has_areas = len(self.box.get("removal_areas", [])) > 0
        self.btn_aplicar_limpieza.setVisible(has_areas)
        self.btn_limpiar_todo.setVisible(has_areas)
        
        has_uniones = len(self.box.get("manual_connections", [])) > 0
        has_any_changes = has_uniones or getattr(self, 'skeleton_has_changes', False)
        has_undo_steps = len(getattr(self, 'undo_stack', [])) > 0
        is_union_on = self.btn_tool_unir.isChecked()
        # Guardar Cambios: visible si hay cambios pendientes o el modo unión está activo, pero solo habilitado si hay cambios reales
        self.btn_aplicar_union.setVisible(has_any_changes or is_union_on)
        self.btn_aplicar_union.setEnabled(has_any_changes)
        # Restablecer / Reset: visible solo cuando unir ramas está OFF y la imagen ya fue modificada permanentemente (amarillo)
        is_modified_saved = self.box.get("esqueleto_modificado", False)
        show_reset = (not is_union_on) and is_modified_saved
        self.btn_reset.setVisible(show_reset)
        # Deshacer paso: habilitado si hay pasos en la pila
        self.btn_deshacer_paso.setEnabled(has_undo_steps)

    def _push_undo_state(self):
        """Guarda el estado actual en la pila de deshacer antes de una acción."""
        import copy
        state = {
            'skeleton': self.skeleton_working.copy() if self.skeleton_working is not None else None,
            'connections': copy.deepcopy(self.box.get("manual_connections", []))
        }
        self.undo_stack.append(state)
        self.skeleton_has_changes = True

    def deshacer_paso(self):
        """Deshace el último paso individual (una acción de goma o una conexión)."""
        if not getattr(self, 'undo_stack', None) or len(self.undo_stack) == 0:
            return
        state = self.undo_stack.pop()
        if state['skeleton'] is not None:
            self.skeleton_working = state['skeleton']
        self.box["manual_connections"] = state['connections']
        self.skeleton_has_changes = len(self.undo_stack) > 0
        self._actualizar_pixmap_esqueleto_memoria()
        self.actualizar_visibilidad_boton_limpieza()
        self.actualizar_vista()
        self.label_imagen.update()

    def restablecer_esqueleto(self):
        """Revierte los cambios al estado original.
        - Si unir ramas está ON: Revierte los cambios temporales en memoria al último backup guardado.
        - Si unir ramas está OFF: Recalcula y restaura el esqueleto original sin modificaciones a partir de la imagen filtrada.
        """
        import numpy as np
        is_union_on = self.btn_tool_unir.isChecked()
        
        if not is_union_on:
            from vistas.utilidades import DialogoConfirmacion
            diag = DialogoConfirmacion(
                "Restaurar Esqueleto",
                "¿Confirmas que deseas restaurar este esqueleto al original?\nSe perderán permanentemente todos los cambios manuales guardados."
            )
            if diag.exec():
                path_filtrado = self.crop_path.replace("\\", "/").replace("/crops/", "/filtradas/")
                import cv2
                from skimage.morphology import skeletonize
                img_raw = cv2.imread(path_filtrado, cv2.IMREAD_GRAYSCALE)
                if img_raw is not None:
                    _, bin_img = cv2.threshold(img_raw, 127, 255, cv2.THRESH_BINARY)
                    img_bool = bin_img > 0
                    skeleton = skeletonize(img_bool)
                    img_final = (skeleton * 255).astype(np.uint8)
                    
                    path_esqueleto = self.crop_path.replace("\\", "/").replace("/crops/", "/esqueletos/")
                    cv2.imwrite(path_esqueleto, img_final)
                    
                    self.skeleton_working = img_final.copy()
                    self.skeleton_backup = img_final.copy()
                    self.box["esqueleto_modificado"] = False
                    self.box["manual_connections"] = []
                    self.undo_stack = []
                    self.skeleton_has_changes = False
                    
                    self._actualizar_pixmap_esqueleto_memoria()
                    self.actualizar_visibilidad_boton_limpieza()
                    self.actualizar_vista()
                    self.label_imagen.update()
                    self._refrescar_imagen_global_esqueleto()
                    self.mostrar_notificacion("Restablecido", "El esqueleto ha sido restaurado al filtrado original.", "info")
            return

        # Si unir ramas está ON: Restaurar esqueleto al último estado guardado en disco
        if getattr(self, 'skeleton_backup', None) is not None:
            self.skeleton_working = self.skeleton_backup.copy()
            self.skeleton_has_changes = False
        # Limpiar todas las conexiones manuales y pila de undo
        self.box["manual_connections"] = []
        self.undo_stack = []
        self._actualizar_pixmap_esqueleto_memoria()
        self.actualizar_visibilidad_boton_limpieza()
        self.actualizar_vista()
        self.label_imagen.update()

    def deshacer_limpieza(self):
        if "removal_areas" in self.box and len(self.box["removal_areas"]) > 0:
            self.box["removal_areas"].pop()
        self.actualizar_visibilidad_boton_limpieza()
        self.actualizar_offsets()
        self.label_imagen.update()

    def aplicar_limpieza(self):
        from vistas.utilidades import DialogoConfirmacion
        diag = DialogoConfirmacion("Eliminar Áreas", "¿Confirmas que deseas eliminar permanentemente estas áreas? Esto modificará la imagen base.")
        if diag.exec():
            # Aplicar permanentemente a la imagen en memoria (BGR)
            nombre = os.path.basename(self.crop_path)
            if hasattr(self.parent(), "crops_en_memoria") and nombre in self.parent().crops_en_memoria:
                img_bgr = self.parent().crops_en_memoria[nombre]
                import cv2
                for area in self.box.get("removal_areas", []):
                    ax, ay, aw, ah = area["x"], area["y"], area["w"], area["h"]
                    cv2.rectangle(img_bgr, (ax, ay), (ax + aw, ay + ah), (0, 0, 0), -1)
                
                # Limpiar áreas ya aplicadas
                self.box["removal_areas"] = []
                self.actualizar_visibilidad_boton_limpieza()
                self.actualizar_offsets() # Esto reprocesará con la nueva imagen base
                self.label_imagen.update()
                self.mostrar_notificacion("Limpieza Aplicada", "Las áreas se han eliminado de la imagen base.", "info")


    def deshacer_limpieza(self):
        self.box["removal_areas"] = []
        self.actualizar_visibilidad_boton_limpieza()
        self.actualizar_offsets()
        self.label_imagen.update()

    def mostrar_notificacion(self, t, m, tipo):
        from vistas.utilidades import DialogoNotificacion
        DialogoNotificacion(t, m, tipo, self).exec()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            if self.window().windowHandle():
                self.window().windowHandle().startSystemMove()
            event.accept()



    def mostrar_anterior(self):
        if self.indice_fase > 0:
            self.indice_fase -= 1
            self.actualizar_vista()

    def mostrar_siguiente(self):
        if self.indice_fase < len(self.fases_disponibles) - 1:
            self.indice_fase += 1
            self.actualizar_vista()

    def mostrar_comparativa(self):
        diag = DialogoComparativo(self.fases_disponibles, self.parent())
        diag.exec()

    def showEvent(self, event):
        super().showEvent(event)
        if self.parent():
            p_geom = self.parent().geometry()
            self.move(p_geom.x() + (p_geom.width() - self.width()) // 2, p_geom.y() + (p_geom.height() - self.height()) // 2)
        else:
            screen = QApplication.primaryScreen().geometry()
            self.move((screen.width() - self.width()) // 2, (screen.height() - self.height()) // 2)


class DialogoConfirmacion(QDialog):
    def __init__(self, titulo, mensaje, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True)
        self.resultado = False
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        if parent:
            parent.installEventFilter(self)

        layout = QVBoxLayout(self)
        frame = QFrame(self)
        frame.setStyleSheet("""
            QFrame { background-color: #FFFFFF; border-radius: 12px; border: 2px solid #cc0000; }
            QLabel { color: #333333; font-size: 15px; padding: 10px; border: none;}
            QPushButton { border-radius: 6px; font-weight: bold; padding: 8px 15px; font-size: 14px; }
            QPushButton#btn_eliminar { background-color: #cc0000; color: white; }
            QPushButton#btn_eliminar:hover { background-color: #aa0000; }
            QPushButton#btn_mantener { background-color: #e0e0e0; color: #333333; }
        """)

        flayout = QVBoxLayout(frame)
        lbl_titulo = QLabel(f"<b>{titulo}</b>")
        lbl_titulo.setStyleSheet("color: #cc0000; font-size: 18px; border: none;")
        lbl_titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)

        lbl_mensaje = QLabel(mensaje)
        lbl_mensaje.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_mensaje.setWordWrap(True)

        btn_layout = QHBoxLayout()
        btn_mantener = QPushButton("Mantener")
        btn_mantener.setObjectName("btn_mantener")
        btn_mantener.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_mantener.clicked.connect(self.cancelar)

        btn_eliminar = QPushButton("Eliminar")
        btn_eliminar.setObjectName("btn_eliminar")
        btn_eliminar.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_eliminar.clicked.connect(self.aceptar)

        btn_layout.addWidget(btn_mantener)
        btn_layout.addWidget(btn_eliminar)

        flayout.addWidget(lbl_titulo)
        flayout.addWidget(lbl_mensaje)
        flayout.addLayout(btn_layout)
        layout.addWidget(frame)
        self.setLayout(layout)

    def eventFilter(self, obj, event):
        from PyQt6.QtCore import QEvent
        if obj == self.parent() and (event.type() == QEvent.Type.Resize or event.type() == QEvent.Type.Move):
            self.centrar_en_padre()
        return super().eventFilter(obj, event)

    def centrar_en_padre(self):
        if self.parent():
            p_geom = self.parent().geometry()
            self.move(p_geom.x() + (p_geom.width() - self.width()) // 2, p_geom.y() + (p_geom.height() - self.height()) // 2)

    def showEvent(self, event):
        from PyQt6.QtWidgets import QApplication
        super().showEvent(event)
        if self.parent():
            self.centrar_en_padre()
        else:
            screen = QApplication.primaryScreen().geometry()
            self.move((screen.width() - self.width()) // 2, (screen.height() - self.height()) // 2)

    def aceptar(self): self.resultado = True; self.accept()
    def cancelar(self): self.resultado = False; self.reject()

class InteractiveImageViewer(QLabel):
    conteo_actualizado = pyqtSignal(int)
    nueva_caja_dibujada = pyqtSignal(int, int, int, int)
    nivel_zoom_cambiado = pyqtSignal(int)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMouseTracking(True)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.original_pixmap = None
        self._current_pixmap = None
        self.boxes = []
        self.hovered_index = -1
        self.active_index = -1
        self.view_mode = "Original"
        self.current_tool = "pointer" 
        self.is_drawing = False
        self.draw_start_pos = None
        self.draw_current_pos = None
        self.is_panning = False
        self.pan_start_pos = None
        self.pan_x = 0
        self.pan_y = 0
        self.zoom_level = 100
        self.zoom_locked = False

    def set_image_and_boxes(self, pixmap, bounding_boxes):
        self.original_pixmap = pixmap
        self.boxes = bounding_boxes
        self.hovered_index = -1
        self.pan_x = 0
        self.pan_y = 0
        self.draw_current_state()
        self.conteo_actualizado.emit(len(self.boxes))

    def set_view_mode(self, mode, new_pixmap):
        self.view_mode = mode
        self.original_pixmap = new_pixmap
        self.draw_current_state()

    def map_mouse_to_original(self, mouse_pos):
        if not self.original_pixmap or self.original_pixmap.isNull(): return None
        pix_w, pix_h = self.original_pixmap.width(), self.original_pixmap.height()
        lbl_w, lbl_h = self.width(), self.height()
        if lbl_w == 0 or lbl_h == 0: return None
        
        base_scale = min(lbl_w / pix_w, lbl_h / pix_h)
        actual_scale = base_scale * (self.zoom_level / 100.0)
        scaled_w = int(pix_w * actual_scale)
        scaled_h = int(pix_h * actual_scale)
        
        draw_x = (lbl_w - scaled_w) // 2 + self.pan_x
        draw_y = (lbl_h - scaled_h) // 2 + self.pan_y
        mx, my = mouse_pos.x(), mouse_pos.y()
        
        if draw_x <= mx <= draw_x + scaled_w and draw_y <= my <= draw_y + scaled_h:
            orig_x = (mx - draw_x) / actual_scale
            orig_y = (my - draw_y) / actual_scale
            return orig_x, orig_y
        return None

    def wheelEvent(self, event):
        if not self.original_pixmap or self.zoom_locked: return
        delta = event.angleDelta().y()
        if delta > 0: new_zoom = min(400, self.zoom_level + 15)
        else: new_zoom = max(50, self.zoom_level - 15)
        if new_zoom != self.zoom_level: self.set_zoom(new_zoom)

    def mousePressEvent(self, event):
        if not self.original_pixmap: return
        

        if event.button() == Qt.MouseButton.LeftButton:
            if self.current_tool == "pointer":
                if self.hovered_index != -1:
                    crop_path_base = self.boxes[self.hovered_index]["crop_path"]
                    crop_path = crop_path_base.replace("\\", "/")
                    pixmap_mem = None
                    if hasattr(self.window(), 'crops_filtrados_temp'):
                        nombre_base = os.path.basename(crop_path)
                        if self.view_mode == "Filtrada" or self.view_mode == "Previsualización":
                            arr = self.window().crops_filtrados_temp.get(nombre_base)
                            if arr is not None:
                                h, w = arr.shape
                                qimg = QImage(arr.data, w, h, w, QImage.Format.Format_Grayscale8)
                                pixmap_mem = QPixmap.fromImage(qimg)
                    
                    if self.view_mode == "Filtrada": crop_path = crop_path.replace("/crops/", "/filtradas/")
                    elif self.view_mode == "Esqueleto": crop_path = crop_path.replace("/crops/", "/esqueletos/")

                    if os.path.exists(crop_path) or pixmap_mem:
                        self.active_index = self.hovered_index
                        self.draw_current_state()
                        DialogoVistaCelular(self.boxes[self.hovered_index], pixmap_mem, self.view_mode, self.window()).exec()
                        self.active_index = -1
                        # Refrescar imagen global de esqueleto tras cerrar el diálogo por si hubo ediciones
                        if self.view_mode == "Esqueleto" and hasattr(self.window(), 'construir_imagen_global'):
                            pixmap_esqueleto = self.window().construir_imagen_global("esqueletos")
                            self.window().pixmaps_globales["Esqueleto"] = pixmap_esqueleto
                            self.window().visor_imagen.set_view_mode("Esqueleto", pixmap_esqueleto)
                        self.draw_current_state()

                else:
                    self.is_panning = True
                    self.pan_start_pos = event.pos()
                    self.setCursor(Qt.CursorShape.ClosedHandCursor)
            elif self.current_tool == "draw":
                self.is_drawing = True
                self.draw_start_pos = event.pos()
                self.draw_current_pos = event.pos()
            elif self.current_tool == "delete":
                if self.hovered_index != -1:
                    index_to_delete = self.hovered_index
                    diag = DialogoConfirmacion("Eliminar Detección", "¿Estás seguro de descartar esta célula?")
                    diag.exec()
                    if diag.resultado:
                        if index_to_delete < len(self.boxes):
                            self.boxes.pop(index_to_delete)
                            self.hovered_index = -1
                            self.draw_current_state()
                            self.conteo_actualizado.emit(len(self.boxes))

    def mouseMoveEvent(self, event):
        if not self.original_pixmap: return
        if self.is_panning:
            delta = event.pos() - self.pan_start_pos
            self.pan_x += delta.x()
            self.pan_y += delta.y()
            self.pan_start_pos = event.pos()
            self.update() 
            return
        if self.is_drawing:
            self.draw_current_pos = event.pos()
            self.draw_current_state()
            return
        if self.current_tool == "draw":
            if self.hovered_index != -1:
                self.hovered_index = -1
                self.draw_current_state()
            return
        if not self.boxes: return
        orig_coords = self.map_mouse_to_original(event.pos())
        new_hovered_index = -1
        if orig_coords:
            ox, oy = orig_coords
            for i in range(len(self.boxes)-1, -1, -1):
                box = self.boxes[i]
                if box["x"] <= ox <= box["x"] + box["w"] and box["y"] <= oy <= box["y"] + box["h"]:
                    new_hovered_index = i
                    break
        if new_hovered_index != self.hovered_index:
            self.hovered_index = new_hovered_index
            self.draw_current_state()

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            if self.is_panning:
                self.is_panning = False
                self.setCursor(Qt.CursorShape.ArrowCursor)
            elif self.is_drawing:
                self.is_drawing = False
                orig_start = self.map_mouse_to_original(self.draw_start_pos)
                orig_end = self.map_mouse_to_original(self.draw_current_pos)
                if orig_start and orig_end:
                    x1, y1 = orig_start; x2, y2 = orig_end
                    x = min(x1, x2); y = min(y1, y2)
                    w = abs(x2 - x1); h = abs(y2 - y1)
                    if w >= MIN_MICROGLIA_SIZE and h >= MIN_MICROGLIA_SIZE: self.nueva_caja_dibujada.emit(int(x), int(y), int(w), int(h))
                self.draw_start_pos = None; self.draw_current_pos = None
                self.draw_current_state()

    def leaveEvent(self, event):
        if self.hovered_index != -1:
            self.hovered_index = -1
            self.draw_current_state()
        super().leaveEvent(event)

    def borrar_rama_esqueleto(self, pos):
        if not self.original_pixmap: return
        
        orig_coords = self.map_mouse_to_original(pos)
        if not orig_coords: return
        
        ox, oy = orig_coords
        
        # Modificar el pixmap original (que es el esqueleto)
        # Pintamos un círculo negro para 'borrar' la rama
        painter = QPainter(self.original_pixmap)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor(0, 0, 0)) # Negro
        painter.drawEllipse(QPoint(int(ox), int(oy)), 8, 8) # Radio de 8 pixeles para facilitar el borrado
        painter.end()
        
        self.draw_current_state()

    def set_zoom(self, value):
        if not self.original_pixmap or self.original_pixmap.isNull(): return
        if not self.zoom_locked: self.zoom_level = value; self.draw_current_state(); self.nivel_zoom_cambiado.emit(value)

    def lock_zoom(self, locked): self.zoom_locked = locked

    def draw_current_state(self):
        """Dibuja la imagen, las cajas normales y la caja resaltada."""
        if not self.original_pixmap or self.original_pixmap.isNull():
            self._current_pixmap = None
            self.update()
            return
        pix_w, pix_h = self.original_pixmap.width(), self.original_pixmap.height()
        lbl_w, lbl_h = self.width(), self.height()
        if lbl_w == 0 or lbl_h == 0: return
        
        base_scale = min(lbl_w / pix_w, lbl_h / pix_h)
        actual_scale = base_scale * (self.zoom_level / 100.0)
        scaled_w = int(pix_w * actual_scale)
        scaled_h = int(pix_h * actual_scale)
        if scaled_w == 0 or scaled_h == 0: return
        
        scaled_pix = self.original_pixmap.scaled(scaled_w, scaled_h, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
        temp_pixmap = scaled_pix.copy()
        painter = QPainter(temp_pixmap)
        
        for i, box in enumerate(self.boxes):
            rect = QRect(int(box["x"] * actual_scale), int(box["y"] * actual_scale), int(box["w"] * actual_scale), int(box["h"] * actual_scale))
            
            # Detectar si tiene offsets (solo en modo previsualización o filtrada)
            has_offset = False
            if self.view_mode in ["Previsualización", "Filtrada"]:
                offs = box.get("offsets", {})
                if offs.get("clahe", 0) != 0 or offs.get("gauss", 0) != 0 or offs.get("otsu", 0) != 0:
                    has_offset = True

            # Detectar si el esqueleto ha sido modificado (en modo esqueleto)
            has_skeleton_modified = False
            if self.view_mode == "Esqueleto":
                if box.get("esqueleto_modificado", False):
                    has_skeleton_modified = True

            # Si es modificado (offset o esqueleto), usar color amarillo/naranja
            is_modified = has_offset or has_skeleton_modified

            if i == self.active_index:
                color = QColor(9, 105, 218) # Azul GitHub
                pen = QPen(color)
                pen.setWidth(max(2, int(pix_w * 0.003 * actual_scale)))
            elif i == self.hovered_index and self.current_tool != "draw":
                color = QColor(255, 215, 0) if is_modified else QColor(0, 255, 0)
                pen = QPen(color)
                pen.setWidth(max(2, int(pix_w * 0.003 * actual_scale)))
            else:
                color = QColor(255, 215, 0, 180) if is_modified else QColor(0, 255, 0, 120)
                pen = QPen(color)
                pen.setWidth(max(1, int(pix_w * 0.0015 * actual_scale)))
                
            painter.setPen(pen)
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRect(rect)
            
            if is_modified:
                painter.drawText(rect.x() + 2, rect.y() + 12, "*")

            
        if self.is_drawing and self.draw_start_pos and self.draw_current_pos:
            orig_start = self.map_mouse_to_original(self.draw_start_pos); orig_end = self.map_mouse_to_original(self.draw_current_pos)
            if orig_start and orig_end:
                ox1, oy1 = orig_start; ox2, oy2 = orig_end
                rx = min(ox1, ox2); ry = min(oy1, oy2); rw = abs(ox2 - ox1); rh = abs(oy2 - oy1)
                pen_draw = QPen(QColor(0, 150, 255)); pen_draw.setWidth(max(2, int(pix_w * 0.003 * actual_scale))); painter.setPen(pen_draw); painter.setBrush(Qt.BrushStyle.NoBrush)
                painter.drawRect(int(rx * actual_scale), int(ry * actual_scale), int(rw * actual_scale), int(rh * actual_scale))
                
        painter.end()
        self._current_pixmap = temp_pixmap; self.update()

    def paintEvent(self, event):
        if self._current_pixmap and not self._current_pixmap.isNull():
            painter = QPainter(self)
            x = (self.width() - self._current_pixmap.width()) // 2 + self.pan_x
            y = (self.height() - self._current_pixmap.height()) // 2 + self.pan_y
            painter.drawPixmap(x, y, self._current_pixmap); painter.end()
        else: super().paintEvent(event)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self.original_pixmap: self.draw_current_state()


class DialogoHistorial(QDialog):
    def __init__(self, id_usuario, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True); self.id_usuario = id_usuario; self.seleccion = None
        from PyQt6.QtGui import QColor, QIcon
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        
        main_layout = QVBoxLayout(self)
        self.frame = QFrame()
        self.frame.setStyleSheet("""
            QFrame { background-color: #FFFFFF; border-radius: 15px; border: 2px solid #3a61a0; }
            QLabel { color: #333333; border: none; }
            QPushButton { border-radius: 8px; font-weight: bold; padding: 10px; }
        """)
        
        flayout = QVBoxLayout(self.frame)
        
        # Layout de encabezado con título e icono de basura
        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(15, 10, 15, 0)
        
        lbl_titulo = QLabel("<b>Historial de Reportes y Análisis</b>")
        lbl_titulo.setStyleSheet("font-size: 15px; color: #3a61a0;")
        lbl_titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.btn_borrar_icon = QPushButton()
        self.btn_borrar_icon.setIcon(QIcon("assets/buttons/borrar.png"))
        self.btn_borrar_icon.setIconSize(QSize(22, 22))
        self.btn_borrar_icon.setFixedSize(35, 35)
        self.btn_borrar_icon.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_borrar_icon.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #ddf4ff; border-radius: 17px; }
            QPushButton:disabled { opacity: 0.3; }
        """)
        self.btn_borrar_icon.setEnabled(False)
        self.btn_borrar_icon.setToolTip("Borrar seleccionados")
        self.btn_borrar_icon.clicked.connect(self.borrar_reportes_seleccionados)


        header_layout.setContentsMargins(10, 10, 10, 0)
        self.btn_cerrar_x = QPushButton()
        self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))
        self.btn_cerrar_x.setIconSize(QSize(20, 20))
        self.btn_cerrar_x.setFixedSize(35, 35)
        self.btn_cerrar_x.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cerrar_x.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #f6f8fa; border-radius: 17px; }
        """)
        self.btn_cerrar_x.enterEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xroja.png"))
        self.btn_cerrar_x.leaveEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))
        self.btn_cerrar_x.clicked.connect(self.reject)
        
        header_layout.addSpacing(70) # Ajuste para centrar título
        header_layout.addStretch()
        header_layout.addWidget(lbl_titulo)
        header_layout.addStretch()
        header_layout.addWidget(self.btn_borrar_icon)
        header_layout.addWidget(self.btn_cerrar_x)
        
        flayout.addLayout(header_layout)
        
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["Reporte / Imagen", "Fecha", "Estado", "Detecciones", "Descargar"])
        self.tree.setColumnWidth(0, 250)
        self.tree.setColumnWidth(1, 130)
        self.tree.setColumnWidth(2, 100)
        self.tree.setColumnWidth(3, 100)
        self.tree.setColumnWidth(4, 80)
        from PyQt6.QtWidgets import QHeaderView
        self.tree.header().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.tree.header().setStretchLastSection(False)
        self.tree.setIndentation(20)
        self.tree.setAnimated(True)
        from PyQt6.QtWidgets import QAbstractItemView
        self.tree.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        self.tree.setStyleSheet("""
            QTreeWidget { border: 1px solid #d0d7de; border-radius: 8px; background-color: #ffffff; alternate-background-color: #f6f8fa; font-size: 11px; outline: none; }
            QTreeWidget::item { height: 32px; border-bottom: 1px solid #f0f0f0; color: #24292f; }
            QTreeWidget::item:selected { background-color: #0969da; color: #ffffff; }
            QHeaderView::section { background-color: #f6f8fa; padding: 6px; font-weight: bold; border: none; border-bottom: 2px solid #d0d7de; color: #57606a; font-size: 11px; }
        """)
        flayout.addWidget(self.tree)
        
        self.cargar_datos()
        self.tree.itemSelectionChanged.connect(self.actualizar_estado_boton_borrar)
        
        btn_layout = QHBoxLayout()
        
        self.btn_cargar = QPushButton("Cargar / Retomar")
        self.btn_cargar.setEnabled(False)
        self.btn_cargar.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cargar.setStyleSheet("""
            QPushButton { 
                background-color: transparent; 
                border: 2px solid #2da44e; 
                color: #2da44e; 
                border-radius: 8px; 
                font-weight: bold; 
                padding: 10px 20px; 
            }
            QPushButton:hover { 
                background-color: #2da44e; 
                color: white; 
            }
            QPushButton:disabled { 
                background-color: transparent; 
                border: 2px solid #d0d7de; 
                color: #8c959f; 
            }
        """)
        self.btn_cargar.clicked.connect(self.aceptar_seleccion)
        
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_cargar)
        
        flayout.addLayout(btn_layout);        main_layout.addWidget(self.frame)
        self.actualizar_estado_boton_borrar() # Asegurar estado inicial
        self.resize(950, 650)

    def descargar_reporte_id(self, id_reporte):
        from bd.database import conectar
        import json
        conn = conectar(); cur = conn.cursor()
        try:
            cur.execute("SELECT datos_persistentes FROM Analisis WHERE id_reporte = ?", (id_reporte,))
            rows = cur.fetchall()
            all_metrics = []
            for r in rows:
                if r[0]:
                    d = json.loads(r[0])
                    m = d.get("metricas_acumuladas", [])
                    if isinstance(m, list): all_metrics.extend(m)
                    else: all_metrics.append(m)
            
            if not all_metrics:
                from vistas.utilidades import DialogoNotificacion
                DialogoNotificacion("Aviso", "Este reporte no tiene métricas completadas para descargar.", "info", self).exec()
                return
                
            # Usar la lógica global de descarga
            original_metrics = self.parent().metricas_reporte
            self.parent().metricas_reporte = all_metrics
            self.parent().descargar_reporte()
            self.parent().metricas_reporte = original_metrics
            
        except Exception as e:
            from vistas.utilidades import DialogoNotificacion
            DialogoNotificacion("Error", f"Error al descargar: {e}", "error", self).exec()
        finally:
            conn.close()

    def actualizar_estado_boton_borrar(self):
        selected_items = self.tree.selectedItems()
        selected_reports = [item for item in selected_items if item.data(0, Qt.ItemDataRole.UserRole) and item.data(0, Qt.ItemDataRole.UserRole).get("type") == "reporte"]
        
        num_seleccionados = len(selected_reports)
        self.btn_borrar_icon.setEnabled(num_seleccionados > 0)
        # Habilitar cargar SOLO si hay exactamente uno seleccionado
        self.btn_cargar.setEnabled(num_seleccionados == 1)

    def showEvent(self, event):
        super().showEvent(event)
        if self.parent():
            p_geom = self.parent().geometry()
            self.move(p_geom.x() + (p_geom.width() - self.width()) // 2, p_geom.y() + (p_geom.height() - self.height()) // 2)

    def cargar_datos(self):
        from bd.database import conectar
        conn = conectar(); cur = conn.cursor()
        try:
            cur.execute("SELECT id_reporte, nombre_reporte, fecha_creacion, estado FROM Reporte WHERE id_usuario = ? ORDER BY fecha_creacion DESC", (self.id_usuario,))
            reportes = cur.fetchall()
            for rep in reportes:
                id_rep, nombre, fecha, estado = rep
                rep_item = QTreeWidgetItem(self.tree, [nombre, str(fecha), estado, ""])
                rep_item.setData(0, Qt.ItemDataRole.UserRole, {"type": "reporte", "id": id_rep})
                
                cur.execute("""
                    SELECT A.id_analisis, I.ruta_archivo, A.fecha_analisis, A.paso_actual, A.cantidad_microglias
                    FROM Analisis A JOIN Imagen I ON A.id_imagen = I.id_imagen
                    WHERE A.id_reporte = ? ORDER BY A.fecha_analisis ASC
                """, (id_rep,))
                análisis = cur.fetchall()
                for an in análisis:
                    id_an, ruta, f_an, paso, cant = an
                    st = "Completado" if paso >= 5 else f"Paso {paso}"
                    an_item = QTreeWidgetItem(rep_item, [os.path.basename(ruta), str(f_an), st, str(cant)])
                    an_item.setData(0, Qt.ItemDataRole.UserRole, {"type": "analisis", "id": id_an, "id_reporte": id_rep})
                    # Deshabilitar interacción totalmente (solo texto informativo)
                    an_item.setDisabled(True)
                rep_item.setExpanded(True)
                
                # Agregar botón de descarga individual si el reporte está completado
                cur.execute("SELECT COUNT(*) FROM Analisis WHERE id_reporte = ? AND paso_actual >= 5", (id_rep,))
                completados = cur.fetchone()[0]
                
                btn_dl = QPushButton()
                btn_dl.setIcon(QIcon("assets/buttons/download.png"))
                btn_dl.setIconSize(QSize(18, 18))
                btn_dl.setFixedSize(30, 30)
                btn_dl.setCursor(Qt.CursorShape.PointingHandCursor)
                btn_dl.setStyleSheet("""
                    QPushButton { background-color: transparent; border: none; }
                    QPushButton:hover { background-color: #f0f0f0; border-radius: 15px; }
                    QPushButton:disabled { opacity: 0.1; }
                """)
                btn_dl.setEnabled(completados > 0)
                btn_dl.setToolTip("Descargar")
                btn_dl.clicked.connect(lambda checked, r_id=id_rep: self.descargar_reporte_id(r_id))
                
                # Contenedor para centrar el botón en la columna
                container = QWidget()
                layout_c = QHBoxLayout(container)
                layout_c.addWidget(btn_dl)
                layout_c.setAlignment(Qt.AlignmentFlag.AlignCenter)
                layout_c.setContentsMargins(0, 0, 0, 0)
                self.tree.setItemWidget(rep_item, 4, container)
        except Exception as e: logging.error(f"Error historial: {e}")
        finally: conn.close()

    def aceptar_seleccion(self):
        selected_items = self.tree.selectedItems()
        selected_reports = [item for item in selected_items if item.data(0, Qt.ItemDataRole.UserRole) and item.data(0, Qt.ItemDataRole.UserRole).get("type") == "reporte"]
        
        if not selected_reports: return
        
        item_seleccionado = selected_reports[0]
        data = item_seleccionado.data(0, Qt.ItemDataRole.UserRole)
        self.seleccion = {"type": "reporte", "id_reporte": data["id"], "estado": item_seleccionado.text(2)}
        self.accept()


    def borrar_reportes_seleccionados(self):
        # Recopilar todos los reportes seleccionados
        selected_items = self.tree.selectedItems()
        items_a_borrar = [item for item in selected_items if item.data(0, Qt.ItemDataRole.UserRole) and item.data(0, Qt.ItemDataRole.UserRole).get("type") == "reporte"]
        
        if not items_a_borrar:
            from vistas.utilidades import DialogoNotificacion
            DialogoNotificacion("Atención", "Selecciona al menos un reporte de la lista para borrar.", "warning", self).exec()
            return
            
        from vistas.utilidades import DialogoConfirmacion
        msg = f"¿Estás seguro de borrar {len(items_a_borrar)} reporte(s) seleccionado(s) y todos sus datos asociados?"
        if not DialogoConfirmacion("Borrar Selección", msg).exec(): return
        
        from bd.database import conectar
        conn = conectar(); cur = conn.cursor()
        try:
            for item in items_a_borrar:
                data = item.data(0, Qt.ItemDataRole.UserRole)
                id_rep = data["id"]
                cur.execute("DELETE FROM Microglia WHERE id_analisis IN (SELECT id_analisis FROM Analisis WHERE id_reporte = ?)", (id_rep,))
                cur.execute("DELETE FROM Analisis WHERE id_reporte = ?", (id_rep,))
                cur.execute("DELETE FROM Reporte WHERE id_reporte = ?", (id_rep,))
                self.tree.takeTopLevelItem(self.tree.indexOfTopLevelItem(item))
            conn.commit()
        except Exception as e: logging.error(f"Error borrado masivo: {e}")
        finally: 
            conn.close()
            self.actualizar_estado_boton_borrar()


class VentanaInvestigador(QMainWindow):
    def mostrar_notificacion(self, titulo, mensaje, tipo="info"):
        from vistas.utilidades import DialogoNotificacion
        DialogoNotificacion(titulo, mensaje, tipo, self).exec()

    def __init__(self, id_usuario, rol, nombre_usuario=""):
        super().__init__()
        self.id_usuario = id_usuario; self.rol = rol; self.nombre_usuario = nombre_usuario
        self.ruta_imagen_actual = None
        self.pixmaps_globales = {"Original": None, "Filtrada": None, "Esqueleto": None}
        self.crops_en_memoria = {}
        self.crops_filtrados_temp = {}
        self.metadatos_imagen = {"campo": "", "tiempo": ""}
        self.metricas_reporte = []
        self.id_reporte_actual = None
        self.id_analisis_actual = None
        self.paso_actual = 0
        self.metricas_extraidas_ciclo_actual = False
        self.setWindowTitle(f"Prototipo Microglías - Panel ({self.rol})")
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        
        screen_geom = QApplication.primaryScreen().geometry()
        self.resize(int(screen_geom.width() * 0.8), int(screen_geom.height() * 0.8))
        self.setMinimumSize(1050, 700)
        self.inicializar_ui()

    def inicializar_ui(self):
        widget_central = QWidget(); layout_principal = QHBoxLayout()
        self.menu_lateral = QVBoxLayout(); self.menu_lateral.setAlignment(Qt.AlignmentFlag.AlignTop)
        
        # Layout superior del menú lateral para el botón de historial (≡) - Absolute Top
        layout_historial_top = QHBoxLayout()
        layout_historial_top.setContentsMargins(0, 0, 0, 0)
        self.btn_historial = QPushButton()
        self.btn_historial.setIcon(QIcon("assets/buttons/historial.png"))
        self.btn_historial.setIconSize(QSize(22, 22))
        self.btn_historial.setFixedSize(35, 35)
        self.btn_historial.setToolTip("Historial")
        self.btn_historial.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_historial.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #ddf4ff; border-radius: 17px; }
        """)
        
        # Información del usuario al lado derecho
        layout_user_info = QVBoxLayout()
        lbl_rango = QLabel(self.rol)
        lbl_rango.setStyleSheet("color: #007bff; font-weight: bold; font-size: 13px; margin-left: 2px; padding: 0;")
        layout_user_info.addWidget(lbl_rango)

        if self.rol != "Invitado":
            lbl_nombre = QLabel(self.nombre_usuario)
            lbl_nombre.setStyleSheet("color: #666; font-size: 11px; margin-left: 2px; padding: 0;")
            layout_user_info.addWidget(lbl_nombre)
        
        layout_user_info.setSpacing(0)
        
        layout_historial_top.addWidget(self.btn_historial)
        layout_historial_top.addLayout(layout_user_info)
        layout_historial_top.addStretch()
        self.menu_lateral.addLayout(layout_historial_top)
        self.menu_lateral.addSpacing(10)

        self.btn_cargar = QPushButton("Cargar Imagen"); label_caract = QLabel("Caracterización:"); label_caract.setStyleSheet("font-weight: bold; margin-top: 15px;")
        self.btn_conteo = QPushButton("Detectar Microglías"); self.btn_filtrar = QPushButton("Filtrar Imagen"); self.btn_ramas = QPushButton("Mostrar Ramas"); label_reporte = QLabel("Reportes:"); label_reporte.setStyleSheet("font-weight: bold; margin-top: 15px;")
        
        self.btn_obtener_metricas = QPushButton("Obtener métricas")
        self.btn_agregar_imagen_reporte = QPushButton("Agregar imagen")
        self.btn_descargar_reporte = QPushButton("Descargar reporte")
        self.btn_finalizar_reporte = QPushButton("Finalizar reporte")
        
        self.btn_corregir_filtrado = QPushButton("Corregir Filtrado")


        estilo_btn_menu = "QPushButton { background-color: transparent; text-align: left; padding: 8px 10px; font-weight: normal; color: #333333; border: none; font-size: 11px;} QPushButton:hover { background-color: #F0F0F0; border-radius: 5px; } QPushButton:disabled { color: #aaaaaa; }"
        lista_botones = [
            self.btn_cargar, self.btn_conteo, self.btn_filtrar, self.btn_ramas, 
            self.btn_obtener_metricas, self.btn_agregar_imagen_reporte, 
            self.btn_descargar_reporte, self.btn_finalizar_reporte, self.btn_corregir_filtrado
        ]
        for btn in lista_botones: 
            btn.setStyleSheet(estilo_btn_menu)
            self.menu_lateral.addWidget(btn)
            
        self.btn_corregir_filtrado.hide()
        self.btn_corregir_filtrado.setStyleSheet(estilo_btn_menu + "QPushButton { color: #0969da; font-weight: bold; }")


        # Conectar el botón de historial
        self.btn_historial.clicked.connect(self.abrir_historial)

        
        self.frame_filtros = QFrame()
        self.frame_filtros.hide()
        layout_filtros = QVBoxLayout(self.frame_filtros)
        layout_filtros.setContentsMargins(0, 10, 0, 10)
        lbl_f_titulo = QLabel("Ajuste de Filtros Globales"); lbl_f_titulo.setStyleSheet("font-weight: bold; color: #003366;")

        layout_filtros.addWidget(lbl_f_titulo)
        lbl_clahe = QLabel("Contraste (CLAHE):")
        self.sld_clahe = QSlider(Qt.Orientation.Horizontal); self.sld_clahe.setRange(0, 10); self.sld_clahe.setValue(2)
        layout_filtros.addWidget(lbl_clahe); layout_filtros.addWidget(self.sld_clahe)
        lbl_gauss = QLabel("Suavizado Gaussiano:")
        self.sld_gauss = QSlider(Qt.Orientation.Horizontal); self.sld_gauss.setRange(1, 15); self.sld_gauss.setSingleStep(2); self.sld_gauss.setValue(5)
        layout_filtros.addWidget(lbl_gauss); layout_filtros.addWidget(self.sld_gauss)
        lbl_otsu = QLabel("Umbral Binarización:")
        self.sld_otsu = QSlider(Qt.Orientation.Horizontal); self.sld_otsu.setRange(-50, 50); self.sld_otsu.setValue(0)
        layout_filtros.addWidget(lbl_otsu); layout_filtros.addWidget(self.sld_otsu)
        estilo_slider = "QSlider::groove:horizontal { border: 1px solid #d0d7de; height: 4px; background: #f6f8fa; margin: 2px 0; border-radius: 2px; } QSlider::handle:horizontal { background: #ffffff; border: 1px solid #3a61a0; width: 12px; height: 12px; margin: -5px 0; border-radius: 6px; } QSlider::handle:horizontal:hover { background: #eaf2ff; }"
        self.sld_clahe.setStyleSheet(estilo_slider); self.sld_gauss.setStyleSheet(estilo_slider); self.sld_otsu.setStyleSheet(estilo_slider)


        btn_f_layout = QHBoxLayout()
        btn_aceptar_filtro = QPushButton("Aceptar")
        btn_aceptar_filtro.setStyleSheet("background-color: #2da44e; color: white; font-weight: bold; padding: 6px; font-size: 10px; border-radius: 4px;")
        btn_cancelar_filtro = QPushButton("Cancelar")
        btn_cancelar_filtro.setStyleSheet("background-color: #cf222e; color: white; font-weight: bold; padding: 6px; font-size: 10px; border-radius: 4px;")
        btn_f_layout.addWidget(btn_cancelar_filtro)
        btn_f_layout.addWidget(btn_aceptar_filtro)
        layout_filtros.addLayout(btn_f_layout)
        self.menu_lateral.addWidget(self.frame_filtros)
        
        self.sld_clahe.valueChanged.connect(self.previsualizar_filtrado)
        self.sld_gauss.valueChanged.connect(self.previsualizar_filtrado)
        self.sld_otsu.valueChanged.connect(self.previsualizar_filtrado)
        btn_aceptar_filtro.clicked.connect(self.confirmar_filtrado)
        btn_cancelar_filtro.clicked.connect(self.cancelar_filtrado)

        self.menu_lateral.addStretch()
        self.btn_cerrar_sesion = QPushButton("Cerrar Sesión"); self.btn_cerrar_sesion.setStyleSheet("QPushButton { background-color: transparent; border: 2px solid #cc0000; color: #cc0000; font-weight: bold; border-radius: 8px; padding: 10px; margin-top: 20px; } QPushButton:hover { background-color: #cc0000; color: white; }"); self.menu_lateral.addWidget(self.btn_cerrar_sesion)
        frame_menu = QFrame(); frame_menu.setObjectName("menu_lateral"); frame_menu.setFixedWidth(200); frame_menu.setLayout(self.menu_lateral)
        
        area_imagen = QVBoxLayout()
        controles_superiores = QHBoxLayout()
        controles_superiores.setContentsMargins(15, 5, 15, 5)
        controles_superiores.setSpacing(10)
        self.combo_vista = QComboBox(); self.combo_vista.addItem("Original"); self.combo_vista.setMinimumWidth(140); self.combo_vista.setStyleSheet("""
            QComboBox { background-color: white; border: 1px solid #d0d7de; border-radius: 6px; padding: 4px 10px; font-size: 11px; font-weight: bold; color: #24292f; min-width: 140px; }
            QComboBox:hover { background-color: #f6f8fa; }
            QComboBox::drop-down { 
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 20px; 
                border: none;
            }
            QComboBox::down-arrow { 
                image: url(assets/buttons/abajo.png);
                width: 12px; 
                height: 12px;
            }
            QComboBox QAbstractItemView { background-color: white; border: 1px solid #d0d7de; selection-background-color: #eaf2ff; selection-color: #0969da; outline: none; }
        """); self.combo_vista.setEnabled(False); self.combo_vista.currentTextChanged.connect(self.cambiar_vista_global)
        
        # Botones de navegación global
        self.btn_ant_global = QPushButton()
        self.btn_ant_global.setIcon(QIcon("assets/buttons/izq.png"))
        self.btn_ant_global.setIconSize(QSize(12, 12))
        self.btn_ant_global.setFixedSize(22, 22)
        self.btn_ant_global.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_ant_global.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #ddf4ff; border-radius: 11px; }
            QPushButton:disabled { opacity: 0.3; }
        """)
        self.btn_ant_global.clicked.connect(self.anterior_vista_global)
        self.btn_ant_global.setEnabled(False)
        
        self.btn_sig_global = QPushButton()
        self.btn_sig_global.setIcon(QIcon("assets/buttons/der.png"))
        self.btn_sig_global.setIconSize(QSize(12, 12))
        self.btn_sig_global.setFixedSize(22, 22)
        self.btn_sig_global.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_sig_global.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #ddf4ff; border-radius: 11px; }
            QPushButton:disabled { opacity: 0.3; }
        """)
        self.btn_sig_global.clicked.connect(self.siguiente_vista_global)
        self.btn_sig_global.setEnabled(False)
        
        controles_superiores.addWidget(self.btn_ant_global)
        controles_superiores.addWidget(self.combo_vista)
        controles_superiores.addWidget(self.btn_sig_global)
        controles_superiores.addStretch()
        
        self.lbl_info_conteo = QLabel("Microglías detectadas: 0"); self.lbl_info_conteo.setStyleSheet("font-size: 11px; font-weight: bold; color: #3a61a0; background-color: white; border: 1px solid #d0d7de; border-radius: 6px; padding: 4px 10px;"); self.lbl_info_conteo.setAlignment(Qt.AlignmentFlag.AlignCenter); controles_superiores.addWidget(self.lbl_info_conteo); controles_superiores.addSpacing(15)
        
        estilo_herramienta = "QPushButton { background-color: transparent; border: none; padding: 2px; } QPushButton:hover { background-color: #ddf4ff; border-radius: 17px; } QPushButton:checked { background-color: #cce5ff; border: 1px solid #007bff; border-radius: 17px; } QPushButton:disabled { opacity: 0.5; }"
        
        self.btn_herramienta_caja = SafeToolTipButton()
        self.btn_herramienta_caja.setFixedSize(35, 35)
        self.btn_herramienta_caja.setIcon(QIcon("assets/buttons/seleccionar.png"))
        self.btn_herramienta_caja.setIconSize(QSize(20, 20))
        self.btn_herramienta_caja.setCustomToolTip("Crear seleccion")
        self.btn_herramienta_caja.setStyleSheet(estilo_herramienta)
        self.btn_herramienta_caja.setCheckable(True)
        self.btn_herramienta_caja.setEnabled(False)
        self.btn_herramienta_caja.hide() 
        
        self.btn_herramienta_eliminar = SafeToolTipButton()
        self.btn_herramienta_eliminar.setFixedSize(35, 35)
        self.btn_herramienta_eliminar.setIcon(QIcon("assets/buttons/borrar.png"))
        self.btn_herramienta_eliminar.setIconSize(QSize(20, 20))
        self.btn_herramienta_eliminar.setCustomToolTip("Eliminar seleccion")
        self.btn_herramienta_eliminar.setStyleSheet(estilo_herramienta)
        self.btn_herramienta_eliminar.setCheckable(True)
        self.btn_herramienta_eliminar.setEnabled(False)
        self.btn_herramienta_eliminar.hide()
        
        self.btn_herramienta_caja.clicked.connect(self.toggle_herramienta_caja)
        self.btn_herramienta_eliminar.clicked.connect(self.toggle_herramienta_eliminar)

        controles_superiores.addWidget(self.btn_herramienta_caja); controles_superiores.addWidget(self.btn_herramienta_eliminar); controles_superiores.addSpacing(15)
        
        lbl_minus = QLabel("-"); lbl_minus.setStyleSheet("font-size: 24px; font-weight: bold; color: #555;"); lbl_plus = QLabel("+"); lbl_plus.setStyleSheet("font-size: 20px; font-weight: bold; color: #555;")
        self.sld_nivel_zoom = QSlider(Qt.Orientation.Horizontal); self.sld_nivel_zoom.setRange(50, 400); self.sld_nivel_zoom.setValue(100); self.sld_nivel_zoom.setFixedWidth(120); self.sld_nivel_zoom.setEnabled(False)
        self.sld_nivel_zoom.setStyleSheet("""
            QSlider::groove:horizontal { border: 1px solid #d0d7de; height: 4px; background: #f6f8fa; margin: 2px 0; border-radius: 2px; } 
            QSlider::handle:horizontal { background: #ffffff; border: 1px solid #3a61a0; width: 10px; height: 10px; margin: -4px 0; border-radius: 5px; }
            QSlider::handle:horizontal:hover { background: #eaf2ff; }
        """)

        
        self.btn_zoom_reset = QPushButton("↺")
        self.btn_zoom_reset.setFixedSize(35, 35)
        self.btn_zoom_reset.setToolTip("Restablecer Zoom")
        self.btn_zoom_reset.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_zoom_reset.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; font-size: 18px; font-weight: bold; color: #3a61a0; padding: 0; text-align: center; }
            QPushButton:hover { background-color: #ddf4ff; border-radius: 17px; }
            QPushButton:disabled { color: #aaaaaa; }
        """)
        self.btn_zoom_reset.setEnabled(False)
        
        self.btn_bloquear_zoom = SafeToolTipButton()
        self.btn_bloquear_zoom.setFixedSize(35, 35)
        self.btn_bloquear_zoom.setIcon(QIcon("assets/buttons/desbloqueado.png"))
        self.btn_bloquear_zoom.setIconSize(QSize(20, 20))
        self.btn_bloquear_zoom.setCustomToolTip("Bloquear zoom")
        self.btn_bloquear_zoom.setStyleSheet(estilo_herramienta)
        self.btn_bloquear_zoom.setCheckable(True)
        self.btn_bloquear_zoom.setEnabled(False)
        self.btn_bloquear_zoom.toggled.connect(self.toggle_bloqueo_zoom)
        
        controles_superiores.addWidget(lbl_minus); controles_superiores.addWidget(self.sld_nivel_zoom); controles_superiores.addWidget(lbl_plus); controles_superiores.addSpacing(5); controles_superiores.addWidget(self.btn_zoom_reset); controles_superiores.addWidget(self.btn_bloquear_zoom)
        
        self.visor_imagen = InteractiveImageViewer(); self.visor_imagen.setText("Sube una imagen .tiff para empezar el análisis..."); self.visor_imagen.setStyleSheet("border: 2px dashed #aaa; background-color: #f0f0f0; font-size: 18px; color: #666;")
        self.visor_imagen.conteo_actualizado.connect(self.conteo_modificado_auto_save); self.visor_imagen.nueva_caja_dibujada.connect(self.agregar_microglia_manual); self.visor_imagen.nivel_zoom_cambiado.connect(self.sld_nivel_zoom.setValue)
        self.sld_nivel_zoom.valueChanged.connect(self.visor_imagen.set_zoom); self.btn_zoom_reset.clicked.connect(self.reset_zoom)
        
        area_imagen.addLayout(controles_superiores); area_imagen.addWidget(self.visor_imagen, stretch=1)
        layout_principal.addWidget(frame_menu); layout_principal.addLayout(area_imagen, stretch=1); widget_central.setLayout(layout_principal); self.setCentralWidget(widget_central)
        
        self.btn_cargar.clicked.connect(self.cargar_imagen); self.btn_cerrar_sesion.clicked.connect(self.cerrar_sesion); self.btn_conteo.clicked.connect(self.execute_microglia_counting); self.btn_filtrar.clicked.connect(self.ejecutar_filtrado); self.btn_ramas.clicked.connect(self.mostrar_ramas_morfologia); self.btn_corregir_filtrado.clicked.connect(self.corregir_filtrado)
        
        self.btn_obtener_metricas.clicked.connect(self.obtener_metricas)
        self.btn_agregar_imagen_reporte.clicked.connect(self.agregar_imagen_reporte)
        self.btn_descargar_reporte.clicked.connect(self.descargar_reporte)
        self.btn_finalizar_reporte.clicked.connect(self.finalizar_reporte)

        self.actualizar_estado_flujo(0)
        
        if self.rol == "Invitado" or self.rol == "Guest": self.btn_historial.hide()

    def toggle_herramienta_caja(self, checked):
        if checked: self.btn_herramienta_eliminar.setChecked(False); self.visor_imagen.current_tool = "draw"; self.visor_imagen.hovered_index = -1; self.visor_imagen.draw_current_state()
        else: self.visor_imagen.current_tool = "pointer"

    def toggle_herramienta_eliminar(self, checked):
        if checked: self.btn_herramienta_caja.setChecked(False); self.visor_imagen.current_tool = "delete"
        else: self.visor_imagen.current_tool = "pointer"

    def reset_zoom(self): self.sld_nivel_zoom.setValue(100); self.visor_imagen.pan_x = 0; self.visor_imagen.pan_y = 0; self.visor_imagen.update()

    def toggle_bloqueo_zoom(self, checked):
        if checked:
            self.btn_bloquear_zoom.setIcon(QIcon("assets/buttons/bloqueado.png"))
            self.btn_bloquear_zoom.setCustomToolTip("Desbloquear zoom")
        else:
            self.btn_bloquear_zoom.setIcon(QIcon("assets/buttons/desbloqueado.png"))
            self.btn_bloquear_zoom.setCustomToolTip("Bloquear zoom")
            
        self.sld_nivel_zoom.setEnabled(not checked); self.btn_zoom_reset.setEnabled(not checked); self.visor_imagen.lock_zoom(checked)

    def conteo_modificado_auto_save(self, conteo):
        self.actualizar_etiqueta_conteo(conteo)
        self.save_current_progress(mostrar_notif=False)

    def actualizar_estado_flujo(self, paso):
        self.paso_actual = paso
        # Reset report buttons by default
        self.btn_obtener_metricas.setEnabled(False)
        self.btn_agregar_imagen_reporte.setEnabled(False)
        self.btn_descargar_reporte.setEnabled(False)
        self.btn_finalizar_reporte.setEnabled(False)

        # Activar/Desactivar controles de zoom a partir del conteo (paso 1 o mayor)
        zoom_activado = (paso >= 1)
        self.sld_nivel_zoom.setEnabled(zoom_activado)
        self.btn_zoom_reset.setEnabled(zoom_activado)
        self.btn_bloquear_zoom.setEnabled(zoom_activado)

        # Habilitación estricta de botones según la etapa actual de análisis (se desactiva la etapa anterior)
        self.btn_cargar.setEnabled(paso == 0)
        self.btn_conteo.setEnabled(paso == 1)
        self.btn_filtrar.setEnabled(paso == 2)
        self.btn_ramas.setEnabled(paso == 3)

        if paso == 0:
            self.combo_vista.setEnabled(False)
            self.btn_herramienta_caja.hide(); self.btn_herramienta_eliminar.hide()
        elif paso == 1:
            self.combo_vista.setEnabled(False)
            self.btn_herramienta_caja.hide(); self.btn_herramienta_eliminar.hide()
        elif paso == 2:
            self.combo_vista.setEnabled(True)
            self.btn_herramienta_caja.show(); self.btn_herramienta_eliminar.show()
            self.btn_herramienta_caja.setEnabled(True); self.btn_herramienta_eliminar.setEnabled(True)
        elif paso == 3:
            self.combo_vista.setEnabled(True)
            self.btn_herramienta_caja.hide(); self.btn_herramienta_caja.setChecked(False)
            self.btn_herramienta_eliminar.hide(); self.btn_herramienta_eliminar.setChecked(False)
            self.visor_imagen.current_tool = "pointer"
        elif paso == 4:
            self.btn_corregir_filtrado.show()
            self.btn_herramienta_caja.hide(); self.btn_herramienta_eliminar.hide()
            self.btn_obtener_metricas.setEnabled(True)
        elif paso == 5:
            self.btn_corregir_filtrado.show()
            self.btn_herramienta_caja.hide(); self.btn_herramienta_eliminar.hide()
            self.btn_obtener_metricas.setEnabled(False)
            self.btn_agregar_imagen_reporte.setEnabled(True)
            self.btn_descargar_reporte.setEnabled(True)
            self.btn_finalizar_reporte.setEnabled(True)
        
        if paso not in [4, 5]: self.btn_corregir_filtrado.hide()
        
        self.actualizar_botones_navegacion()
        
        # Guardado automático de progreso al cambiar de fase
        if paso > 0:
            self.save_current_progress(mostrar_notif=False)

    def save_current_progress(self, mostrar_notif=True):
        if not self.ruta_imagen_actual: return
        
        from bd.database import conectar
        import json
        conn = conectar(); cur = conn.cursor()
        
        try:
            # 1. Asegurar Reporte (Grupo)
            if not self.id_reporte_actual:
                cur.execute("INSERT INTO Reporte (id_usuario, nombre_reporte) VALUES (?, ?)", 
                           (self.id_usuario, f"Reporte {datetime.now().strftime('%Y-%m-%d %H:%M')}"))
                self.id_reporte_actual = cur.lastrowid

            # 2. Asegurar Imagen
            cur.execute("SELECT id_imagen FROM Imagen WHERE ruta_archivo = ?", (self.ruta_imagen_actual,))
            res = cur.fetchone()
            if not res:
                cur.execute("INSERT INTO Imagen (id_usuario, ruta_archivo, formato, campo, tiempo_muestra) VALUES (?,?,?,?,?)",
                           (self.id_usuario, self.ruta_imagen_actual, 'TIFF', self.metadatos_imagen.get("campo"), self.metadatos_imagen.get("tiempo")))
                id_img = cur.lastrowid
            else: id_img = res[0]

            # 3. Guardar Análisis con Estado Persistente
            datos = {
                "boxes": self.visor_imagen.boxes,
                "metricas_acumuladas": self.metricas_reporte
            }
            datos_json = json.dumps(datos)

            if not self.id_analisis_actual:
                cur.execute("INSERT INTO Analisis (id_reporte, id_imagen, cantidad_microglias, paso_actual, datos_persistentes) VALUES (?,?,?,?,?)",
                           (self.id_reporte_actual, id_img, len(self.visor_imagen.boxes), self.paso_actual, datos_json))
                self.id_analisis_actual = cur.lastrowid
            else:
                cur.execute("UPDATE Analisis SET cantidad_microglias = ?, paso_actual = ?, datos_persistentes = ? WHERE id_analisis = ?",
                           (len(self.visor_imagen.boxes), self.paso_actual, datos_json, self.id_analisis_actual))

            # 4. Sincronizar Microglias (Detecciones individuales)
            cur.execute("DELETE FROM Microglia WHERE id_analisis = ?", (self.id_analisis_actual,))
            for box in self.visor_imagen.boxes:
                cur.execute("INSERT INTO Microglia (id_analisis, centroide_x, centroide_y, area_total_pixeles, perimetro, bbox_x, bbox_y, bbox_w, bbox_h, crop_path) VALUES (?,0,0,0,0,?,?,?,?,?)",
                           (self.id_analisis_actual, box['x'], box['y'], box['w'], box['h'], box.get('crop_path', '')))
            
            conn.commit()
            if mostrar_notif:
                self.mostrar_notificacion("Éxito", "Progreso guardado correctamente.", "info")
        except Exception as e:
            self.mostrar_notificacion("Error", f"No se pudo guardar: {e}", "error")
        finally: conn.close()

    def abrir_historial(self):
        diag = DialogoHistorial(self.id_usuario, self)
        if diag.exec() and diag.seleccion:
            self.cargar_reporte_especifico(diag.seleccion)

    def cargar_reporte_especifico(self, seleccion):
        id_reporte = seleccion["id_reporte"]
        from bd.database import conectar
        import json
        conn = conectar(); cur = conn.cursor()
        try:
            # 1. Obtener el último análisis de este reporte
            cur.execute("""
                SELECT A.id_analisis, I.ruta_archivo, I.campo, I.tiempo_muestra, A.paso_actual, A.datos_persistentes
                FROM Analisis A JOIN Imagen I ON A.id_imagen = I.id_imagen
                WHERE A.id_reporte = ? ORDER BY A.id_analisis DESC LIMIT 1
            """, (id_reporte,))
            res = cur.fetchone()
            
            self.id_reporte_actual = id_reporte
            
            if not res:
                # Reporte vacío? (No debería pasar)
                self.id_analisis_actual = None; self.metricas_reporte = []
                self.actualizar_estado_flujo(0)
                return

            id_an, ruta, campo, tiempo, paso, datos_json = res
            
            # 2. Decidir si retomar o empezar nueva imagen
            if paso < 5:
                # RETOMAR ANÁLISIS INCOMPLETO
                self.ruta_imagen_actual = ruta
                self.metadatos_imagen = {"campo": campo, "tiempo": tiempo}
                self.id_analisis_actual = id_an
                self.paso_actual = paso

                if datos_json:
                    datos = json.loads(datos_json)
                    boxes = datos.get("boxes", [])
                    for b in boxes:
                        if "crop_path" in b and isinstance(b["crop_path"], str):
                            b["crop_path"] = b["crop_path"].replace("\\", "/")
                    self.metricas_reporte = datos.get("metricas_acumuladas", [])
                else: boxes = []; self.metricas_reporte = []

                # Cargar imagen
                from PyQt6.QtGui import QImage, QPixmap
                import cv2; import numpy as np
                cv_img = cv2.imread(ruta, cv2.IMREAD_UNCHANGED)
                if cv_img is not None:
                    if cv_img.dtype == np.uint16: cv_img = ((cv_img-cv_img.min())/(cv_img.max()-cv_img.min())*255).astype(np.uint8)
                    if len(cv_img.shape) == 2: h, w = cv_img.shape; qimg = QImage(cv_img.data, w, h, w, QImage.Format.Format_Grayscale8)
                    else: cv_img = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB); h, w, ch = cv_img.shape; qimg = QImage(cv_img.data, w, h, ch * w, QImage.Format.Format_RGB888)
                    pixmap = QPixmap.fromImage(qimg)
                else: pixmap = QPixmap(ruta)
                
                self.pixmaps_globales = {"Original": pixmap, "Filtrada": None, "Esqueleto": None}
                self.visor_imagen.set_image_and_boxes(pixmap, boxes)
                
                # Reconstruir la lista del combo_vista y las imágenes globales según el paso alcanzado
                self.combo_vista.blockSignals(True)
                # Limpiar opciones extra del combo (mantener solo Original)
                while self.combo_vista.count() > 1:
                    self.combo_vista.removeItem(1)
                
                if paso >= 3:
                    pix_f = self.construir_imagen_global("filtradas")
                    if pix_f:
                        self.pixmaps_globales["Filtrada"] = pix_f
                        self.combo_vista.addItem("Filtrada")
                
                if paso >= 4:
                    pix_e = self.construir_imagen_global("esqueletos")
                    if pix_e:
                        self.pixmaps_globales["Esqueleto"] = pix_e
                        self.combo_vista.addItem("Esqueleto")
                
                # Seleccionar la vista correspondiente al paso donde se quedó el análisis
                if paso >= 4:
                    modo_default = "Esqueleto"
                elif paso == 3:
                    modo_default = "Filtrada"
                else:
                    modo_default = "Original"

                self.combo_vista.setCurrentText(modo_default)
                self.combo_vista.blockSignals(False)
                self.combo_vista.setEnabled(paso >= 2)
                
                # Sincronizar el visor con el modo cargado
                self.visor_imagen.set_view_mode(modo_default, self.pixmaps_globales.get(modo_default))
                self.actualizar_estado_flujo(paso)
                self.actualizar_botones_navegacion()
                self.mostrar_notificacion("Éxito", f"Continuando análisis: {os.path.basename(ruta)}", "info")
            else:
                # ÚLTIMA IMAGEN COMPLETADA -> CARGAR MÉTRICAS Y PEDIR NUEVA IMAGEN
                if datos_json:
                    datos = json.loads(datos_json)
                    self.metricas_reporte = datos.get("metricas_acumuladas", [])
                    # Importante: si la última imagen se terminó, sus métricas deben estar en el acumulado.
                    # Si no están, las buscamos en el propio análisis (esto es una salvaguarda)
                else: self.metricas_reporte = []
                
                self.id_analisis_actual = None; self.ruta_imagen_actual = None; self.paso_actual = 0
                self.visor_imagen.set_image_and_boxes(None, [])
                self.actualizar_estado_flujo(0)
                self.mostrar_notificacion("Reporte Cargado", "Última imagen completada. Por favor, añade una nueva imagen para continuar el reporte.", "info")

        except Exception as e: self.mostrar_notificacion("Error", f"Fallo al cargar reporte: {e}", "error")
        finally: conn.close()


    def actualizar_etiqueta_conteo(self, conteo): self.lbl_info_conteo.setText(f"Microglías detectadas: {conteo}")

    def cargar_imagen(self):
        dialogo_intermedio = DialogoCargarImagen(self)
        if dialogo_intermedio.exec():
            ruta_archivo = dialogo_intermedio.ruta_seleccionada
            campo = dialogo_intermedio.campo_val
            tiempo = dialogo_intermedio.tiempo_val
            
            # Check for duplicates in current session report
            for item in self.metricas_reporte:
                if item["campo"] == campo and item["tiempo"] == tiempo:
                    from vistas.utilidades import DialogoConfirmacion
                    diag = DialogoConfirmacion("Advertencia de Duplicado", f"Ya existen métricas guardadas para el Campo '{campo}' y Tiempo '{tiempo}'.\n\n¿Deseas continuar de todos modos?")
                    if not diag.exec():
                        return # User cancelled
                    break
                    
            self.metadatos_imagen["campo"] = campo; self.metadatos_imagen["tiempo"] = tiempo
            if ruta_archivo:
                self.ruta_imagen_actual = ruta_archivo; pixmap = QPixmap(ruta_archivo)
                if pixmap.isNull():
                    try:
                        import cv2; import numpy as np
                        cv_img = cv2.imread(ruta_archivo, cv2.IMREAD_UNCHANGED)
                        if cv_img is not None:
                            if cv_img.dtype == np.uint16: cv_img = ((cv_img - cv_img.min()) / (cv_img.max() - cv_img.min()) * 255).astype(np.uint8)
                            if len(cv_img.shape) == 2: h, w = cv_img.shape; qimg = QImage(cv_img.data, w, h, w, QImage.Format.Format_Grayscale8)
                            else: cv_img = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB); h, w, ch = cv_img.shape; qimg = QImage(cv_img.data, w, h, ch * w, QImage.Format.Format_RGB888)
                            pixmap = QPixmap.fromImage(qimg)
                    except Exception as e: logging.error(f"Error al cargar imagen: {e}")
                if not pixmap.isNull():
                    self.pixmaps_globales["Original"] = pixmap; self.pixmaps_globales["Filtrada"] = None; self.pixmaps_globales["Esqueleto"] = None; self.btn_herramienta_caja.setChecked(False); self.btn_herramienta_eliminar.setChecked(False); self.visor_imagen.current_tool = "pointer"; self.btn_bloquear_zoom.setChecked(False); self.reset_zoom(); self.visor_imagen.set_image_and_boxes(pixmap, []); self.actualizar_estado_flujo(1); self.combo_vista.blockSignals(True); self.combo_vista.clear(); self.combo_vista.addItem("Original"); self.combo_vista.setCurrentText("Original"); self.combo_vista.blockSignals(False); self.actualizar_botones_navegacion(); self.visor_imagen.view_mode = "Original"; self.mostrar_notificacion("Imagen cargada", "Imagen lista para el análisis.", "info")
                else: self.mostrar_notificacion("Error", "El archivo está corrupto o no es válido.", "error")

    def cambiar_vista_global(self, texto_vista):
        pixmap_guardado = self.pixmaps_globales.get(texto_vista)
        if pixmap_guardado:
            self.visor_imagen.set_view_mode(texto_vista, pixmap_guardado)
            self.actualizar_botones_navegacion()
        else:
            self.mostrar_notificacion("Aviso", f"Aún no has generado el paso: {texto_vista}.", "warning")
            self.combo_vista.blockSignals(True)
            self.combo_vista.setCurrentText(self.visor_imagen.view_mode)
            self.combo_vista.blockSignals(False)

    def anterior_vista_global(self):
        idx = self.combo_vista.currentIndex()
        if idx > 0:
            self.combo_vista.setCurrentIndex(idx - 1)

    def siguiente_vista_global(self):
        idx = self.combo_vista.currentIndex()
        if idx < self.combo_vista.count() - 1:
            self.combo_vista.setCurrentIndex(idx + 1)

    def actualizar_botones_navegacion(self):
        idx = self.combo_vista.currentIndex()
        self.btn_ant_global.setEnabled(idx > 0)
        self.btn_sig_global.setEnabled(idx < self.combo_vista.count() - 1)

    def cerrar_sesion(self):
        from vistas.login import VentanaLogin
        self.ventana_login = VentanaLogin(); self.ventana_login.setObjectName("ventana_login"); self.ventana_login.show(); self.close()

    def execute_microglia_counting(self):
        if not self.ruta_imagen_actual: self.mostrar_notificacion("Advertencia", "Por favor, carga una imagen primero.", "warning"); return
        
        # Si ya se ejecutó y estamos en un paso avanzado, dar opción de sólo volver a editar o reprocesar con IA
        if self.paso_actual >= 2:
            from PyQt6.QtWidgets import QMessageBox
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Question)
            msg.setWindowTitle("Regresar a Detección")
            msg.setText("¿Qué deseas hacer al regresar al paso de Detección (Conteo)?")
            
            btn_edit = msg.addButton("Editar detecciones manualmente", QMessageBox.ButtonRole.ActionRole)
            btn_yolo = msg.addButton("Volver a procesar con IA (YOLO)", QMessageBox.ButtonRole.ActionRole)
            btn_cancel = msg.addButton("Cancelar", QMessageBox.ButtonRole.RejectRole)
            
            msg.exec()
            clicked = msg.clickedButton()
            if clicked == btn_edit:
                self.actualizar_estado_flujo(2)
                self.combo_vista.blockSignals(True)
                self.combo_vista.setCurrentText("Original")
                self.combo_vista.blockSignals(False)
                self.cambiar_vista_global("Original")
                return
            elif clicked == btn_cancel or clicked is None:
                return
                
        dialogo = DialogoCarga("Cargando IA y aplicando conteo...\nPor favor, espera.", self); dialogo.show()
        from PyQt6.QtWidgets import QApplication; QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor); QApplication.processEvents()
        try:
            from ia.modelo_yolo import MicrogliaProcessor
            model_path = os.path.join(os.getcwd(), "ia", "entrenamiento_resultados", "modelo_microglias5", "weights", "best.pt"); output_dir = os.path.join(os.getcwd(), "analisis_resultados")
            processor = MicrogliaProcessor(model_path=model_path); resultado = processor.process_and_crop(self.ruta_imagen_actual, base_output_folder=output_dir)
            if len(resultado) == 3:
                crops_folder, count, boxes_data = resultado
                for box in boxes_data: box["offsets"] = {"clahe": 0, "gauss": 0, "otsu": 0}; box["removal_areas"] = []

                self.visor_imagen.set_image_and_boxes(self.pixmaps_globales["Original"], boxes_data)
            else:
                crops_folder, count = resultado; self.visor_imagen.set_image_and_boxes(self.pixmaps_globales["Original"], [])


            dialogo.close(); QApplication.restoreOverrideCursor(); self.actualizar_estado_flujo(2); self.mostrar_notificacion("2. Detección", f"Se detectaron {count} posibles microglías.\n\nUsa las herramientas superiores si necesitas agregar o eliminar selecciones.", "info")
        except Exception as e: dialogo.close(); QApplication.restoreOverrideCursor(); self.mostrar_notificacion("Error", str(e), "error")

    def agregar_microglia_manual(self, x, y, w, h):
        if not self.ruta_imagen_actual: return
        area_nueva = w * h
        for box in self.visor_imagen.boxes:
            bx, by, bw, bh = box["x"], box["y"], box["w"], box["h"]; ix1 = max(x, bx); iy1 = max(y, by); ix2 = min(x + w, bx + bw); iy2 = min(y + h, by + bh); iw_inter = max(0, ix2 - ix1); ih_inter = max(0, iy2 - iy1); inter_area = iw_inter * ih_inter
            if inter_area > 0:
                ioa_nueva = inter_area / area_nueva; ioa_existente = inter_area / (bw * bh)
                if ioa_nueva >= 0.40 or ioa_existente >= 0.40: self.mostrar_notificacion("Acción bloqueada", "No puedes crear una microglía que esté contenida dentro de otra existente o que se superponga fuertemente.", "warning"); return
        base_name = Path(self.ruta_imagen_actual).stem; crops_folder = os.path.join(os.getcwd(), "analisis_resultados", base_name, "crops"); os.makedirs(crops_folder, exist_ok=True)
        orig_pixmap = self.pixmaps_globales["Original"]
        if not orig_pixmap: return
        rect_recorte = QRect(x, y, w, h); pixmap_recorte = orig_pixmap.copy(rect_recorte); nombre_archivo = f"manual_{uuid.uuid4().hex[:6]}.png"; ruta_guardado = os.path.join(crops_folder, nombre_archivo); pixmap_recorte.save(ruta_guardado, "PNG")
        nueva_caja = {"x": x, "y": y, "w": w, "h": h, "crop_path": ruta_guardado, "offsets": {"clahe": 0, "gauss": 0, "otsu": 0}, "removal_areas": []}

        self.visor_imagen.boxes.append(nueva_caja); self.visor_imagen.draw_current_state(); self.actualizar_etiqueta_conteo(len(self.visor_imagen.boxes))
        self.save_current_progress(mostrar_notif=False)


    def construir_imagen_global(self, carpeta_origen):
        import cv2; import numpy as np
        orig_pixmap = self.pixmaps_globales["Original"]; orig_w = orig_pixmap.width(); orig_h = orig_pixmap.height(); lienzo = np.zeros((orig_h, orig_w), dtype=np.uint8); base_name = Path(self.ruta_imagen_actual).stem; base_dir = os.path.join(os.getcwd(), "analisis_resultados", base_name)
        for box in self.visor_imagen.boxes:
            x, y, w, h = int(box["x"]), int(box["y"]), int(box["w"]), int(box["h"]); nombre_archivo = os.path.basename(box["crop_path"]); ruta_recorte = os.path.join(base_dir, carpeta_origen, nombre_archivo)
            if os.path.exists(ruta_recorte):
                with open(ruta_recorte, "rb") as f: file_bytes = bytearray(f.read())
                img_array = np.asarray(file_bytes, dtype=np.uint8); recorte = cv2.imdecode(img_array, cv2.IMREAD_GRAYSCALE)
                if recorte is not None: rh, rw = recorte.shape; y_fin = min(y + rh, orig_h); x_fin = min(x + rw, orig_w); h_real = y_fin - y; w_real = x_fin - x; lienzo[y:y_fin, x:x_fin] = recorte[:h_real, :w_real]
        qimg = QImage(lienzo.data, orig_w, orig_h, orig_w, QImage.Format.Format_Grayscale8).copy()
        return QPixmap.fromImage(qimg)

    def construir_imagen_global_memoria(self):
        import numpy as np
        orig_pixmap = self.pixmaps_globales["Original"]; orig_w = orig_pixmap.width(); orig_h = orig_pixmap.height(); lienzo = np.zeros((orig_h, orig_w), dtype=np.uint8)
        for box in self.visor_imagen.boxes:
            x, y, w, h = int(box["x"]), int(box["y"]), int(box["w"]), int(box["h"])
            nombre_archivo = os.path.basename(box["crop_path"])
            recorte = self.crops_filtrados_temp.get(nombre_archivo)
            if recorte is not None:
                rh, rw = recorte.shape; y_fin = min(y + rh, orig_h); x_fin = min(x + rw, orig_w); h_real = y_fin - y; w_real = x_fin - x
                lienzo[y:y_fin, x:x_fin] = recorte[:h_real, :w_real]
        qimg = QImage(lienzo.data, orig_w, orig_h, orig_w, QImage.Format.Format_Grayscale8).copy()
        return QPixmap.fromImage(qimg)

    def ejecutar_filtrado(self):
        if not self.ruta_imagen_actual or not self.visor_imagen.boxes: self.mostrar_notificacion("Advertencia", "Aplica el conteo primero.", "warning"); return
        
        # Guardar paso previo para poder regresar al cancelar
        self.paso_previo = self.paso_actual
        
        import cv2; import numpy as np
        self.crops_en_memoria.clear()
        self.crops_filtrados_temp.clear()
        
        from PyQt6.QtWidgets import QApplication
        dialogo = DialogoCarga("Cargando imágenes a memoria...\nPor favor, espera.", self); dialogo.show(); QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor); QApplication.processEvents()
        
        for box in self.visor_imagen.boxes:
            crop_path = box["crop_path"]
            if os.path.exists(crop_path):
                with open(crop_path, "rb") as f: file_bytes = bytearray(f.read())
                img_array = np.asarray(file_bytes, dtype=np.uint8)
                img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
                if img is not None:
                    nombre = os.path.basename(crop_path)
                    self.crops_en_memoria[nombre] = img
                    
        dialogo.close(); QApplication.restoreOverrideCursor()
        
        self.frame_filtros.show()
        for btn in [self.btn_cargar, self.btn_conteo, self.btn_filtrar, self.btn_ramas, self.btn_obtener_metricas, self.btn_agregar_imagen_reporte, self.btn_descargar_reporte, self.btn_finalizar_reporte]: btn.setEnabled(False)
        
        # Ocultar herramientas manuales al iniciar el filtrado
        self.btn_herramienta_caja.setChecked(False)
        self.btn_herramienta_eliminar.setChecked(False)
        self.btn_herramienta_caja.hide()
        self.btn_herramienta_eliminar.hide()
        self.visor_imagen.current_tool = "pointer"
        
        self.combo_vista.setEnabled(True)
        items_combo = [self.combo_vista.itemText(i) for i in range(self.combo_vista.count())]
        if "Previsualización" not in items_combo:
            self.combo_vista.addItem("Previsualización")
        
        self.combo_vista.blockSignals(True)
        self.combo_vista.setCurrentText("Previsualización")
        self.combo_vista.blockSignals(False)
        self.actualizar_botones_navegacion()
        self.visor_imagen.view_mode = "Previsualización"
        
        self.previsualizar_filtrado()

    def previsualizar_filtrado(self, *args):
        if not self.crops_en_memoria: return
        import cv2; import numpy as np
        g_clahe_clip = self.sld_clahe.value()
        g_k_val = self.sld_gauss.value()
        g_otsu_offset = self.sld_otsu.value()
        
        # Mapa de nombre a info (offsets y areas) para eficiencia
        mapa_info = {}
        for box in self.visor_imagen.boxes:
            nombre = os.path.basename(box["crop_path"])
            mapa_info[nombre] = {
                "offsets": box.get("offsets", {"clahe":0, "gauss":0, "otsu":0}),
                "removal_areas": box.get("removal_areas", [])
            }

        for nombre, img in self.crops_en_memoria.items():
            info = mapa_info.get(nombre, {"offsets": {"clahe":0, "gauss":0, "otsu":0}, "removal_areas": []})
            offsets = info["offsets"]
            
            # Aplicar Global + Offset
            c_clip = max(0, min(10, g_clahe_clip + offsets["clahe"]))
            k_val = max(1, min(15, g_k_val + offsets["gauss"]))
            k = k_val if k_val % 2 != 0 else k_val + 1
            o_offset = g_otsu_offset + offsets["otsu"]
            
            clahe = cv2.createCLAHE(clipLimit=float(c_clip), tileGridSize=(8, 8)) if c_clip > 0 else None
            
            if clahe is not None:
                hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
                h_c, s_c, v_c = cv2.split(hsv)
                v_clahe = clahe.apply(v_c)
                hsv_clahe = cv2.merge((h_c, s_c, v_clahe))
                bgr_proc = cv2.cvtColor(hsv_clahe, cv2.COLOR_HSV2BGR)
            else:
                bgr_proc = img.copy()
                
            gray = cv2.cvtColor(bgr_proc, cv2.COLOR_BGR2GRAY)
            blur = cv2.GaussianBlur(gray, (k, k), 0)
            ret, _ = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            threshold_val = max(0, min(255, ret + o_offset))
            _, bin_img = cv2.threshold(blur, threshold_val, 255, cv2.THRESH_BINARY)
            
            # Aplicar eliminación de áreas manuales
            for area in info["removal_areas"]:
                ax, ay, aw, ah = area["x"], area["y"], area["w"], area["h"]
                cv2.rectangle(bin_img, (ax, ay), (ax + aw, ay + ah), 0, -1)
            
            self.crops_filtrados_temp[nombre] = bin_img

            
        pixmap_filtrada = self.construir_imagen_global_memoria()
        self.pixmaps_globales["Previsualización"] = pixmap_filtrada
        self.visor_imagen.set_view_mode("Previsualización", pixmap_filtrada)


    def confirmar_filtrado(self):
        base_name = Path(self.ruta_imagen_actual).stem; filtradas_dir = os.path.join(os.getcwd(), "analisis_resultados", base_name, "filtradas"); os.makedirs(filtradas_dir, exist_ok=True); import cv2; count = 0
        from PyQt6.QtWidgets import QApplication
        try:
            dialogo = DialogoCarga("Guardando filtros aplicados...\nPor favor, espera.", self); dialogo.show(); QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor); QApplication.processEvents()
            for nombre, bin_img in self.crops_filtrados_temp.items():
                out_path = os.path.join(filtradas_dir, nombre)
                is_success, im_buf_arr = cv2.imencode(".png", bin_img)
                if is_success: im_buf_arr.tofile(out_path); count += 1
            dialogo.close(); QApplication.restoreOverrideCursor()
            
            if count > 0:
                self.pixmaps_globales["Filtrada"] = self.pixmaps_globales["Previsualización"]
                self.combo_vista.blockSignals(True)
                items_combo = [self.combo_vista.itemText(i) for i in range(self.combo_vista.count())]
                if "Filtrada" not in items_combo: self.combo_vista.addItem("Filtrada")
                idx = self.combo_vista.findText("Previsualización")
                if idx >= 0: self.combo_vista.removeItem(idx)
                self.combo_vista.setCurrentText("Filtrada")
                self.combo_vista.blockSignals(False)
                self.visor_imagen.view_mode = "Filtrada"
                
                self.frame_filtros.hide()
                self.actualizar_estado_flujo(3)
                self.mostrar_notificacion("3. Filtrado", f"Se aplicaron los filtros a {count} microglías.", "info")
            else: self.mostrar_notificacion("Error", "No se guardó ninguna imagen.", "error")
        except Exception as error: dialogo.close(); QApplication.restoreOverrideCursor(); self.mostrar_notificacion("Error", f"Falló el guardado: {str(error)}", "error")

    def cancelar_filtrado(self):
        self.crops_en_memoria.clear()
        self.crops_filtrados_temp.clear()
        self.frame_filtros.hide()
        
        self.combo_vista.blockSignals(True)
        idx = self.combo_vista.findText("Previsualización")
        if idx >= 0: self.combo_vista.removeItem(idx)
        
        # Volver a la vista previa del paso restaurado
        paso_restaurar = getattr(self, "paso_previo", 2)
        vista_restaurar = "Original"
        if paso_restaurar >= 3:
            vista_restaurar = "Filtrada"
        if paso_restaurar >= 4:
            vista_restaurar = "Esqueleto"
            
        self.combo_vista.setCurrentText(vista_restaurar)
        self.combo_vista.blockSignals(False)
        self.cambiar_vista_global(vista_restaurar)
        
        self.actualizar_estado_flujo(paso_restaurar)

    def mostrar_ramas_morfologia(self):
        if not self.ruta_imagen_actual or not self.visor_imagen.boxes: self.mostrar_notificacion("Advertencia", "Aplica el conteo y filtrado primero.", "warning"); return
        base_name = Path(self.ruta_imagen_actual).stem; filtradas_dir = os.path.join(os.getcwd(), "analisis_resultados", base_name, "filtradas"); esqueletos_dir = os.path.join(os.getcwd(), "analisis_resultados", base_name, "esqueletos"); os.makedirs(esqueletos_dir, exist_ok=True); import cv2; import numpy as np; from skimage.morphology import skeletonize
        from PyQt6.QtWidgets import QApplication; count = 0
        try:
            dialogo = DialogoCarga("Generando esqueletos topológicos...\nPor favor, espera.", self); dialogo.show(); QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor); QApplication.processEvents()
            for box in self.visor_imagen.boxes:
                crop_path = box["crop_path"]; nombre = os.path.basename(crop_path); fil_path = os.path.join(filtradas_dir, nombre)
                if os.path.exists(fil_path):
                    with open(fil_path, "rb") as f: file_bytes = bytearray(f.read())
                    img_array = np.asarray(file_bytes, dtype=np.uint8); img_raw = cv2.imdecode(img_array, cv2.IMREAD_GRAYSCALE)
                    if img_raw is not None: _, bin_img = cv2.threshold(img_raw, 127, 255, cv2.THRESH_BINARY); img_bool = bin_img > 0; skeleton = skeletonize(img_bool); skeleton_img = (skeleton * 255).astype(np.uint8); out_path = os.path.join(esqueletos_dir, nombre); is_success, im_buf_arr = cv2.imencode(".png", skeleton_img)
                    if is_success: im_buf_arr.tofile(out_path); count += 1
            dialogo.close(); QApplication.restoreOverrideCursor()
            if count > 0: pixmap_esqueleto = self.construir_imagen_global("esqueletos"); self.pixmaps_globales["Esqueleto"] = pixmap_esqueleto; self.actualizar_estado_flujo(4); self.combo_vista.addItem("Esqueleto"); self.combo_vista.setCurrentText("Esqueleto"); self.mostrar_notificacion("4. Esqueleto (Ramas)", f"Se generaron {count} esqueletos topológicos.\n\nYa puedes avanzar a obtener las métricas finales.", "info")
            else: self.mostrar_notificacion("Advertencia", "No se generaron esqueletos. Verifica la carpeta de filtrado.", "warning")
        except Exception as error: dialogo.close(); QApplication.restoreOverrideCursor(); self.mostrar_notificacion("Error de Procesamiento", f"Falló el cálculo:\n{str(error)}", "error")

    def corregir_filtrado(self):
        from vistas.utilidades import DialogoConfirmacion
        diag = DialogoConfirmacion("Corregir Filtrado", "¿Estás seguro de que deseas eliminar el filtrado y esqueletizado actual para volver a ajustar los parámetros?")
        if not diag.exec(): return
        
        self.pixmaps_globales["Filtrada"] = None
        self.pixmaps_globales["Esqueleto"] = None
        
        self.combo_vista.blockSignals(True)
        for label in ["Esqueleto", "Filtrada"]:
            idx = self.combo_vista.findText(label)
            if idx >= 0: self.combo_vista.removeItem(idx)
        self.combo_vista.setCurrentText("Original")
        self.combo_vista.blockSignals(False)
        self.cambiar_vista_global("Original")
        
        if self.ruta_imagen_actual:
            try:
                import shutil
                base_name = Path(self.ruta_imagen_actual).stem
                base_dir = os.path.join(os.getcwd(), "analisis_resultados", base_name)
                for sub in ["filtradas", "esqueletos"]:
                    folder = os.path.join(base_dir, sub)
                    if os.path.exists(folder): shutil.rmtree(folder); os.makedirs(folder, exist_ok=True)
            except Exception as e: logging.error(f"Error al limpiar carpetas: {e}")
            
        # Resetear sliders globales y offsets individuales
        self.sld_clahe.setValue(2)
        self.sld_gauss.setValue(5)
        self.sld_otsu.setValue(0)
        for box in self.visor_imagen.boxes:
            box["offsets"] = {"clahe": 0, "gauss": 0, "otsu": 0}
            box["removal_areas"] = []
            box["esqueleto_modificado"] = False
            
        self.actualizar_estado_flujo(2)


        self.mostrar_notificacion("Corregir Filtrado", "Se han eliminado las fases anteriores. Puedes ajustar los filtros nuevamente.", "info")


    def obtener_metricas(self):
        if not self.ruta_imagen_actual or not self.visor_imagen.boxes:
            self.mostrar_notificacion("Advertencia", "No hay datos para extraer métricas.", "warning")
            return
            
        base_name = Path(self.ruta_imagen_actual).stem
        esqueletos_dir = os.path.join(os.getcwd(), "analisis_resultados", base_name, "esqueletos")
        
        if not os.path.exists(esqueletos_dir):
            self.mostrar_notificacion("Advertencia", "No se encontraron esqueletos generados.", "warning")
            return
            
        from ia.extract_microglia_metrics import extract_microglia_metrics
        from PyQt6.QtWidgets import QApplication
        
        dialogo = DialogoCarga("Extrayendo métricas morfológicas...\nPor favor, espera.", self)
        dialogo.show()
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        QApplication.processEvents()
        
        metricas_imagen = []
        for box in self.visor_imagen.boxes:
            nombre = os.path.basename(box["crop_path"])
            out_path = os.path.join(esqueletos_dir, nombre)
            if os.path.exists(out_path):
                try:
                    met = extract_microglia_metrics(out_path)
                    metricas_imagen.append(met)
                except Exception as e:
                    logging.error(f"Error extrayendo métricas de {nombre}: {e}")
                    
        dialogo.close()
        QApplication.restoreOverrideCursor()
        
        if not metricas_imagen:
            self.mostrar_notificacion("Error", "No se pudieron extraer métricas de ninguna microglía.", "error")
            return
            
        self.metricas_reporte.append({
            "campo": self.metadatos_imagen.get("campo", ""),
            "tiempo": self.metadatos_imagen.get("tiempo", ""),
            "metricas": metricas_imagen
        })
        
        # Limpiar marcas de esqueleto modificado para quitar el resaltado amarillo
        for box in self.visor_imagen.boxes:
            box["esqueleto_modificado"] = False
        self.visor_imagen.draw_current_state()
        
        self.metricas_extraidas_ciclo_actual = True
        self.actualizar_estado_flujo(5)
        self.mostrar_notificacion("5. Métricas", "Métricas extraídas y análisis completado exitosamente.", "info")

    def agregar_imagen_reporte(self):
        self.visor_imagen.set_image_and_boxes(None, [])
        self.ruta_imagen_actual = None
        self.id_analisis_actual = None
        self.pixmaps_globales = {"Original": None, "Filtrada": None, "Esqueleto": None}
        self.combo_vista.blockSignals(True)
        self.combo_vista.clear()
        self.combo_vista.blockSignals(False)
        
        self.metricas_extraidas_ciclo_actual = False
        self.actualizar_estado_flujo(0)
        self.mostrar_notificacion("Info", "Sesión lista para cargar otra imagen y agregar al reporte.", "info")

    def descargar_reporte(self):
        if not self.metricas_reporte:
            self.mostrar_notificacion("Advertencia", "No hay métricas acumuladas para descargar.", "warning")
            return
            
        # Depuración: imprimir estructura de datos
        print(f"DEBUG: metricas_reporte type: {type(self.metricas_reporte)}")
        if isinstance(self.metricas_reporte, list) and len(self.metricas_reporte) > 0:
            print(f"DEBUG: First item type: {type(self.metricas_reporte[0])}")
            
        from datetime import datetime; from PyQt6.QtWidgets import QFileDialog; from pathlib import Path
        fecha_str = datetime.now().strftime("%Y%m%d_%H%M"); default_name = f"Reporte_{fecha_str}.xlsx"
        
        filepath, filter_selected = QFileDialog.getSaveFileName(self, "Guardar Reporte", default_name, "Excel Files (*.xlsx);;PDF Files (*.pdf);;Both Formats (*.xlsx *.pdf)")
        if not filepath: return
            
        try:
            reporte_por_tiempo = {}
            for img_data in self.metricas_reporte:
                t = str(img_data.get("tiempo", "X HORA")).upper()
                if t not in reporte_por_tiempo: reporte_por_tiempo[t] = []
                reporte_por_tiempo[t].append(img_data)

            columnas_labels = ["No.", "Lines", "Junction Points", "End Points", "Junction Voxels", "Slab Voxels", "Avg. Branch Length", "Triple points", "Quadruple points", "Max Branch Length", "Longest Shortest path"]
            metric_keys = ["lines", "junction points", "end points", "junction voxels", "slab voxels", "average branch length", "triple points", "quadruple points", "maximum branch length", "longest shortest path"]

            save_xlsx = "Excel" in filter_selected or "Both" in filter_selected or filepath.endswith(".xlsx")
            save_pdf = "PDF" in filter_selected or "Both" in filter_selected or filepath.endswith(".pdf")

            if save_xlsx:
                xlsx_path = filepath if filepath.endswith(".xlsx") else str(Path(filepath).with_suffix(".xlsx"))
                import openpyxl; from openpyxl.styles import PatternFill, Font, Alignment; from openpyxl.utils import get_column_letter
                wb = openpyxl.Workbook(); wb.remove(wb.active)
                
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                header_bg_fill = PatternFill(start_color="3A61A0", end_color="3A61A0", fill_type="solid")
                light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                bold_font = Font(bold=True); header_black_font = Font(color="000000", bold=True); center_alignment = Alignment(horizontal="center", vertical="center")
                anchos_fijos = [9.3, 18.0, 22.6, 18.6, 22.6, 16.6, 26.6, 20.0, 26.6, 26.6, 32.0]
                
                for tiempo, lista_campos in reporte_por_tiempo.items():
                    ws = wb.create_sheet(title=tiempo[:31])
                    for col_idx, width in enumerate(anchos_fijos, start=1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = width
                    
                    row_idx = 1
                    for img_data in lista_campos:
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(columnas_labels))
                        for c in range(1, len(columnas_labels) + 1):
                            ws.cell(row=row_idx, column=c).fill = yellow_fill
                        
                        campo_val = img_data.get("campo", "")
                        if isinstance(campo_val, dict): campo_val = str(campo_val)
                        cell_title = ws.cell(row=row_idx, column=1, value=str(campo_val))
                        cell_title.font = bold_font; cell_title.alignment = center_alignment; row_idx += 1
                        
                        for col_idx, label in enumerate(columnas_labels, start=1):
                            cell_h = ws.cell(row=row_idx, column=col_idx, value=label)
                            cell_h.font = header_black_font; cell_h.fill = header_bg_fill; cell_h.alignment = center_alignment
                        row_idx += 1
                        
                        for i, met in enumerate(img_data["metricas"], start=1):
                            cell_num = ws.cell(row=row_idx, column=1, value=i); cell_num.alignment = center_alignment
                            for col_idx, key in enumerate(metric_keys, start=2):
                                val = met.get(key, "")
                                if isinstance(val, (dict, list)): val = str(val)
                                cell_m = ws.cell(row=row_idx, column=col_idx, value=val); cell_m.alignment = center_alignment
                            if i % 2 != 0:
                                for c in range(1, len(columnas_labels) + 1):
                                    ws.cell(row=row_idx, column=c).fill = light_gray_fill
                            row_idx += 1
                        row_idx += 1
                wb.save(xlsx_path)

            if save_pdf:
                pdf_path = filepath if filepath.endswith(".pdf") else str(Path(filepath).with_suffix(".pdf"))
                try:
                    from fpdf import FPDF
                    class PDFReport(FPDF):
                        def header(self):
                            self.set_font("Arial", "B", 14)
                            self.cell(0, 10, "Reporte de Métricas Morfológicas - Microglías", 0, 1, "C")
                            self.ln(5)

                    pdf = PDFReport(orientation="L", unit="mm", format="A4")
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf_widths = [12, 22, 26, 22, 26, 20, 31, 23, 31, 31, 36] 

                    for tiempo, lista_campos in reporte_por_tiempo.items():
                        pdf.add_page()
                        pdf.set_font("Arial", "B", 12)
                        pdf.cell(0, 10, f"TIEMPO: {tiempo}", 0, 1, "L")
                        
                        for img_data in lista_campos:
                            pdf.set_fill_color(255, 255, 0); pdf.set_font("Arial", "B", 10)
                            pdf.cell(sum(pdf_widths), 8, f"Campo: {img_data['campo']}", 1, 1, "C", True)
                            
                            pdf.set_fill_color(58, 97, 160); pdf.set_text_color(0, 0, 0); pdf.set_font("Arial", "B", 7)
                            for i, label in enumerate(columnas_labels):
                                pdf.cell(pdf_widths[i], 8, label, 1, 0, "C", True)
                            pdf.ln()
                            
                            pdf.set_font("Arial", "", 9); pdf.set_text_color(0, 0, 0)
                            for idx, met in enumerate(img_data["metricas"], start=1):
                                if idx % 2 != 0: pdf.set_fill_color(242, 242, 242)
                                else: pdf.set_fill_color(255, 255, 255)
                                
                                pdf.cell(pdf_widths[0], 7, str(idx), 1, 0, "C", True)
                                values = []
                                for k in metric_keys:
                                    v = met.get(k, "")
                                    values.append(str(v) if not isinstance(v, (int, float)) else v)
                                    
                                for i_v, val in enumerate(values):
                                    pdf.cell(pdf_widths[i_v+1], 7, str(val), 1, 0, "C", True)
                                pdf.ln()
                            pdf.ln(5)
                    pdf.output(pdf_path)
                except Exception as e:
                    import logging
                    logging.error(f"Error generando PDF: {e}")

            self.mostrar_notificacion("Éxito", f"Reporte guardado en: {os.path.basename(filepath)}", "info")
        except Exception as error:
            self.mostrar_notificacion("Error", f"Falló la exportación: {str(error)}", "error")

    def finalizar_reporte(self):
        from vistas.utilidades import DialogoConfirmacion
        msg = "¿Estás seguro de finalizar el reporte actual? Se limpiarán todas las métricas acumuladas."
        if not DialogoConfirmacion("Finalizar Reporte", msg).exec(): return
        self.metricas_reporte.clear(); self.metricas_extraidas_ciclo_actual = False
        self.visor_imagen.set_image_and_boxes(None, []); self.ruta_imagen_actual = None; self.id_analisis_actual = None
        self.pixmaps_globales = {"Original": None, "Filtrada": None, "Esqueleto": None}
        self.combo_vista.blockSignals(True); self.combo_vista.clear(); self.combo_vista.blockSignals(False); self.actualizar_estado_flujo(0)
        self.mostrar_notificacion("Reporte Finalizado", "Sistema reiniciado.", "info")
