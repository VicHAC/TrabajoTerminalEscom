import logging
import os
import subprocess
import uuid
from pathlib import Path
from datetime import datetime
import json

from ia.constants import MIN_MICROGLIA_SIZE

from bd.database import conectar
from procesamiento.validacion_reporte import (
    ValidacionReporteMixin,
    construir_boton_validar,
    construir_boton_descargar,
    debe_bloquear_carga,
)

from PyQt6.QtCore import pyqtSignal, QRect, Qt, QSize, QEvent, QPoint
from PyQt6.QtGui import QColor, QImage, QPainter, QPen, QPixmap, QIcon, QIntValidator, QFont
from PyQt6.QtWidgets import (
    QApplication,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QPushButton,
    QVBoxLayout,
    QWidget,
    QSlider,
    QCheckBox,
    QLineEdit,
    QGridLayout,
    QMessageBox,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QAbstractItemView,
    QToolTip,
    QTreeWidget,
    QTreeWidgetItem,
    QScrollArea,
    QStackedWidget,
)
from PyQt6.QtCore import QTimer

# os.environ["QT_QPA_PLATFORM"] = "xcb"


# MorphologyAnalyzer will be imported inside methods that use it

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
# Silence chatty libraries
logging.getLogger("PIL").setLevel(logging.WARNING)
logging.getLogger("fpdf").setLevel(logging.WARNING)

# =========================================================
# BOTÓN PERSONALIZADO PARA TOOLTIPS SEGUROS (ANTI-CORTES)
# =========================================================
class SafeToolTipButton(QPushButton):
    """
    Un botón que fuerza al ToolTip a mostrarse desplazado hacia la izquierda.
    Garantiza que el texto nunca se corte en el borde derecho de la pantalla (WSL/Pantalla Completa).
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.custom_tooltip = ""

    def setCustomToolTip(self, text):
        self.custom_tooltip = text
        self.setToolTip(text)

    def event(self, e):
        if e.type() == QEvent.Type.ToolTip:
            if self.custom_tooltip:
                global_pos = self.mapToGlobal(QPoint(0, self.height() + 2))
                global_pos.setX(global_pos.x() - 80)
                QToolTip.showText(global_pos, self.custom_tooltip, self)
            return True
        return super().event(e)


# =========================================================
# ZONA DE ARRASTRAR Y SOLTAR (DRAG & DROP)
# =========================================================
class DropZone(QFrame):
    file_selected = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        
        self.setObjectName("DropZoneObj")
        
        self.layout = QVBoxLayout(self)
        self.layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.layout.setContentsMargins(10, 10, 10, 15) 
        
        self.lbl_icono = QLabel()
        self.lbl_icono.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_icono.setStyleSheet("border: none; background: transparent;")
        self.lbl_icono.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        
        self.lbl_texto = QLabel("Haz clic para cargar una imagen o arrástrala aquí")
        self.lbl_texto.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_texto.setWordWrap(True)
        self.lbl_texto.setStyleSheet("color: #888888; font-size: 12px; font-weight: normal; border: none; background: transparent;")
        self.lbl_texto.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        
        self.layout.addWidget(self.lbl_icono, stretch=1)
        self.layout.addWidget(self.lbl_texto)
        
        self.pixmap_original = None
        self.is_preview = False
        
        self.mostrar_placeholder()

    def mostrar_placeholder(self):
        self.is_preview = False
        self.pixmap_original = QPixmap("assets/cargar.png")
        
        self.lbl_texto.setText("Haz clic para cargar una imagen o arrástrala aquí")
        self.lbl_texto.setStyleSheet("color: #888888; font-size: 12px; font-weight: normal; border: none; background: transparent;")
        
        self.setStyleSheet("""
            #DropZoneObj {
                border: 2px dashed #0969da;
                border-radius: 8px;
                background-color: #ffffff;
            }
            #DropZoneObj:hover {
                background-color: #f6f8fa;
                border: 2px dashed #0550ae;
            }
        """)
        self.actualizar_imagen()

    def mostrar_imagen(self, pixmap, nombre_archivo):
        self.is_preview = True
        self.pixmap_original = pixmap
        
        self.lbl_texto.setText(f"Archivo seleccionado: <b>{nombre_archivo}</b>")
        self.lbl_texto.setStyleSheet("color: #333333; font-size: 12px; border: none; background: transparent;")
        
        self.setStyleSheet("""
            #DropZoneObj {
                border: 2px solid #28a745;
                border-radius: 8px;
                background-color: #ffffff;
            }
        """)
        self.actualizar_imagen()

    def actualizar_imagen(self):
        if self.pixmap_original and not self.pixmap_original.isNull():
            w = self.width() - 20
            h = self.height() - 60 
            
            if w > 0 and h > 0:
                self.lbl_icono.setPixmap(self.pixmap_original.scaled(
                    w, h, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation
                ))

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.actualizar_imagen()

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls(): event.accept()
        else: event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls(): event.accept()
        else: event.ignore()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            event.setDropAction(Qt.DropAction.CopyAction)
            event.accept()
            urls = event.mimeData().urls()
            if urls and urls[0].isLocalFile():
                file_path = urls[0].toLocalFile()
                ext = Path(file_path).suffix.lower()
                if ext in [".tif", ".tiff", ".png", ".jpg", ".jpeg"]:
                    self.file_selected.emit(file_path)
        else:
            event.ignore()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            from PyQt6.QtCore import QStandardPaths
            import os
            docs_dir = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.DocumentsLocation)
            default_dir = docs_dir if docs_dir and os.path.exists(docs_dir) else os.path.expanduser("~")
            ruta_archivo, _ = QFileDialog.getOpenFileName(
                self, "Seleccionar imagen", default_dir, "Imágenes TIFF (*.tiff *.tif);;Todas las imágenes (*.png *.jpg *.jpeg)",
                options=QFileDialog.Option.DontUseNativeDialog
            )
            if ruta_archivo:
                self.file_selected.emit(ruta_archivo)


# =========================================================
# DIÁLOGO INTERMEDIO PARA CARGA DE DATOS
# =========================================================
class DialogoCargarImagen(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Detalles del Campo")
        self.setFixedSize(580, 520) 
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True)
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        
        self.ruta_seleccionada = None
        self.campo_val = ""
        self.tiempo_val = ""

        main_layout = QVBoxLayout(self)
        frame = QFrame(self)
        frame.setStyleSheet("QFrame { background-color: #FFFFFF; border-radius: 12px; border: 1px solid #d0d7de; } QLabel { border: none; }")
        layout = QVBoxLayout(frame)
        
        header_layout = QHBoxLayout()
        lbl_titulo = QLabel("Registro del Campo")
        lbl_titulo.setStyleSheet("font-size: 18px; font-weight: bold; color: #0969da; border: none;")
        lbl_titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        header_layout.setContentsMargins(10, 10, 10, 0)
        self.btn_cerrar_x = QPushButton()
        self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))
        self.btn_cerrar_x.setIconSize(QSize(20, 20))
        self.btn_cerrar_x.setFixedSize(35, 35)
        self.btn_cerrar_x.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cerrar_x.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #f6f8fa; border-radius: 17px; }
        """)
        self.btn_cerrar_x.enterEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xroja.png"))
        self.btn_cerrar_x.leaveEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))
        self.btn_cerrar_x.clicked.connect(self.reject)
        
        header_layout.addStretch()
        header_layout.addWidget(lbl_titulo)
        header_layout.addStretch()
        header_layout.addWidget(self.btn_cerrar_x)
        layout.addLayout(header_layout)
        layout.addSpacing(10)

        layout_campos = QHBoxLayout()
        
        layout_campo = QVBoxLayout()
        lbl_campo = QLabel("Campo:")
        lbl_campo.setStyleSheet("font-weight: bold; font-size: 13px;")
        self.input_campo = QLineEdit()
        self.input_campo.setPlaceholderText("Ej. Campo A, 1...")
        layout_campo.addWidget(lbl_campo)
        layout_campo.addWidget(self.input_campo)

        layout_tiempo = QVBoxLayout()
        lbl_tiempo = QLabel("Tiempo:")
        lbl_tiempo.setStyleSheet("font-weight: bold; font-size: 12px; color: #57606a;")
        self.input_tiempo = QLineEdit()
        self.input_tiempo.setPlaceholderText("Ej. 1hr, 2hrs...")
        layout_tiempo.addWidget(lbl_tiempo)
        layout_tiempo.addWidget(self.input_tiempo)

        layout_campos.addLayout(layout_campo)
        layout_campos.addSpacing(15)
        layout_campos.addLayout(layout_tiempo)
        
        layout.addLayout(layout_campos)
        layout.addSpacing(15)

        self.drop_zone = DropZone(self)
        self.drop_zone.file_selected.connect(self.procesar_archivo)
        layout.addWidget(self.drop_zone, stretch=1)
        layout.addSpacing(10)

        layout_botones = QHBoxLayout()
        
        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_cancelar.clicked.connect(self.reject)
        
        btn_continuar = QPushButton("Cargar Imagen")
        btn_continuar.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_continuar.setStyleSheet("""
            QPushButton { 
                background-color: transparent; 
                border: 2px solid #2da44e; 
                color: #2da44e; 
                border-radius: 6px; 
                padding: 8px 20px; 
                font-weight: bold; 
                font-size: 13px; 
            }
            QPushButton:hover { 
                background-color: #2da44e; 
                color: white; 
            }
        """)
        btn_continuar.clicked.connect(self.validar_y_continuar)
        
        layout_botones.addStretch()
        layout_botones.addWidget(btn_continuar)
        layout_botones.addStretch()
        
        layout.addLayout(layout_botones)
        main_layout.addWidget(frame)

    def showEvent(self, event):
        super().showEvent(event)
        if self.parent():
            p_geom = self.parent().geometry()
            self.move(p_geom.x() + (p_geom.width() - self.width()) // 2, p_geom.y() + (p_geom.height() - self.height()) // 2)
        else:
            screen = QApplication.primaryScreen().geometry()
            self.move((screen.width() - self.width()) // 2, (screen.height() - self.height()) // 2)

    def mousePressEvent(self, event):
        from PyQt6.QtCore import Qt
        if event.button() == Qt.MouseButton.LeftButton:
            if self.windowHandle():
                self.windowHandle().startSystemMove()
            else:
                self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            event.accept()
        else:
            super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        from PyQt6.QtCore import Qt
        if not self.windowHandle() and hasattr(self, "_drag_pos") and event.buttons() == Qt.MouseButton.LeftButton:
            self.move(event.globalPosition().toPoint() - self._drag_pos)
            event.accept()
        else:
            super().mouseMoveEvent(event)

    def procesar_archivo(self, ruta):
        self.ruta_seleccionada = ruta
        nombre_archivo = os.path.basename(ruta)
        
        try:
            import cv2
            import numpy as np
            try:
                with open(ruta, "rb") as f:
                    file_bytes = np.frombuffer(f.read(), dtype=np.uint8)
                cv_img = cv2.imdecode(file_bytes, cv2.IMREAD_UNCHANGED)
            except Exception:
                cv_img = None
            if cv_img is not None:
                if cv_img.dtype == np.uint16:
                    cv_img = ((cv_img - cv_img.min()) / (cv_img.max() - cv_img.min()) * 255).astype(np.uint8)
                if len(cv_img.shape) == 2:
                    h, w = cv_img.shape
                    qimg = QImage(cv_img.data, w, h, w, QImage.Format.Format_Grayscale8)
                else:
                    cv_img = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
                    h, w, ch = cv_img.shape
                    qimg = QImage(cv_img.data, w, h, ch * w, QImage.Format.Format_RGB888)
                pixmap_preview = QPixmap.fromImage(qimg)
            else:
                pixmap_preview = QPixmap(ruta)
        except Exception as e:
            logging.error(f"Error al previsualizar: {e}")
            pixmap_preview = QPixmap(ruta)

        if not pixmap_preview.isNull():
            self.drop_zone.mostrar_imagen(pixmap_preview, nombre_archivo)

    def validar_y_continuar(self):
        if not self.input_campo.text().strip() or not self.input_tiempo.text().strip() or not self.ruta_seleccionada:
            from vistas.utilidades import DialogoNotificacion
            DialogoNotificacion("Error", "Campos incompletos", "warning", self).exec()
            return
        
        self.campo_val = self.input_campo.text().strip()
        self.tiempo_val = self.input_tiempo.text().strip()
        self.accept()


# =========================================================
# CLASES AUXILIARES Y VISOR INTERACTIVO
# =========================================================
class DialogoCarga(QDialog):
    def __init__(self, mensaje="Procesando...", parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True)
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        
        layout = QVBoxLayout(self)
        frame = QFrame(self)
        frame.setStyleSheet("""
            QFrame { background-color: #FFFFFF; border-radius: 12px; border: 2px solid #003366; }
            QLabel { color: #003366; font-size: 16px; font-weight: bold; padding: 20px; }
        """)
        flayout = QVBoxLayout(frame)
        label = QLabel(mensaje)
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        flayout.addWidget(label)
        layout.addWidget(frame)
        self.setLayout(layout)

    def showEvent(self, event):
        super().showEvent(event)
        if self.parent():
            p_geom = self.parent().geometry()
            self.move(p_geom.x() + (p_geom.width() - self.width()) // 2, p_geom.y() + (p_geom.height() - self.height()) // 2)
        else:
            screen = QApplication.primaryScreen().geometry()
            self.move((screen.width() - self.width()) // 2, (screen.height() - self.height()) // 2)


class DialogoComparativo(QDialog):
    """
    Muestra las 3 fases del proceso lado a lado para una microglía,
    o una versión sobrepuesta (esqueleto sobre original).
    """
    def __init__(self, fases, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Proceso de la Microglía")
        self.setFixedSize(950, 500) 
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        
        self.fases = fases
        
        main_layout = QVBoxLayout(self)
        self.frame = QFrame(self)
        self.frame.setStyleSheet("QFrame { background-color: #FFFFFF; border-radius: 12px; border: 2px solid #003366; } QLabel { border: none; }")
        self.layout_principal = QVBoxLayout(self.frame)
        
        header_layout = QHBoxLayout()
        self.lbl_titulo = QLabel("<b>Comparativa del Proceso</b>")
        self.lbl_titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_titulo.setStyleSheet("font-size: 18px; color: #003366; border: none;")
        
        header_layout.setContentsMargins(10, 10, 10, 0)
        self.btn_cerrar_x = QPushButton()
        self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))
        self.btn_cerrar_x.setIconSize(QSize(20, 20))
        self.btn_cerrar_x.setFixedSize(35, 35)
        self.btn_cerrar_x.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cerrar_x.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #f6f8fa; border-radius: 17px; }
        """)
        self.btn_cerrar_x.enterEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xroja.png"))
        self.btn_cerrar_x.leaveEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))

        self.btn_cerrar_x.clicked.connect(self.accept)
        
        header_layout.addStretch()
        header_layout.addWidget(self.lbl_titulo)
        header_layout.addStretch()
        header_layout.addWidget(self.btn_cerrar_x)
        self.layout_principal.addLayout(header_layout)
        self.layout_principal.addSpacing(10)

        # Widget para vista lado a lado
        self.widget_lado_lado = QWidget()
        self.layout_lado_lado = QHBoxLayout(self.widget_lado_lado)
        for fase in self.fases:
            v_layout = QVBoxLayout()
            lbl_nombre_fase = QLabel(fase["nombre"])
            lbl_nombre_fase.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl_nombre_fase.setStyleSheet("font-weight: bold; color: #555; margin-bottom: 5px;")
            
            lbl_img = QLabel()
            lbl_img.setAlignment(Qt.AlignmentFlag.AlignCenter)
            pixmap = fase["pixmap"] if fase["pixmap"] else QPixmap(fase["path"])
            if pixmap and not pixmap.isNull():
                lbl_img.setPixmap(pixmap.scaled(280, 280, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
            else:
                lbl_img.setText("N/D")
            v_layout.addWidget(lbl_nombre_fase)
            v_layout.addWidget(lbl_img)
            self.layout_lado_lado.addLayout(v_layout)
        
        self.layout_principal.addWidget(self.widget_lado_lado)
        
        self.layout_principal.addSpacing(15)
        
        main_layout.addWidget(self.frame)
        self.setLayout(main_layout)

    def ajustar_posicion(self):
        if self.parent():
            p_geom = self.parent().geometry()
            self.move(p_geom.x() + (p_geom.width() - self.width()) // 2, p_geom.y() + (p_geom.height() - self.height()) // 2)

    def showEvent(self, event):
        super().showEvent(event)
        self.ajustar_posicion()

    def mousePressEvent(self, event):
        from PyQt6.QtCore import Qt
        if event.button() == Qt.MouseButton.LeftButton:
            if self.windowHandle():
                self.windowHandle().startSystemMove()
            else:
                self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            event.accept()
        else:
            super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        from PyQt6.QtCore import Qt
        if not self.windowHandle() and hasattr(self, "_drag_pos") and event.buttons() == Qt.MouseButton.LeftButton:
            self.move(event.globalPosition().toPoint() - self._drag_pos)
            event.accept()
        else:
            super().mouseMoveEvent(event)


class InteractiveLabelDetail(QLabel):
    def __init__(self, parent_dialog):
        super().__init__()
        self.parent_dialog = parent_dialog
        
        # Crear cursor de lápiz
        from PyQt6.QtGui import QPixmap, QPainter, QPen, QCursor
        from PyQt6.QtCore import Qt
        pixmap = QPixmap(16, 16)
        pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(pixmap)
        painter.setPen(QPen(Qt.GlobalColor.white, 1.5))
        # Dibujar un lapiz simple blanco
        painter.drawLine(0, 15, 5, 10)
        painter.drawLine(5, 10, 15, 0)
        painter.drawLine(15, 0, 16, 1)
        painter.drawLine(16, 1, 6, 11)
        painter.drawLine(6, 11, 0, 15)
        painter.end()
        self.pencil_cursor = QCursor(pixmap, 0, 15)
        self.is_drawing = False
        self.start_pos = None
        self.current_pos = None
        self.mode = "pointer" 
        self.setMouseTracking(True)
        self.pencil_path = [] # Lista de puntos para unir ramas
        self.is_dragging_point = False
        self.dragging_point = None

    def get_point_in_label_coords(self, ox, oy):
        from PyQt6.QtCore import QPoint
        pix = self.pixmap()
        if not pix or pix.isNull(): return None
        lbl_w, lbl_h = self.width(), self.height()
        pix_w, pix_h = pix.width(), pix.height()
        dx = (lbl_w - pix_w) // 2; dy = (lbl_h - pix_h) // 2
        orig_w, orig_h = self.parent_dialog.get_original_crop_size()
        if orig_w == 0 or orig_h == 0: return None
        scale_x = pix_w / orig_w; scale_y = pix_h / orig_h
        return QPoint(int(ox * scale_x) + dx, int(oy * scale_y) + dy)
        
    def point_to_segment_dist_squared(self, p, a, b):
        x0, y0 = p.x(), p.y()
        x1, y1 = a.x(), a.y()
        x2, y2 = b.x(), b.y()
        dx = x2 - x1
        dy = y2 - y1
        if dx == 0 and dy == 0:
            return (x0 - x1)**2 + (y0 - y1)**2
        t = ((x0 - x1) * dx + (y0 - y1) * dy) / float(dx*dx + dy*dy)
        t = max(0.0, min(1.0, t))
        px = x1 + t * dx
        py = y1 + t * dy
        return (x0 - px)**2 + (y0 - py)**2

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            event.accept()
            if self.mode == "eraser":
                # Primero verificar si el clic es para eliminar un área existente
                coords = self.get_rect_in_original_coords(event.pos(), event.pos())
                if coords:
                    ox, oy = coords["x"], coords["y"]
                    areas = self.parent_dialog.box.get("removal_areas", [])
                    for i, area in enumerate(areas):
                        if area["x"] <= ox <= area["x"]+area["w"] and area["y"] <= oy <= area["y"]+area["h"]:
                            areas.pop(i)
                            self.parent_dialog.actualizar_visibilidad_boton_limpieza()
                            self.update()
                            return
                
                self.is_drawing = True
                self.start_pos = event.pos()
                self.current_pos = event.pos()
                self.update()
            elif self.mode == "pencil_line":
                click_pos = event.pos()
                connections = self.parent_dialog.box.get("manual_connections", [])
                
                for p_idx, path in enumerate(connections):
                    for pt_idx, pt in enumerate(path):
                        lbl_pt = self.get_point_in_label_coords(pt[0], pt[1])
                        if lbl_pt:
                            dist_sq = (click_pos.x() - lbl_pt.x())**2 + (click_pos.y() - lbl_pt.y())**2
                            if dist_sq <= 36:
                                self.is_dragging_point = True
                                self.dragging_point = (p_idx, pt_idx)
                                return
                                
                for p_idx, path in enumerate(connections):
                    for pt_idx in range(len(path) - 1):
                        pA = self.get_point_in_label_coords(path[pt_idx][0], path[pt_idx][1])
                        pB = self.get_point_in_label_coords(path[pt_idx+1][0], path[pt_idx+1][1])
                        if pA and pB:
                            dist_sq = self.point_to_segment_dist_squared(click_pos, pA, pB)
                            if dist_sq <= 25:
                                c = self.get_rect_in_original_coords(click_pos, click_pos)
                                if c:
                                    path.insert(pt_idx + 1, (c["x"], c["y"]))
                                    self.is_dragging_point = True
                                    self.dragging_point = (p_idx, pt_idx + 1)
                                    self.update()
                                    return
                                    
                if not getattr(self, "is_waiting_second_click", False):
                    self.is_waiting_second_click = True
                    self.line_start_point = event.pos()
                    self.current_pos = event.pos()
                else:
                    self.is_waiting_second_click = False
                    self.current_pos = event.pos()
                    p1 = self.get_rect_in_original_coords(self.line_start_point, self.line_start_point)
                    p2 = self.get_rect_in_original_coords(self.current_pos, self.current_pos)
                    if p1 and p2:
                        self.parent_dialog._push_undo_state()
                        orig_points = [(p1["x"], p1["y"]), (p2["x"], p2["y"])]
                        if "manual_connections" not in self.parent_dialog.box: self.parent_dialog.box["manual_connections"] = []
                        self.parent_dialog.box["manual_connections"].append(orig_points)
                        self.parent_dialog.actualizar_visibilidad_boton_limpieza()
                    self.line_start_point = None
                    self.current_pos = None
                self.update()
                
            elif self.mode == "pencil_freehand":
                self.is_drawing_freehand = True
                self.pencil_path = [event.pos()]
                self.update()
                
            elif self.mode == "eraser_union":
                self.is_drawing_eraser = True
                self.current_pos = event.pos()
                self.parent_dialog._push_undo_state()
                self.erase_at_position(event.pos())
                self.update()


    def mouseMoveEvent(self, event):
        event.accept()
        if getattr(self, "is_dragging_point", False) and getattr(self, "dragging_point", None):
            c = self.get_rect_in_original_coords(event.pos(), event.pos())
            if c:
                p_idx, pt_idx = self.dragging_point
                self.parent_dialog.box["manual_connections"][p_idx][pt_idx] = (c["x"], c["y"])
                self.update()
        elif self.mode == "pencil_line" and getattr(self, "is_waiting_second_click", False):
            self.current_pos = event.pos()
            self.update()
        elif getattr(self, "is_drawing_freehand", False):
            self.pencil_path.append(event.pos())
            self.update()
        elif getattr(self, "is_drawing_eraser", False) or self.is_drawing:
            self.current_pos = event.pos()
            if self.mode == "eraser_union" and getattr(self, "is_drawing_eraser", False):
                self.erase_at_position(event.pos())
            self.update()
        elif self.mode == "eraser_union":
            self.current_pos = event.pos()
            self.update()

    def leaveEvent(self, event):
        self.current_pos = None
        self.update()

    def erase_at_position(self, pos):
        """Borra píxeles en la imagen de esqueleto EN MEMORIA (no en disco)."""
        import numpy as np
        skeleton_img = getattr(self.parent_dialog, 'skeleton_working', None)
        if skeleton_img is None:
            return
        from PyQt6.QtCore import QPoint
        # Map the 16x16 label square to original crop coordinates
        p1 = QPoint(pos.x() - 8, pos.y() - 8)
        p2 = QPoint(pos.x() + 8, pos.y() + 8)
        c = self.get_rect_in_original_coords(p1, p2)
        if c:
            x, y, w, h = c["x"], c["y"], c["w"], c["h"]
            # Poner a 0 (negro) los píxeles dentro del cuadro de la goma
            self.parent_dialog.skeleton_working[y:y+h, x:x+w] = 0
            self.parent_dialog.skeleton_has_changes = True
            
            # Filtrar puntos de manual_connections que estén dentro del cuadro de la goma
            if "manual_connections" in self.parent_dialog.box:
                new_connections = []
                for path in self.parent_dialog.box["manual_connections"]:
                    new_path = []
                    for pt in path:
                        if not (x <= pt[0] <= x + w and y <= pt[1] <= y + h):
                            new_path.append(pt)
                    if len(new_path) > 1:
                        new_connections.append(new_path)
                self.parent_dialog.box["manual_connections"] = new_connections
            
            # Actualizar el pixmap en memoria y refrescar vista
            self.parent_dialog._actualizar_pixmap_esqueleto_memoria()
            self.parent_dialog.actualizar_vista()
            self.parent_dialog.actualizar_visibilidad_boton_limpieza()

    def mouseReleaseEvent(self, event):
        event.accept()
        if self.is_drawing:
            self.is_drawing = False
            if self.mode == "eraser":
                area = self.get_rect_in_original_coords(self.start_pos, self.current_pos)
                if area:
                    if "removal_areas" not in self.parent_dialog.box: self.parent_dialog.box["removal_areas"] = []
                    self.parent_dialog.box["removal_areas"].append(area)
                    self.parent_dialog.actualizar_visibilidad_boton_limpieza()
                    self.parent_dialog.actualizar_offsets()
        
        if self.mode == "pencil_line":
            if getattr(self, "is_dragging_point", False):
                self.is_dragging_point = False
                self.dragging_point = None
        elif self.mode == "pencil_freehand":
            if getattr(self, "is_drawing_freehand", False):
                self.is_drawing_freehand = False
                orig_points = []
                for p in getattr(self, "pencil_path", []):
                    c = self.get_rect_in_original_coords(p, p)
                    if c: orig_points.append((c["x"], c["y"]))
                if len(orig_points) > 1:
                    self.parent_dialog._push_undo_state()
                    if "manual_connections" not in self.parent_dialog.box: self.parent_dialog.box["manual_connections"] = []
                    self.parent_dialog.box["manual_connections"].append(orig_points)
                    self.parent_dialog.actualizar_visibilidad_boton_limpieza()

        elif self.mode == "eraser_union":
            if getattr(self, "is_drawing_eraser", False):
                self.is_drawing_eraser = False
                    
        self.start_pos = None; self.current_pos = None
        self.pencil_path = []
        self.update()

    def get_rect_in_original_coords(self, p1, p2):
        pix = self.pixmap()
        if not pix or pix.isNull(): return None
        lbl_w, lbl_h = self.width(), self.height()
        pix_w, pix_h = pix.width(), pix.height()
        dx = (lbl_w - pix_w) // 2; dy = (lbl_h - pix_h) // 2
        x1 = p1.x() - dx; y1 = p1.y() - dy
        x2 = p2.x() - dx; y2 = p2.y() - dy
        orig_w, orig_h = self.parent_dialog.get_original_crop_size()
        if orig_w == 0 or orig_h == 0: return None
        scale_x = orig_w / pix_w; scale_y = orig_h / pix_h
        rx = min(x1, x2); ry = min(y1, y2); rw = abs(x2 - x1); rh = abs(y2 - y1)
        return {"x": int(rx * scale_x), "y": int(ry * scale_y), "w": int(rw * scale_x), "h": int(rh * scale_y)}

    def paintEvent(self, event):
        super().paintEvent(event)
        pix = self.pixmap()
        if not pix or pix.isNull(): return
        painter = QPainter(self)
        lbl_w, lbl_h = self.width(), self.height()
        pix_w, pix_h = pix.width(), pix.height()
        dx = (lbl_w - pix_w) // 2; dy = (lbl_h - pix_h) // 2
        orig_w, orig_h = self.parent_dialog.get_original_crop_size()
        if orig_w > 0:
            scale_x = pix_w / orig_w; scale_y = pix_h / orig_h
            painter.setPen(QPen(QColor(220, 53, 69), 2))
            for area in self.parent_dialog.box.get("removal_areas", []):
                if area.get("baked", False): continue
                rx = int(area["x"] * scale_x) + dx; ry = int(area["y"] * scale_y) + dy
                rw = int(area["w"] * scale_x); rh = int(area["h"] * scale_y)
                painter.drawRect(rx, ry, rw, rh)
                painter.drawLine(rx, ry, rx + rw, ry + rh); painter.drawLine(rx + rw, ry, rx, ry + rh)
        if self.mode == "eraser" and self.start_pos and self.current_pos:
            # print("Drawing eraser rect")
            painter.setPen(QPen(QColor(255, 0, 0, 100), 1, Qt.PenStyle.DashLine))
            painter.setBrush(QColor(255, 0, 0, 50))
            rx = min(self.start_pos.x(), self.current_pos.x())
            ry = min(self.start_pos.y(), self.current_pos.y())
            rw = abs(self.current_pos.x() - self.start_pos.x())
            rh = abs(self.current_pos.y() - self.start_pos.y())
            painter.drawRect(rx, ry, rw, rh)
            
        # Dibujar cuadrito azul de la goma pixel-por-pixel
        if self.mode == "eraser_union" and getattr(self, "current_pos", None):
            painter.setPen(QPen(QColor(0, 120, 255), 1.5))
            painter.setBrush(QColor(0, 120, 255, 80)) # Azul semi-transparente
            painter.drawRect(self.current_pos.x() - 8, self.current_pos.y() - 8, 16, 16)
        
        # Dibujar pincel libre
        if getattr(self, "is_drawing_freehand", False) and getattr(self, "pencil_path", None) and len(self.pencil_path) > 1:
            painter.setPen(QPen(QColor(255, 255, 0), 2, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap, Qt.PenJoinStyle.RoundJoin))
            for i in range(len(self.pencil_path) - 1):
                painter.drawLine(self.pencil_path[i], self.pencil_path[i+1])
                
        # Dibujar la línea recta del lápiz mientras se espera el segundo clic
        if self.mode == "pencil_line" and getattr(self, "is_waiting_second_click", False) and getattr(self, "line_start_point", None) and self.current_pos:
            painter.setPen(QPen(QColor(255, 255, 0), 2, Qt.PenStyle.DashLine, Qt.PenCapStyle.RoundCap, Qt.PenJoinStyle.RoundJoin))
            painter.drawLine(self.line_start_point, self.current_pos)

        # Dibujar conexiones ya guardadas (solo si estamos en modo esqueleto)
        if self.parent_dialog.fases_disponibles[self.parent_dialog.indice_fase]["nombre"] == "ESQUELETIZADO":
            connections = self.parent_dialog.box.get("manual_connections", [])
            if connections:
                painter.setPen(QPen(Qt.GlobalColor.white, 2))
                # Necesitamos mapear de original a label
                pix = self.pixmap()
                lbl_w, lbl_h = self.width(), self.height()
                pix_w, pix_h = pix.width(), pix.height()
                dx = (lbl_w - pix_w) // 2; dy = (lbl_h - pix_h) // 2
                orig_w, orig_h = self.parent_dialog.get_original_crop_size()
                if orig_w > 0 and orig_h > 0:
                    scale_x = pix_w / orig_w; scale_y = pix_h / orig_h
                    for path in connections:
                        for i in range(len(path) - 1):
                            p1 = QPoint(int(path[i][0] * scale_x) + dx, int(path[i][1] * scale_y) + dy)
                            p2 = QPoint(int(path[i+1][0] * scale_x) + dx, int(path[i+1][1] * scale_y) + dy)
                            painter.drawLine(p1, p2)
        painter.end()

class DialogoVistaCelular(QDialog):

    """
    Ventana detallada que permite navegar entre las fases de procesamiento
    de una microglía específica (Original < Filtrado > Esqueletizado).
    """
    def __init__(self, box, pixmap_mem=None, modo_inicial="Original", parent=None):
        super().__init__(parent)
        self.setWindowTitle("Vista Detallada de la Célula")
        self.resize(500, 750) 
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        
        self.box = box
        self.crop_path = box["crop_path"].replace("\\", "/")
        self.pixmap_mem_filtrado = pixmap_mem
        self.drag_position = None

        self.fases_disponibles = []
        self.preparar_fases()
        
        # Establecer índice inicial basado en el modo de la ventana principal
        self.indice_fase = 0
        mapping = {"Original": "ORIGINAL", "Filtrada": "FILTRADO", "Esqueleto": "ESQUELETIZADO", "Previsualización": "FILTRADO"}
        nombre_buscado = mapping.get(modo_inicial, "ORIGINAL")
        for i, f in enumerate(self.fases_disponibles):
            if f["nombre"] == nombre_buscado:
                self.indice_fase = i
                break
                
        # Asegurar que no empiece en una fase no completada
        paso_padre = getattr(self.parent(), "paso_actual", 0)
        
        # Determinar si la fase de FILTRADO está disponible
        filtrado_disponible = False
        if paso_padre >= 3:
            filtrado_disponible = True
        elif paso_padre == 2:
            if self.parent() and hasattr(self.parent(), "combo_vista"):
                if self.parent().combo_vista.currentText() == "Previsualización":
                    filtrado_disponible = True
            if self.parent() and hasattr(self.parent(), "frame_filtros"):
                if self.parent().frame_filtros.isVisible():
                    filtrado_disponible = True
                    
        esqueleto_disponible = (paso_padre >= 4)
        
        if self.indice_fase == 1 and not filtrado_disponible:
            self.indice_fase = 0
        elif self.indice_fase == 2 and not esqueleto_disponible:
            self.indice_fase = 1 if filtrado_disponible else 0
        
        main_layout = QVBoxLayout(self)
        frame = QFrame(self)
        frame.setStyleSheet("QFrame { background-color: #FFFFFF; border-radius: 12px; border: 2px solid #003366; } QLabel { border: none; }")
        
        def start_drag(event):
            from PyQt6.QtCore import Qt
            if event.button() == Qt.MouseButton.LeftButton:
                if self.window().windowHandle():
                    self.window().windowHandle().startSystemMove()
                event.accept()

        frame.mousePressEvent = start_drag
        layout = QVBoxLayout(frame)
        
        nombre_archivo = os.path.basename(self.crop_path)
        header_layout = QHBoxLayout()
        lbl_nombre = QLabel(f"Identificador: <b>{nombre_archivo}</b>")
        lbl_nombre.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_nombre.setStyleSheet("font-size: 15px; color: #003366; border: none;")
        lbl_nombre.mousePressEvent = start_drag
        
        header_layout.setContentsMargins(10, 10, 10, 0)
        self.btn_cerrar_x = QPushButton()
        self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))
        self.btn_cerrar_x.setIconSize(QSize(20, 20))
        self.btn_cerrar_x.setFixedSize(35, 35)
        self.btn_cerrar_x.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cerrar_x.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #f6f8fa; border-radius: 17px; }
        """)
        self.btn_cerrar_x.enterEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xroja.png"))
        self.btn_cerrar_x.leaveEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))

        self.btn_cerrar_x.clicked.connect(self.accept)
        
        header_layout.addStretch()
        header_layout.addWidget(lbl_nombre)
        header_layout.addStretch()
        header_layout.addWidget(self.btn_cerrar_x)
        layout.addLayout(header_layout)
        layout.addSpacing(10)

        # Nombre de la fase actual
        self.lbl_fase = QLabel("FASE: ORIGINAL")
        self.lbl_fase.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_fase.setStyleSheet("font-size: 13px; font-weight: bold; color: #555; margin-bottom: 5px;")
        self.lbl_fase.mousePressEvent = start_drag
        layout.addWidget(self.lbl_fase)

        # Contenedor de imagen con botones de navegación lateral
        layout_imagen_nav = QHBoxLayout()
        
        self.btn_ant = QPushButton()
        self.btn_ant.setIcon(QIcon("assets/buttons/izq.png"))
        self.btn_ant.setIconSize(QSize(12, 12))
        self.btn_ant.setFixedSize(22, 22)
        self.btn_ant.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_ant.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #f0f0f0; border-radius: 11px; }
            QPushButton:disabled { opacity: 0.3; }
        """)
        self.btn_ant.clicked.connect(self.mostrar_anterior)
        
        self.label_imagen = InteractiveLabelDetail(self)
        self.label_imagen.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label_imagen.setFixedSize(380, 380)

        
        self.btn_sig = QPushButton()
        self.btn_sig.setIcon(QIcon("assets/buttons/der.png"))
        self.btn_sig.setIconSize(QSize(12, 12))
        self.btn_sig.setFixedSize(22, 22)
        self.btn_sig.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_sig.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; }
            QPushButton:hover { background-color: #f0f0f0; border-radius: 11px; }
            QPushButton:disabled { opacity: 0.3; }
        """)
        self.btn_sig.clicked.connect(self.mostrar_siguiente)
        
        layout_imagen_nav.addWidget(self.btn_ant)
        layout_imagen_nav.addWidget(self.label_imagen, stretch=1)
        layout_imagen_nav.addWidget(self.btn_sig)
        
        layout.addLayout(layout_imagen_nav)
        
        # Herramientas de limpieza
        self.frame_tools_limpieza = QFrame()
        self.frame_tools_limpieza.setStyleSheet("QFrame { border: none; background: transparent; }")
        layout_tools_limpieza = QHBoxLayout(self.frame_tools_limpieza)
        layout_tools_limpieza.setContentsMargins(0,0,0,0)
        
        self.btn_tool_limpieza = QPushButton("Off")
        self.btn_tool_limpieza.setCheckable(True)
        self.btn_tool_limpieza.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_tool_limpieza.setStyleSheet("QPushButton { padding: 4px 16px; background-color: #e1e4e8; border: 2px solid #d1d5da; border-radius: 12px; font-weight: bold; color: #586069; font-size: 11px; } QPushButton:checked { background-color: #2da44e; border-color: #2da44e; color: white; }")
        self.btn_tool_limpieza.clicked.connect(self.toggle_modo_limpieza)
        
        self.btn_aplicar_limpieza = QPushButton("Eliminar áreas")
        self.btn_aplicar_limpieza.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_aplicar_limpieza.setStyleSheet("""
            QPushButton { 
                padding: 5px 15px; 
                background-color: transparent; 
                border: 2px solid #0969da; 
                color: #0969da; 
                border-radius: 4px; 
                font-weight: bold; 
                font-size: 11px; 
            }
            QPushButton:hover { 
                background-color: #0969da; 
                color: white; 
            }
        """)
        self.btn_aplicar_limpieza.hide()
        self.btn_aplicar_limpieza.clicked.connect(self.aplicar_limpieza)
        
        self.btn_limpiar_todo = QPushButton("Deshacer limpieza")
        self.btn_limpiar_todo.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_limpiar_todo.setStyleSheet("""
            QPushButton { 
                padding: 5px 10px; 
                background-color: transparent; 
                border: 2px solid #57606a; 
                color: #57606a; 
                border-radius: 4px; 
                font-weight: bold; 
                font-size: 11px; 
            }
            QPushButton:hover { 
                background-color: #57606a; 
                color: white; 
            }
        """)
        self.btn_limpiar_todo.hide()
        self.btn_limpiar_todo.clicked.connect(self.deshacer_limpieza)
        
        # Herramientas de unión de ramas (Esqueleto)
        self.frame_tools_esqueleto = QFrame()
        self.frame_tools_esqueleto.setStyleSheet("QFrame { border: none; background: transparent; }")
        layout_tools_esqueleto = QHBoxLayout(self.frame_tools_esqueleto)
        layout_tools_esqueleto.setContentsMargins(0,0,0,0)
        
        self.frame_subtools_union = QFrame()
        self.frame_subtools_union.setStyleSheet("QFrame { background-color: #f1f3f5; border-radius: 6px; margin-left: 10px; }")
        layout_sub = QHBoxLayout(self.frame_subtools_union)
        layout_sub.setContentsMargins(5,2,5,2)
        
        self.btn_sub_pincel = QPushButton()
        self.btn_sub_pincel.setIcon(QIcon("assets/buttons/editar.png"))
        self.btn_sub_pincel.setIconSize(QSize(20, 20))
        self.btn_sub_pincel.setToolTip("Pincel")
        
        self.btn_sub_linea = QPushButton()
        self.btn_sub_linea.setIcon(QIcon("assets/buttons/recta.png"))
        self.btn_sub_linea.setIconSize(QSize(20, 20))
        self.btn_sub_linea.setToolTip("Línea Recta")
        
        self.btn_sub_goma = QPushButton()
        self.btn_sub_goma.setIcon(QIcon("assets/buttons/goma.png"))
        self.btn_sub_goma.setIconSize(QSize(20, 20))
        self.btn_sub_goma.setToolTip("Goma")
        
        estilo_sub = "QPushButton { background-color: transparent; border: none; padding: 2px; } QPushButton:hover { background-color: #eaf2ff; border-radius: 17px; } QPushButton:checked { background-color: #cce5ff; border: 1px solid #007bff; border-radius: 17px; }"
        for btn in [self.btn_sub_pincel, self.btn_sub_linea, self.btn_sub_goma]:
            btn.setCheckable(True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet(estilo_sub)
            btn.setFixedSize(34, 34)
            btn.clicked.connect(self.cambiar_subherramienta)
            layout_sub.addWidget(btn)
        
        # Botón de deshacer paso a paso (dentro del sub-frame de herramientas)
        self.btn_deshacer_paso = QPushButton()
        self.btn_deshacer_paso.setIcon(QIcon("assets/buttons/deshacer.png"))
        self.btn_deshacer_paso.setIconSize(QSize(20, 20))
        self.btn_deshacer_paso.setToolTip("Deshacer último paso")
        self.btn_deshacer_paso.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_deshacer_paso.setStyleSheet("QPushButton { background-color: transparent; border: none; padding: 2px; } QPushButton:hover { background-color: #eaf2ff; border-radius: 17px; } QPushButton:disabled { opacity: 0.4; }")
        self.btn_deshacer_paso.setFixedSize(34, 34)
        self.btn_deshacer_paso.clicked.connect(self.deshacer_paso)
        self.btn_deshacer_paso.setEnabled(False)
        layout_sub.addWidget(self.btn_deshacer_paso)
        
        self.btn_sub_linea.setChecked(False)
        self.frame_subtools_union.hide()
        
        lbl_unir = QLabel("Unir Ramas:")
        lbl_unir.setStyleSheet("font-size: 11px; font-weight: bold; color: #555;")
        
        self.btn_tool_unir = QPushButton("Off")
        self.btn_tool_unir.setCheckable(True)
        self.btn_tool_unir.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_tool_unir.setStyleSheet("QPushButton { padding: 4px 16px; background-color: #e1e4e8; border: 2px solid #d1d5da; border-radius: 12px; font-weight: bold; color: #586069; font-size: 11px; } QPushButton:checked { background-color: #2da44e; border-color: #2da44e; color: white; }")
        self.btn_tool_unir.clicked.connect(self.toggle_modo_union)
        
        self.btn_ver_sobrepuesta = QPushButton("Ver Original")
        self.btn_ver_sobrepuesta.setCheckable(True)
        self.btn_ver_sobrepuesta.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_ver_sobrepuesta.setStyleSheet("QPushButton { padding: 4px 12px; background-color: #e1e4e8; border: 2px solid #d1d5da; border-radius: 12px; font-weight: bold; color: #586069; font-size: 11px; } QPushButton:checked { background-color: #0969da; border-color: #0969da; color: white; }")
        self.btn_ver_sobrepuesta.clicked.connect(self.actualizar_vista)
        
        self.btn_aplicar_union = QPushButton("Guardar")
        self.btn_aplicar_union.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_aplicar_union.setStyleSheet("""
            QPushButton { 
                padding: 5px 15px; 
                background-color: transparent; 
                border: 2px solid #0969da; 
                color: #0969da; 
                border-radius: 4px; 
                font-weight: bold; 
                font-size: 11px; 
            }
            QPushButton:hover { 
                background-color: #0969da; 
                color: white; 
            }
            QPushButton:disabled { 
                background-color: transparent; 
                border: 2px solid #eaeff2; 
                color: #949da3; 
            }
        """)
        self.btn_aplicar_union.hide()
        self.btn_aplicar_union.setEnabled(False)
        self.btn_aplicar_union.clicked.connect(self.aplicar_union_esqueleto)
        
        self.btn_reset = QPushButton()
        self.btn_reset.setIcon(QIcon("assets/buttons/reset.png"))
        self.btn_reset.setIconSize(QSize(20, 20))
        self.btn_reset.setToolTip("Restablecer")
        self.btn_reset.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_reset.setStyleSheet("QPushButton { background-color: transparent; border: none; padding: 2px; } QPushButton:hover { background-color: #ffdddd; border-radius: 17px; }")
        self.btn_reset.setFixedSize(34, 34)
        self.btn_reset.hide()
        self.btn_reset.clicked.connect(self.restablecer_esqueleto)
        
        layout_tools_esqueleto.addWidget(lbl_unir)
        layout_tools_esqueleto.addWidget(self.btn_tool_unir)
        layout_tools_esqueleto.addWidget(self.btn_ver_sobrepuesta)
        layout_tools_esqueleto.addWidget(self.frame_subtools_union)
        layout_tools_esqueleto.addWidget(self.btn_aplicar_union)
        layout_tools_esqueleto.addWidget(self.btn_reset)
        layout_tools_esqueleto.addStretch()
        
        layout_tools_limpieza.addWidget(QLabel("Limpieza:"))
        layout_tools_limpieza.addWidget(self.btn_tool_limpieza)
        layout_tools_limpieza.addStretch()
        layout_tools_limpieza.addWidget(self.btn_limpiar_todo)
        layout_tools_limpieza.addWidget(self.btn_aplicar_limpieza)
        layout.addWidget(self.frame_tools_limpieza)
        layout.addWidget(self.frame_tools_esqueleto)


        
        # Filtros Individuales (Offsets) - Solo si estamos en modo investigador y no es modo lectura
        self.frame_offsets = QFrame()
        self.frame_offsets.setStyleSheet("QFrame { background-color: #f9f9f9; border: 1px solid #ddd; border-radius: 8px; margin: 5px; } QLabel { color: #333; font-weight: bold; font-size: 11px; }")
        layout_offsets = QVBoxLayout(self.frame_offsets)
        
        lbl_off_tit = QLabel("Ajuste de Filtros Individuales (Offsets)")
        lbl_off_tit.setStyleSheet("color: #003366; font-size: 12px; margin-bottom: 5px;")
        layout_offsets.addWidget(lbl_off_tit)
        
        grid_offsets = QGridLayout()
        
        # CLAHE Offset
        grid_offsets.addWidget(QLabel("Contraste:"), 0, 0)
        self.sld_o_clahe = QSlider(Qt.Orientation.Horizontal); self.sld_o_clahe.setRange(-5, 5)
        self.sld_o_clahe.setValue(self.box["offsets"]["clahe"])
        grid_offsets.addWidget(self.sld_o_clahe, 0, 1)
        
        # Gauss Offset
        grid_offsets.addWidget(QLabel("Suavizado:"), 1, 0)
        self.sld_o_gauss = QSlider(Qt.Orientation.Horizontal); self.sld_o_gauss.setRange(-6, 6)
        self.sld_o_gauss.setValue(self.box["offsets"]["gauss"])
        grid_offsets.addWidget(self.sld_o_gauss, 1, 1)
        
        # Otsu Offset
        grid_offsets.addWidget(QLabel("Umbral:"), 2, 0)
        self.sld_o_otsu = QSlider(Qt.Orientation.Horizontal); self.sld_o_otsu.setRange(-50, 50)
        self.sld_o_otsu.setValue(self.box["offsets"]["otsu"])
        grid_offsets.addWidget(self.sld_o_otsu, 2, 1)
        
        # Ruido Offset
        grid_offsets.addWidget(QLabel("Ruido:"), 3, 0)
        self.sld_o_ruido = QSlider(Qt.Orientation.Horizontal); self.sld_o_ruido.setRange(-100, 100)
        self.sld_o_ruido.setValue(self.box["offsets"].get("ruido", 0))
        grid_offsets.addWidget(self.sld_o_ruido, 3, 1)
        
        layout_offsets.addLayout(grid_offsets)
        
        # Estilo para sliders de offset
        estilo_off = "QSlider::groove:horizontal { height: 4px; background: #ddd; } QSlider::handle:horizontal { background: #3a61a0; width: 12px; height: 12px; margin: -4px 0; border-radius: 6px; }"
        for s in [self.sld_o_clahe, self.sld_o_gauss, self.sld_o_otsu, self.sld_o_ruido]:
            s.setStyleSheet(estilo_off)
            s.valueChanged.connect(self.actualizar_offsets)
            
        layout.addWidget(self.frame_offsets)
        
        # La visibilidad se controla en actualizar_vista()


        layout.addSpacing(10)


        
        # Botones inferiores
        layout_inferior = QHBoxLayout()
        
        self.btn_comparativa = QPushButton("Ver Proceso Completo")
        self.btn_comparativa.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_comparativa.setStyleSheet("""
            QPushButton { 
                padding: 8px 15px; 
                background-color: transparent; 
                border: 2px solid #2da44e; 
                color: #2da44e; 
                border-radius: 6px; 
                font-weight: bold; 
                font-size: 13px;
            }
            QPushButton:hover { 
                background-color: #2da44e; 
                color: white; 
            }
            QPushButton:disabled { 
                background-color: transparent; 
                border: 2px solid #eaeff2; 
                color: #949da3; 
            }
        """)
        self.btn_comparativa.clicked.connect(self.mostrar_comparativa)
        # Solo habilitar si el proceso está terminado (las 3 fases existen)
        self.btn_comparativa.setEnabled(len(self.fases_disponibles) == 3)
        
        self.actualizar_visibilidad_boton_limpieza()

        
        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_cerrar.setStyleSheet("""
            QPushButton { 
                padding: 8px 20px; 
                background-color: transparent; 
                border: 2px solid #cf222e; 
                color: #cf222e; 
                border-radius: 6px; 
                font-weight: bold; 
                font-size: 14px;
            }
            QPushButton:hover { 
                background-color: #cf222e; 
                color: white; 
            }
        """)
        btn_cerrar.clicked.connect(self.accept)
        
        layout_inferior.addStretch()
        layout_inferior.addWidget(self.btn_comparativa)
        layout_inferior.addStretch()
        
        layout.addLayout(layout_inferior)
        main_layout.addWidget(frame)
        self.setLayout(main_layout)
        
        self.actualizar_vista()

    def preparar_fases(self):
        from red.config import es_cliente
        if es_cliente():
            from red.cliente import asegurar_archivo_local
            asegurar_archivo_local(self.crop_path)
            
        self.undo_stack = []  # Pila de estados para deshacer paso a paso
        # Fase 0: Original
        self.fases_disponibles.append({"nombre": "ORIGINAL", "path": self.crop_path, "pixmap": None})
        
        # Obtener paso actual de la ventana principal
        paso_padre = getattr(self.parent(), "paso_actual", 0)
        
        # Fase 1: Filtrado
        path_filtrado = self.crop_path.replace("\\", "/").replace("/crops/", "/filtradas/")
        
        if es_cliente():
            from red.cliente import asegurar_archivo_local
            asegurar_archivo_local(path_filtrado)

        if self.pixmap_mem_filtrado:
            self.fases_disponibles.append({"nombre": "FILTRADO", "path": path_filtrado, "pixmap": self.pixmap_mem_filtrado})
        else:
            self.fases_disponibles.append({"nombre": "FILTRADO", "path": path_filtrado, "pixmap": None})
            
        # Fase 2: Esqueletizado
        path_esqueleto = self.crop_path.replace("\\", "/").replace("/crops/", "/esqueletos/")
        
        if es_cliente():
            from red.cliente import asegurar_archivo_local
            asegurar_archivo_local(path_esqueleto)

        # Cargar en memoria el esqueleto si existe y el proceso lo permite
        import cv2; import numpy as np
        try:
            if paso_padre >= 4 and os.path.exists(path_esqueleto):
                with open(path_esqueleto, "rb") as f:
                    file_bytes = np.frombuffer(f.read(), dtype=np.uint8)
                self.skeleton_backup = cv2.imdecode(file_bytes, cv2.IMREAD_GRAYSCALE)
            else:
                self.skeleton_backup = None
        except Exception:
            self.skeleton_backup = None
            
        self.skeleton_working = self.skeleton_backup.copy() if self.skeleton_backup is not None else None
        self.skeleton_has_changes = False
        self.fases_disponibles.append({"nombre": "ESQUELETIZADO", "path": path_esqueleto, "pixmap": None})

    def actualizar_vista(self):
        from PyQt6.QtCore import Qt
        from PyQt6.QtGui import QPainter
        
        fase = self.fases_disponibles[self.indice_fase]
        self.lbl_fase.setText(f"FASE: {fase['nombre']}")
        
        pixmap = fase["pixmap"]
        if not pixmap:
            pixmap = QPixmap(fase["path"])
            
        # Si estamos en ESQUELETIZADO, usar la imagen en memoria (skeleton_working)
        if fase["nombre"] == "ESQUELETIZADO" and getattr(self, 'skeleton_working', None) is not None:
            from PyQt6.QtGui import QImage
            h_img, w_img = self.skeleton_working.shape[:2]
            qimg = QImage(self.skeleton_working.data, w_img, h_img, w_img, QImage.Format.Format_Grayscale8)
            pixmap = QPixmap.fromImage(qimg)
        
        if pixmap and not pixmap.isNull():
            # Si estamos en ESQUELETIZADO y el botón de ver sobrepuesta está activo, combinar con la original
            if fase["nombre"] == "ESQUELETIZADO" and getattr(self, "btn_ver_sobrepuesta", None) and self.btn_ver_sobrepuesta.isChecked():
                pix_orig = QPixmap(self.crop_path)
                if pix_orig and not pix_orig.isNull():
                    combined = QPixmap(pix_orig.size())
                    combined.fill(Qt.GlobalColor.black)
                    
                    painter = QPainter(combined)
                    painter.drawPixmap(0, 0, pix_orig)
                    painter.setCompositionMode(QPainter.CompositionMode.CompositionMode_Screen)
                    painter.setOpacity(0.8)
                    painter.drawPixmap(0, 0, pixmap)
                    painter.end()
                    
                    self.label_imagen.setPixmap(combined.scaled(380, 380, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
                else:
                    self.label_imagen.setPixmap(pixmap.scaled(380, 380, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
            else:
                self.label_imagen.setPixmap(pixmap.scaled(380, 380, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
        else:
            self.label_imagen.setText(f"No se pudo cargar la imagen de {fase['nombre']}.")

        # Obtener paso actual de la ventana principal y determinar disponibilidad de fases
        paso_padre = getattr(self.parent(), "paso_actual", 0)
        
        # Determinar si la fase de FILTRADO está disponible
        filtrado_disponible = False
        if paso_padre >= 3:
            filtrado_disponible = True
        elif paso_padre == 2:
            if self.parent() and hasattr(self.parent(), "combo_vista"):
                if self.parent().combo_vista.currentText() == "Previsualización":
                    filtrado_disponible = True
            if self.parent() and hasattr(self.parent(), "frame_filtros"):
                if self.parent().frame_filtros.isVisible():
                    filtrado_disponible = True
                    
        esqueleto_disponible = (paso_padre >= 4)

        # Botones siempre visibles
        self.btn_ant.setVisible(True)
        self.btn_sig.setVisible(True)
        self.btn_comparativa.setVisible(fase["nombre"] == "ESQUELETIZADO")

        # Habilitar navegación anterior (<)
        if self.indice_fase == 0:
            self.set_button_enabled(self.btn_ant, False)
        elif self.indice_fase == 1:
            self.set_button_enabled(self.btn_ant, True)
        elif self.indice_fase == 2:
            self.set_button_enabled(self.btn_ant, filtrado_disponible)
        else:
            self.set_button_enabled(self.btn_ant, False)

        # Habilitar navegación siguiente (>)
        if self.indice_fase == 0:
            self.set_button_enabled(self.btn_sig, filtrado_disponible)
        elif self.indice_fase == 1:
            self.set_button_enabled(self.btn_sig, esqueleto_disponible)
        else:
            self.set_button_enabled(self.btn_sig, False)

        self.set_button_enabled(self.btn_comparativa, fase["nombre"] == "ESQUELETIZADO" and esqueleto_disponible)

        # Actualizar visibilidad de herramientas según la fase
        if fase["nombre"] == "FILTRADO":
            filtros_activos = True
            if self.parent() and hasattr(self.parent(), "combo_vista"):
                if self.parent().combo_vista.currentText() != "Previsualización":
                    filtros_activos = False
            elif len(self.fases_disponibles) >= 3:
                filtros_activos = False

            if not filtros_activos:
                self.frame_offsets.hide()
                self.frame_tools_limpieza.hide()
                self.resize(450, 520)
            else:
                self.frame_offsets.show()
                self.frame_tools_limpieza.show()
                self.resize(500, 780)
            self.frame_tools_esqueleto.hide()
        elif fase["nombre"] == "ESQUELETIZADO":
            self.frame_offsets.hide()
            self.frame_tools_limpieza.hide()
            self.frame_tools_esqueleto.show()
            self.resize(500, 600)
        else:
            self.frame_offsets.hide()
            self.frame_tools_limpieza.hide()
            self.frame_tools_esqueleto.hide()
            self.resize(450, 520)


    def actualizar_offsets(self):
        # 1. Actualizar valores en el box
        self.box["offsets"]["clahe"] = self.sld_o_clahe.value()
        self.box["offsets"]["ruido"] = self.sld_o_ruido.value()
        
        # 2. Notificar al padre para que reprocese globalmente
        if hasattr(self.parent(), "previsualizar_filtrado"):
            self.parent().previsualizar_filtrado()
            
            # 3. Obtener el nuevo pixmap filtrado para este crop
            nombre = os.path.basename(self.crop_path)
            if nombre in self.parent().crops_filtrados_temp:
                bin_img = self.parent().crops_filtrados_temp[nombre]
                import cv2; from PyQt6.QtGui import QImage
                h, w = bin_img.shape
                qimg = QImage(bin_img.data, w, h, w, QImage.Format.Format_Grayscale8)
                nuevo_pixmap = QPixmap.fromImage(qimg)
                
                # Actualizar en fases_disponibles
                for f in self.fases_disponibles:
                    if f["nombre"] == "FILTRADO":
                        f["pixmap"] = nuevo_pixmap
                        break
                
                # Si estamos viendo la fase filtrado, refrescar
                if self.fases_disponibles[self.indice_fase]["nombre"] == "FILTRADO":
                    self.actualizar_vista()

    def get_original_crop_size(self):
        pix = QPixmap(self.crop_path)
        if not pix.isNull(): return pix.width(), pix.height()
        return 0, 0

    def toggle_modo_limpieza(self, checked):
        if checked:
            self.label_imagen.mode = "eraser"
            self.label_imagen.setCursor(Qt.CursorShape.CrossCursor)
            self.btn_tool_limpieza.setText("On")
            self.btn_tool_unir.blockSignals(True)
            self.btn_tool_unir.setChecked(False)
            self.btn_tool_unir.setText("Off")
            self.btn_tool_unir.blockSignals(False)
        else:
            self.label_imagen.mode = "pointer"
            self.label_imagen.setCursor(Qt.CursorShape.ArrowCursor)
            self.btn_tool_limpieza.setText("Off")

    def cambiar_subherramienta(self):
        sender = self.sender()
        if not sender:
            # Desmarcar todo si no se llamó desde un botón específico
            self.btn_sub_pincel.setChecked(False)
            self.btn_sub_linea.setChecked(False)
            self.btn_sub_goma.setChecked(False)
            self.label_imagen.mode = "pointer"
            self.label_imagen.setCursor(Qt.CursorShape.ArrowCursor)
            return

        self.btn_sub_pincel.setChecked(sender == self.btn_sub_pincel)
        self.btn_sub_linea.setChecked(sender == self.btn_sub_linea)
        self.btn_sub_goma.setChecked(sender == self.btn_sub_goma)
        
        if sender == self.btn_sub_goma:
            self.label_imagen.mode = "eraser_union"
            self.label_imagen.setCursor(Qt.CursorShape.BlankCursor)
        elif sender == self.btn_sub_pincel:
            self.label_imagen.mode = "pencil_freehand"
            self.label_imagen.setCursor(self.label_imagen.pencil_cursor)
        elif sender == self.btn_sub_linea:
            self.label_imagen.mode = "pencil_line"
            self.label_imagen.setCursor(self.label_imagen.pencil_cursor)

    def toggle_modo_union(self, checked):
        if checked:
            self.frame_subtools_union.show()
            # En lugar de seleccionar un modo por defecto, desmarcamos todos para obligar al usuario a elegir
            self.btn_sub_pincel.setChecked(False)
            self.btn_sub_linea.setChecked(False)
            self.btn_sub_goma.setChecked(False)
            self.label_imagen.mode = "pointer"
            self.label_imagen.setCursor(Qt.CursorShape.ArrowCursor)
            
            self.btn_tool_unir.setText("On")
            self.actualizar_visibilidad_boton_limpieza()
            self.btn_tool_limpieza.blockSignals(True)
            self.btn_tool_limpieza.setChecked(False)
            self.btn_tool_limpieza.setText("Off")
            self.btn_tool_limpieza.blockSignals(False)
        else:
            self.frame_subtools_union.hide()
            self.label_imagen.mode = "pointer"
            self.label_imagen.setCursor(Qt.CursorShape.ArrowCursor)
            self.btn_tool_unir.setText("Off")
            self.actualizar_visibilidad_boton_limpieza()

    def _actualizar_pixmap_esqueleto_memoria(self):
        """Genera un QPixmap a partir de skeleton_working y lo asigna a la fase ESQUELETIZADO."""
        if self.skeleton_working is None:
            return
        from PyQt6.QtGui import QImage
        h_img, w_img = self.skeleton_working.shape[:2]
        qimg = QImage(self.skeleton_working.data, w_img, h_img, w_img, QImage.Format.Format_Grayscale8)
        pix = QPixmap.fromImage(qimg)
        for f in self.fases_disponibles:
            if f["nombre"] == "ESQUELETIZADO":
                f["pixmap"] = pix
                break

    def aplicar_union_esqueleto(self):
        """Guarda todos los cambios (goma + uniones) al esqueleto en disco."""
        import cv2; import numpy as np
        from skimage.morphology import skeletonize

        if self.skeleton_working is None:
            return

        # Generar la imagen final con las uniones de ramas aplicadas sobre el working actual
        img_final = self.skeleton_working.copy()
        for conn_path in self.box.get("manual_connections", []):
            for i in range(len(conn_path) - 1):
                p1 = (conn_path[i][0], conn_path[i][1])
                p2 = (conn_path[i+1][0], conn_path[i+1][1])
                cv2.line(img_final, p1, p2, 255, 2)

        # Skeletonize para limpiar uniones a 1px
        if len(self.box.get("manual_connections", [])) > 0:
            img_bool = img_final > 0
            skeleton = skeletonize(img_bool)
            img_final = (skeleton * 255).astype(np.uint8)

        # Mostrar previsualización del resultado final
        self.skeleton_working = img_final
        self._actualizar_pixmap_esqueleto_memoria()
        self.actualizar_vista()
        self.label_imagen.update()

        # Preguntar si desea guardar permanentemente
        from vistas.utilidades import DialogoConfirmacion
        diag = DialogoConfirmacion("Guardar Cambios", "¿Confirmas que deseas guardar estos cambios?")
        if diag.exec():
            # Guardar permanentemente en disco
            path_esqueleto = self.crop_path.replace("\\", "/").replace("/crops/", "/esqueletos/")
            is_success, im_buf_arr = cv2.imencode(".png", img_final)
            if is_success:
                im_buf_arr.tofile(path_esqueleto)
            else:
                cv2.imwrite(path_esqueleto, img_final)
            self.box["esqueleto_modificado"] = True
            # Limpiar conexiones ya aplicadas
            self.box["manual_connections"] = []
            # Actualizar backup a la versión guardada
            self.skeleton_backup = img_final.copy()
            self.skeleton_has_changes = False
            self.undo_stack = []
            self.actualizar_visibilidad_boton_limpieza()
            self.actualizar_vista()
            self.label_imagen.update()

            # Desactivar modo unión tras guardar cambios
            self.btn_tool_unir.setChecked(False)
            self.toggle_modo_union(False)

            # Refrescar imagen global
            self._refrescar_imagen_global_esqueleto()

            self.mostrar_notificacion("Éxito", "Los cambios se han guardado correctamente.", "info")
        else:
            # Revertir: recalcular working desde el estado actual sin las uniones
            # (las uniones no se habían aplicado al working antes, solo temporalmente)
            # Restaurar skeleton_working al estado previo sin uniones aplicadas
            # Re-leer desde la versión que teníamos antes de aplicar uniones
            self.skeleton_working = self.skeleton_backup.copy()
            self.skeleton_has_changes = False
            self.box["manual_connections"] = []
            self.undo_stack = []
            self._actualizar_pixmap_esqueleto_memoria()
            self.actualizar_vista()
            self.label_imagen.update()
            self.actualizar_visibilidad_boton_limpieza()

    def _refrescar_imagen_global_esqueleto(self):
        """Método auxiliar para actualizar la imagen global de esqueleto en la ventana principal."""
        parent_win = self.parent()
        if parent_win and hasattr(parent_win, "pixmaps_globales") and hasattr(parent_win, "construir_imagen_global"):
            pixmap_esqueleto = parent_win.construir_imagen_global("esqueletos")
            parent_win.pixmaps_globales["Esqueleto"] = pixmap_esqueleto
            if hasattr(parent_win, "combo_vista") and parent_win.combo_vista.currentText() == "Esqueleto":
                parent_win.visor_imagen.set_view_mode("Esqueleto", pixmap_esqueleto)

    def actualizar_visibilidad_boton_limpieza(self):
        areas = self.box.get("removal_areas", [])
        has_unbaked = any(not a.get("baked", False) for a in areas)
        has_any = len(areas) > 0
        self.btn_aplicar_limpieza.setVisible(has_unbaked)
        self.btn_limpiar_todo.setVisible(has_any)
        
        has_uniones = len(self.box.get("manual_connections", [])) > 0
        has_any_changes = has_uniones or getattr(self, 'skeleton_has_changes', False)
        has_undo_steps = len(getattr(self, 'undo_stack', [])) > 0
        is_union_on = self.btn_tool_unir.isChecked()
        # Guardar Cambios: visible si hay cambios pendientes o el modo unión está activo, pero solo habilitado si hay cambios reales
        self.btn_aplicar_union.setVisible(has_any_changes or is_union_on)
        self.btn_aplicar_union.setEnabled(has_any_changes)
        # Restablecer / Reset: visible solo cuando unir ramas está OFF y la imagen ya fue modificada permanentemente (amarillo)
        is_modified_saved = self.box.get("esqueleto_modificado", False)
        show_reset = (not is_union_on) and is_modified_saved
        self.btn_reset.setVisible(show_reset)
        # Deshacer paso: habilitado si hay pasos en la pila
        self.btn_deshacer_paso.setEnabled(has_undo_steps)

    def _push_undo_state(self):
        """Guarda el estado actual en la pila de deshacer antes de una acción."""
        import copy
        state = {
            'skeleton': self.skeleton_working.copy() if self.skeleton_working is not None else None,
            'connections': copy.deepcopy(self.box.get("manual_connections", []))
        }
        self.undo_stack.append(state)
        self.skeleton_has_changes = True

    def deshacer_paso(self):
        """Deshace el último paso individual (una acción de goma o una conexión)."""
        if not getattr(self, 'undo_stack', None) or len(self.undo_stack) == 0:
            return
        state = self.undo_stack.pop()
        if state['skeleton'] is not None:
            self.skeleton_working = state['skeleton']
        self.box["manual_connections"] = state['connections']
        self.skeleton_has_changes = len(self.undo_stack) > 0
        self._actualizar_pixmap_esqueleto_memoria()
        self.actualizar_visibilidad_boton_limpieza()
        self.actualizar_vista()
        self.label_imagen.update()

    def restablecer_esqueleto(self):
        """Revierte los cambios al estado original.
        - Si unir ramas está ON: Revierte los cambios temporales en memoria al último backup guardado.
        - Si unir ramas está OFF: Recalcula y restaura el esqueleto original sin modificaciones a partir de la imagen filtrada.
        """
        import numpy as np
        is_union_on = self.btn_tool_unir.isChecked()
        
        if not is_union_on:
            from vistas.utilidades import DialogoConfirmacion
            diag = DialogoConfirmacion(
                "Restaurar Esqueleto",
                "¿Confirmas que deseas restaurar este esqueleto al original?\nSe perderán permanentemente todos los cambios manuales guardados."
            )
            if diag.exec():
                path_filtrado = self.crop_path.replace("\\", "/").replace("/crops/", "/filtradas/")
                import cv2
                from skimage.morphology import skeletonize
                try:
                    with open(path_filtrado, "rb") as f:
                        file_bytes = np.frombuffer(f.read(), dtype=np.uint8)
                    img_raw = cv2.imdecode(file_bytes, cv2.IMREAD_GRAYSCALE)
                except Exception:
                    img_raw = None
                if img_raw is not None:
                    _, bin_img = cv2.threshold(img_raw, 127, 255, cv2.THRESH_BINARY)
                    img_bool = bin_img > 0
                    skeleton = skeletonize(img_bool)
                    img_final = (skeleton * 255).astype(np.uint8)
                    
                    path_esqueleto = self.crop_path.replace("\\", "/").replace("/crops/", "/esqueletos/")
                    is_success, im_buf_arr = cv2.imencode(".png", img_final)
                    if is_success:
                        im_buf_arr.tofile(path_esqueleto)
                    else:
                        cv2.imwrite(path_esqueleto, img_final)
                    
                    self.skeleton_working = img_final.copy()
                    self.skeleton_backup = img_final.copy()
                    self.box["esqueleto_modificado"] = False
                    self.box["manual_connections"] = []
                    self.undo_stack = []
                    self.skeleton_has_changes = False
                    
                    self._actualizar_pixmap_esqueleto_memoria()
                    self.actualizar_visibilidad_boton_limpieza()
                    self.actualizar_vista()
                    self.label_imagen.update()
                    self._refrescar_imagen_global_esqueleto()
                    self.mostrar_notificacion("Restablecido", "El esqueleto ha sido restaurado al filtrado original.", "info")
            return

        # Si unir ramas está ON: Restaurar esqueleto al último estado guardado en disco
        if getattr(self, 'skeleton_backup', None) is not None:
            self.skeleton_working = self.skeleton_backup.copy()
            self.skeleton_has_changes = False
        # Limpiar todas las conexiones manuales y pila de undo
        self.box["manual_connections"] = []
        self.undo_stack = []
        self._actualizar_pixmap_esqueleto_memoria()
        self.actualizar_visibilidad_boton_limpieza()
        self.actualizar_vista()
        self.label_imagen.update()

    def aplicar_limpieza(self):
        from vistas.utilidades import DialogoConfirmacion
        diag = DialogoConfirmacion("Eliminar Áreas", "¿Confirmas que deseas eliminar permanentemente estas áreas? Esto modificará la imagen base.")
        if diag.exec():
            # Aplicar permanentemente a la imagen en memoria (BGR)
            nombre = os.path.basename(self.crop_path)
            if hasattr(self.parent(), "crops_en_memoria") and nombre in self.parent().crops_en_memoria:
                img_bgr = self.parent().crops_en_memoria[nombre]
                import cv2
                for area in self.box.get("removal_areas", []):
                    if not area.get("baked", False):
                        ax, ay, aw, ah = area["x"], area["y"], area["w"], area["h"]
                        cv2.rectangle(img_bgr, (ax, ay), (ax + aw, ay + ah), (0, 0, 0), -1)
                        area["baked"] = True
                
                self.actualizar_visibilidad_boton_limpieza()
                self.actualizar_offsets() # Esto reprocesará con la nueva imagen base
                self.label_imagen.update()
                self.mostrar_notificacion("Limpieza Aplicada", "Las áreas se han eliminado de la imagen base.", "info")

    def deshacer_limpieza(self):
        areas = self.box.get("removal_areas", [])
        has_unbaked = any(not a.get("baked", False) for a in areas)
        if has_unbaked:
            # Eliminar la última área unbaked
            for idx in range(len(areas) - 1, -1, -1):
                if not areas[idx].get("baked", False):
                    areas.pop(idx)
                    break
        else:
            # No hay áreas sin aplicar, restaurar todo al original
            self.box["removal_areas"] = []
            nombre = os.path.basename(self.crop_path)
            if os.path.exists(self.crop_path):
                import cv2
                img = cv2.imread(self.crop_path)
                if img is not None and hasattr(self.parent(), "crops_en_memoria"):
                    self.parent().crops_en_memoria[nombre] = img
        
        self.actualizar_visibilidad_boton_limpieza()
        self.actualizar_offsets()
        self.label_imagen.update()

    def mostrar_notificacion(self, t, m, tipo):
        from vistas.utilidades import DialogoNotificacion
        DialogoNotificacion(t, m, tipo, self).exec()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            if self.window().windowHandle():
                self.window().windowHandle().startSystemMove()
            event.accept()



    def set_button_enabled(self, button, enabled):
        button.setEnabled(enabled)
        if enabled:
            button.setGraphicsEffect(None)
        else:
            from PyQt6.QtWidgets import QGraphicsOpacityEffect
            effect = QGraphicsOpacityEffect()
            effect.setOpacity(0.3)
            button.setGraphicsEffect(effect)

    def mostrar_anterior(self):
        if self.indice_fase > 0:
            self.indice_fase -= 1
            self.actualizar_vista()

    def mostrar_siguiente(self):
        if self.indice_fase < len(self.fases_disponibles) - 1:
            self.indice_fase += 1
            self.actualizar_vista()

    def mostrar_comparativa(self):
        diag = DialogoComparativo(self.fases_disponibles, self.parent())
        diag.exec()

    def showEvent(self, event):
        super().showEvent(event)
        if self.parent():
            p_geom = self.parent().geometry()
            self.move(p_geom.x() + (p_geom.width() - self.width()) // 2, p_geom.y() + (p_geom.height() - self.height()) // 2)
        else:
            screen = QApplication.primaryScreen().geometry()
            self.move((screen.width() - self.width()) // 2, (screen.height() - self.height()) // 2)


class DialogoConfirmacion(QDialog):
    def __init__(self, titulo, mensaje, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True)
        self.resultado = False
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        if parent:
            parent.installEventFilter(self)

        layout = QVBoxLayout(self)
        frame = QFrame(self)
        frame.setStyleSheet("""
            QFrame { background-color: #FFFFFF; border-radius: 15px; border: 2px solid #cf222e; }
            QLabel { color: #333333; font-size: 15px; padding: 10px; border: none;}
            QPushButton { border-radius: 8px; font-size: 13px; outline: none; }
            
            QPushButton#btn_eliminar { 
                background-color: transparent; 
                color: #cf222e; 
                border: 2px solid #cf222e; 
                font-weight: bold;
                padding: 8px 20px;
            }
            QPushButton#btn_eliminar:hover { 
                background-color: #cf222e; 
                color: white; 
                border-color: #cf222e;
            }
            
            QPushButton#btn_mantener { 
                background-color: transparent; 
                color: #57606a; 
                border: 2px solid #d0d7de; 
                font-weight: bold;
                padding: 8px 20px;
            }
            QPushButton#btn_mantener:hover { 
                background-color: #f3f4f6; 
                color: #24292f; 
                border-color: #8c959f;
            }
        """)

        flayout = QVBoxLayout(frame)
        lbl_titulo = QLabel(f"<b>{titulo}</b>")
        lbl_titulo.setStyleSheet("color: #cc0000; font-size: 18px; border: none;")
        lbl_titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)

        lbl_mensaje = QLabel(mensaje)
        lbl_mensaje.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_mensaje.setWordWrap(True)

        btn_layout = QHBoxLayout()
        btn_mantener = QPushButton("Mantener")
        btn_mantener.setObjectName("btn_mantener")
        btn_mantener.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_mantener.clicked.connect(self.cancelar)

        btn_eliminar = QPushButton("Eliminar")
        btn_eliminar.setObjectName("btn_eliminar")
        btn_eliminar.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_eliminar.clicked.connect(self.aceptar)

        btn_layout.addWidget(btn_mantener)
        btn_layout.addWidget(btn_eliminar)

        flayout.addWidget(lbl_titulo)
        flayout.addWidget(lbl_mensaje)
        flayout.addLayout(btn_layout)
        layout.addWidget(frame)
        self.setLayout(layout)

    def eventFilter(self, obj, event):
        from PyQt6.QtCore import QEvent
        if obj == self.parent() and (event.type() == QEvent.Type.Resize or event.type() == QEvent.Type.Move):
            self.centrar_en_padre()
        return super().eventFilter(obj, event)

    def centrar_en_padre(self):
        if self.parent():
            p_geom = self.parent().geometry()
            self.move(p_geom.x() + (p_geom.width() - self.width()) // 2, p_geom.y() + (p_geom.height() - self.height()) // 2)

    def showEvent(self, event):
        from PyQt6.QtWidgets import QApplication
        super().showEvent(event)
        if self.parent():
            self.centrar_en_padre()
        else:
            screen = QApplication.primaryScreen().geometry()
            self.move((screen.width() - self.width()) // 2, (screen.height() - self.height()) // 2)

    def mousePressEvent(self, event):
        from PyQt6.QtCore import Qt
        if event.button() == Qt.MouseButton.LeftButton:
            if self.windowHandle():
                self.windowHandle().startSystemMove()
            else:
                self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            event.accept()
        else:
            super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        from PyQt6.QtCore import Qt
        if not self.windowHandle() and hasattr(self, "_drag_pos") and event.buttons() == Qt.MouseButton.LeftButton:
            self.move(event.globalPosition().toPoint() - self._drag_pos)
            event.accept()
        else:
            super().mouseMoveEvent(event)

    def aceptar(self): self.resultado = True; self.accept()
    def cancelar(self): self.resultado = False; self.reject()

class InteractiveImageViewer(QLabel):
    conteo_actualizado = pyqtSignal(int)
    nueva_caja_dibujada = pyqtSignal(int, int, int, int)
    nivel_zoom_cambiado = pyqtSignal(int)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMouseTracking(True)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.original_pixmap = None
        self._current_pixmap = None
        self.boxes = []
        self.hovered_index = -1
        self.active_index = -1
        self.view_mode = "Original"
        self.current_tool = "pointer" 
        self.is_drawing = False
        self.draw_start_pos = None
        self.draw_current_pos = None
        self.is_panning = False
        self.pan_start_pos = None
        self.pan_x = 0
        self.pan_y = 0
        self.zoom_level = 100
        self.zoom_locked = False

    def set_image_and_boxes(self, pixmap, bounding_boxes):
        self.original_pixmap = pixmap
        self.boxes = bounding_boxes
        self.hovered_index = -1
        self.pan_x = 0
        self.pan_y = 0
        self.draw_current_state()
        self.conteo_actualizado.emit(len(self.boxes))

    def set_view_mode(self, mode, new_pixmap):
        self.view_mode = mode
        self.original_pixmap = new_pixmap
        self.draw_current_state()

    def map_mouse_to_original(self, mouse_pos):
        if not self.original_pixmap or self.original_pixmap.isNull(): return None
        pix_w, pix_h = self.original_pixmap.width(), self.original_pixmap.height()
        lbl_w, lbl_h = self.width(), self.height()
        if lbl_w == 0 or lbl_h == 0: return None
        
        base_scale = min(lbl_w / pix_w, lbl_h / pix_h)
        actual_scale = base_scale * (self.zoom_level / 100.0)
        scaled_w = int(pix_w * actual_scale)
        scaled_h = int(pix_h * actual_scale)
        
        draw_x = (lbl_w - scaled_w) // 2 + self.pan_x
        draw_y = (lbl_h - scaled_h) // 2 + self.pan_y
        mx, my = mouse_pos.x(), mouse_pos.y()
        
        if draw_x <= mx <= draw_x + scaled_w and draw_y <= my <= draw_y + scaled_h:
            orig_x = (mx - draw_x) / actual_scale
            orig_y = (my - draw_y) / actual_scale
            return orig_x, orig_y
        return None

    def wheelEvent(self, event):
        if not self.original_pixmap or self.zoom_locked: return
        delta = event.angleDelta().y()
        if delta > 0: new_zoom = min(400, self.zoom_level + 15)
        else: new_zoom = max(50, self.zoom_level - 15)
        if new_zoom != self.zoom_level: self.set_zoom(new_zoom)

    def mousePressEvent(self, event):
        if not self.original_pixmap: return
        

        if event.button() == Qt.MouseButton.LeftButton:
            if self.current_tool == "pointer":
                if self.hovered_index != -1:
                    crop_path_base = self.boxes[self.hovered_index]["crop_path"]
                    crop_path = crop_path_base.replace("\\", "/")
                    pixmap_mem = None
                    if hasattr(self.window(), 'crops_filtrados_temp'):
                        nombre_base = os.path.basename(crop_path)
                        if self.view_mode == "Filtrada" or self.view_mode == "Previsualización":
                            arr = self.window().crops_filtrados_temp.get(nombre_base)
                            if arr is not None:
                                h, w = arr.shape
                                qimg = QImage(arr.data, w, h, w, QImage.Format.Format_Grayscale8)
                                pixmap_mem = QPixmap.fromImage(qimg)
                    
                    if self.view_mode == "Filtrada": crop_path = crop_path.replace("/crops/", "/filtradas/")
                    elif self.view_mode == "Esqueleto": crop_path = crop_path.replace("/crops/", "/esqueletos/")

                    from red.config import es_cliente
                    if es_cliente():
                        from red.cliente import asegurar_archivo_local
                        asegurar_archivo_local(crop_path)

                    if os.path.exists(crop_path) or pixmap_mem:
                        self.active_index = self.hovered_index
                        self.draw_current_state()
                        DialogoVistaCelular(self.boxes[self.hovered_index], pixmap_mem, self.view_mode, self.window()).exec()
                        self.active_index = -1
                        # Refrescar imagen global de esqueleto tras cerrar el diálogo por si hubo ediciones
                        if self.view_mode == "Esqueleto" and hasattr(self.window(), 'construir_imagen_global'):
                            pixmap_esqueleto = self.window().construir_imagen_global("esqueletos")
                            self.window().pixmaps_globales["Esqueleto"] = pixmap_esqueleto
                            self.window().visor_imagen.set_view_mode("Esqueleto", pixmap_esqueleto)
                        self.draw_current_state()

                else:
                    self.is_panning = True
                    self.pan_start_pos = event.pos()
                    self.setCursor(Qt.CursorShape.ClosedHandCursor)
            elif self.current_tool == "draw":
                self.is_drawing = True
                self.draw_start_pos = event.pos()
                self.draw_current_pos = event.pos()
            elif self.current_tool == "delete":
                if self.hovered_index != -1:
                    index_to_delete = self.hovered_index
                    diag = DialogoConfirmacion("Eliminar Detección", "¿Estás seguro de descartar esta célula?")
                    diag.exec()
                    if diag.resultado:
                        if index_to_delete < len(self.boxes):
                            self.boxes.pop(index_to_delete)
                            self.hovered_index = -1
                            self.draw_current_state()
                            self.conteo_actualizado.emit(len(self.boxes))

    def mouseMoveEvent(self, event):
        if not self.original_pixmap: return
        if self.is_panning:
            delta = event.pos() - self.pan_start_pos
            self.pan_x += delta.x()
            self.pan_y += delta.y()
            self.pan_start_pos = event.pos()
            self.update() 
            return
        if self.is_drawing:
            self.draw_current_pos = event.pos()
            self.draw_current_state()
            return
        if self.current_tool == "draw":
            if self.hovered_index != -1:
                self.hovered_index = -1
                self.draw_current_state()
            return
        if not self.boxes: return
        orig_coords = self.map_mouse_to_original(event.pos())
        new_hovered_index = -1
        if orig_coords:
            ox, oy = orig_coords
            for i in range(len(self.boxes)-1, -1, -1):
                box = self.boxes[i]
                if box["x"] <= ox <= box["x"] + box["w"] and box["y"] <= oy <= box["y"] + box["h"]:
                    new_hovered_index = i
                    break
        if new_hovered_index != self.hovered_index:
            self.hovered_index = new_hovered_index
            self.draw_current_state()

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            if self.is_panning:
                self.is_panning = False
                self.setCursor(Qt.CursorShape.ArrowCursor)
            elif self.is_drawing:
                self.is_drawing = False
                orig_start = self.map_mouse_to_original(self.draw_start_pos)
                orig_end = self.map_mouse_to_original(self.draw_current_pos)
                if orig_start and orig_end:
                    x1, y1 = orig_start; x2, y2 = orig_end
                    x = min(x1, x2); y = min(y1, y2)
                    w = abs(x2 - x1); h = abs(y2 - y1)
                    if w >= MIN_MICROGLIA_SIZE and h >= MIN_MICROGLIA_SIZE: self.nueva_caja_dibujada.emit(int(x), int(y), int(w), int(h))
                self.draw_start_pos = None; self.draw_current_pos = None
                self.draw_current_state()

    def leaveEvent(self, event):
        if self.hovered_index != -1:
            self.hovered_index = -1
            self.draw_current_state()
        super().leaveEvent(event)

    def borrar_rama_esqueleto(self, pos):
        if not self.original_pixmap: return
        
        orig_coords = self.map_mouse_to_original(pos)
        if not orig_coords: return
        
        ox, oy = orig_coords
        
        # Modificar el pixmap original (que es el esqueleto)
        # Pintamos un círculo negro para 'borrar' la rama
        painter = QPainter(self.original_pixmap)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor(0, 0, 0)) # Negro
        painter.drawEllipse(QPoint(int(ox), int(oy)), 8, 8) # Radio de 8 pixeles para facilitar el borrado
        painter.end()
        
        self.draw_current_state()

    def set_zoom(self, value):
        if not self.original_pixmap or self.original_pixmap.isNull(): return
        if not self.zoom_locked: self.zoom_level = value; self.draw_current_state(); self.nivel_zoom_cambiado.emit(value)

    def lock_zoom(self, locked): self.zoom_locked = locked

    def draw_current_state(self):
        """Dibuja la imagen, las cajas normales y la caja resaltada."""
        if not self.original_pixmap or self.original_pixmap.isNull():
            self._current_pixmap = None
            self.update()
            return
        pix_w, pix_h = self.original_pixmap.width(), self.original_pixmap.height()
        lbl_w, lbl_h = self.width(), self.height()
        if lbl_w == 0 or lbl_h == 0: return
        
        base_scale = min(lbl_w / pix_w, lbl_h / pix_h)
        actual_scale = base_scale * (self.zoom_level / 100.0)
        scaled_w = int(pix_w * actual_scale)
        scaled_h = int(pix_h * actual_scale)
        if scaled_w == 0 or scaled_h == 0: return
        
        scaled_pix = self.original_pixmap.scaled(scaled_w, scaled_h, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
        temp_pixmap = scaled_pix.copy()
        painter = QPainter(temp_pixmap)
        
        for i, box in enumerate(self.boxes):
            rect = QRect(int(box["x"] * actual_scale), int(box["y"] * actual_scale), int(box["w"] * actual_scale), int(box["h"] * actual_scale))
            
            # Detectar si tiene offsets o limpiezas (solo en modo previsualización o filtrada)
            has_offset = False
            if self.view_mode in ["Previsualización", "Filtrada"]:
                offs = box.get("offsets", {})
                if offs.get("clahe", 0) != 0 or offs.get("gauss", 0) != 0 or offs.get("otsu", 0) != 0:
                    has_offset = True
                if len(box.get("removal_areas", [])) > 0:
                    has_offset = True

            # Detectar si el esqueleto ha sido modificado (en modo esqueleto)
            has_skeleton_modified = False
            if self.view_mode == "Esqueleto":
                if box.get("esqueleto_modificado", False):
                    has_skeleton_modified = True

            # Si es modificado (offset o esqueleto), usar color amarillo/naranja
            is_modified = has_offset or has_skeleton_modified

            if i == self.active_index:
                color = QColor(9, 105, 218) # Azul GitHub
                pen = QPen(color)
                pen.setWidth(max(2, int(pix_w * 0.003 * actual_scale)))
            elif i == self.hovered_index and self.current_tool != "draw":
                color = QColor(255, 215, 0) if is_modified else QColor(0, 255, 0)
                pen = QPen(color)
                pen.setWidth(max(2, int(pix_w * 0.003 * actual_scale)))
            else:
                color = QColor(255, 215, 0, 180) if is_modified else QColor(0, 255, 0, 120)
                pen = QPen(color)
                pen.setWidth(max(1, int(pix_w * 0.0015 * actual_scale)))
                
            painter.setPen(pen)
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRect(rect)
            
            if is_modified:
                painter.drawText(rect.x() + 2, rect.y() + 12, "*")

            
        if self.is_drawing and self.draw_start_pos and self.draw_current_pos:
            orig_start = self.map_mouse_to_original(self.draw_start_pos); orig_end = self.map_mouse_to_original(self.draw_current_pos)
            if orig_start and orig_end:
                ox1, oy1 = orig_start; ox2, oy2 = orig_end
                rx = min(ox1, ox2); ry = min(oy1, oy2); rw = abs(ox2 - ox1); rh = abs(oy2 - oy1)
                pen_draw = QPen(QColor(0, 150, 255)); pen_draw.setWidth(max(2, int(pix_w * 0.003 * actual_scale))); painter.setPen(pen_draw); painter.setBrush(Qt.BrushStyle.NoBrush)
                painter.drawRect(int(rx * actual_scale), int(ry * actual_scale), int(rw * actual_scale), int(rh * actual_scale))
                
        painter.end()
        self._current_pixmap = temp_pixmap; self.update()

    def paintEvent(self, event):
        if self._current_pixmap and not self._current_pixmap.isNull():
            painter = QPainter(self)
            x = (self.width() - self._current_pixmap.width()) // 2 + self.pan_x
            y = (self.height() - self._current_pixmap.height()) // 2 + self.pan_y
            painter.drawPixmap(x, y, self._current_pixmap); painter.end()
        else: super().paintEvent(event)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self.original_pixmap: self.draw_current_state()


from procesamiento.compartir_reporte import DialogoCompartirReporte


class DialogoHistorial(QDialog):
    def __init__(self, id_usuario, rol, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True); self.id_usuario = id_usuario; self.rol = rol; self.seleccion = None
        from PyQt6.QtGui import QColor, QIcon
        from PyQt6.QtWidgets import QTabWidget, QWidget
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        
        main_layout = QVBoxLayout(self)
        self.frame = QFrame()
        self.frame.setStyleSheet("""
            QFrame { background-color: #FFFFFF; border-radius: 0px; border: 2px solid #3a61a0; }
            QLabel { color: #333333; border: none; }
            QPushButton { border-radius: 0px; font-weight: bold; padding: 10px; }
            QToolTip {
                background-color: #24292f;
                color: #ffffff;
                border: 1px solid #24292f;
                border-radius: 4px;
                padding: 3px 6px;
                font-size: 10px;
            }
        """)
        
        flayout = QVBoxLayout(self.frame)
        
        # Layout de encabezado con título e icono de basura
        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(15, 10, 15, 0)
        
        lbl_titulo = QLabel("<b>Historial de Reportes</b>")
        lbl_titulo.setStyleSheet("font-size: 15px; color: #3a61a0;")
        lbl_titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.btn_compartir_icon = QPushButton()
        self.btn_compartir_icon.setIcon(QIcon("assets/buttons/compartir.png"))
        self.btn_compartir_icon.setIconSize(QSize(22, 22))
        self.btn_compartir_icon.setFixedSize(35, 35)
        self.btn_compartir_icon.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_compartir_icon.setStyleSheet("""
            QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
            QPushButton:hover { background-color: #eaf2ff; border-radius: 17px; }
            QPushButton:disabled { opacity: 0.3; }
        """)
        self.btn_compartir_icon.setEnabled(False)
        self.btn_compartir_icon.setVisible(self.rol == "Investigador")
        self.btn_compartir_icon.setToolTip("Compartir seleccionados")
        self.btn_compartir_icon.clicked.connect(self.abrir_compartir)

        self.btn_borrar_icon = QPushButton()
        self.btn_borrar_icon.setIcon(QIcon("assets/buttons/borrar.png"))
        self.btn_borrar_icon.setIconSize(QSize(22, 22))
        self.btn_borrar_icon.setFixedSize(35, 35)
        self.btn_borrar_icon.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_borrar_icon.setStyleSheet("""
            QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
            QPushButton:hover { background-color: #eaf2ff; border-radius: 17px; }
            QPushButton:disabled { opacity: 0.3; }
        """)
        self.btn_borrar_icon.setEnabled(False)
        self.btn_borrar_icon.setToolTip("Borrar seleccionados")
        self.btn_borrar_icon.clicked.connect(self.borrar_reportes_seleccionados)


        header_layout.setContentsMargins(10, 10, 10, 0)
        self.btn_cerrar_x = QPushButton()
        self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))
        self.btn_cerrar_x.setIconSize(QSize(20, 20))
        self.btn_cerrar_x.setFixedSize(35, 35)
        self.btn_cerrar_x.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cerrar_x.setStyleSheet("""
            QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
            QPushButton:hover { background-color: #eaf2ff; border-radius: 17px; }
        """)
        self.btn_cerrar_x.enterEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xroja.png"))
        self.btn_cerrar_x.leaveEvent = lambda e: self.btn_cerrar_x.setIcon(QIcon("assets/buttons/xneg.png"))
        self.btn_cerrar_x.clicked.connect(self.reject)
        self.btn_cerrar_x.setToolTip("Cerrar")
        
        self.btn_compartir_icon.installEventFilter(self)
        self.btn_borrar_icon.installEventFilter(self)
        self.btn_cerrar_x.installEventFilter(self)
        
        header_layout.addSpacing(70) # Ajuste para centrar título
        header_layout.addStretch()
        header_layout.addWidget(lbl_titulo)
        header_layout.addStretch()
        header_layout.addWidget(self.btn_compartir_icon)
        header_layout.addWidget(self.btn_borrar_icon)
        header_layout.addWidget(self.btn_cerrar_x)
        
        flayout.addLayout(header_layout)

        # Tab Widget para separar reportes propios de compartidos
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::panel { border: 1px solid #d0d7de; border-radius: 0px; background-color: white; }
            QTabBar::tab { background: #f6f8fa; border: 1px solid #d0d7de; border-bottom: none; border-top-left-radius: 0px; border-top-right-radius: 0px; padding: 8px 16px; font-weight: bold; font-size: 11px; color: #57606a; }
            QTabBar::tab:selected { background: white; border-color: #d0d7de; border-bottom: 2px solid white; color: #0969da; }
            QTabBar::tab:hover { background: #eaf2ff; }
        """)
        
        # TAB 1: MIS REPORTES
        self.tab_propios = QWidget()
        layout_p = QVBoxLayout(self.tab_propios)
        layout_p.setContentsMargins(10, 10, 10, 10)
        
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["Reporte / Imagen", "Fecha de Creación", "Estado", "Detecciones", "Validación", "Descargar"])
        from PyQt6.QtWidgets import QHeaderView
        self.tree.header().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.tree.header().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.tree.setColumnWidth(1, 130)
        self.tree.setColumnWidth(2, 160)
        self.tree.setColumnWidth(3, 100)
        self.tree.setColumnWidth(4, 90)
        self.tree.setColumnWidth(5, 90)
        self.tree.header().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.tree.header().setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self.tree.header().setStretchLastSection(False)
        self.tree.setIndentation(20)
        self.tree.setAnimated(True)
        from PyQt6.QtWidgets import QAbstractItemView
        self.tree.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        self.tree.setStyleSheet("""
            QTreeWidget { border: none; background-color: #ffffff; alternate-background-color: #f6f8fa; font-size: 11px; outline: none; }
            QTreeWidget:focus { border: none; outline: none; }
            QTreeWidget::item { height: 32px; border-bottom: 1px solid #f0f0f0; color: #24292f; }
            QTreeWidget::item:selected { background-color: #eaf2ff; color: #24292f; }
            QHeaderView::section { background-color: #f6f8fa; padding: 6px; font-weight: bold; border: none; border-bottom: 2px solid #d0d7de; color: #57606a; font-size: 11px; }
        """)
        layout_p.addWidget(self.tree)
        if self.rol != "Tesista":
            self.tabs.addTab(self.tab_propios, "Mis Reportes")

        # TAB 2: COMPARTIDOS CONMIGO
        self.tab_compartidos = QWidget()
        layout_c = QVBoxLayout(self.tab_compartidos)
        layout_c.setContentsMargins(10, 10, 10, 10)
        
        self.tree_compartidos = QTreeWidget()
        self.tree_compartidos.setHeaderLabels(["Reporte / Imagen", "Propietario / Colaborador", "Fecha Compartido", "Estado de Seguimiento", "Detecciones", "Validación", "Descargar"])
        from PyQt6.QtWidgets import QHeaderView
        self.tree_compartidos.header().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.tree_compartidos.header().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.tree_compartidos.setColumnWidth(1, 150)
        self.tree_compartidos.setColumnWidth(2, 120)
        self.tree_compartidos.setColumnWidth(3, 150)
        self.tree_compartidos.setColumnWidth(4, 100)
        self.tree_compartidos.setColumnWidth(5, 90)
        self.tree_compartidos.setColumnWidth(6, 90)
        self.tree_compartidos.header().setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self.tree_compartidos.header().setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        self.tree_compartidos.header().setStretchLastSection(False)
        self.tree_compartidos.setIndentation(20)
        self.tree_compartidos.setAnimated(True)
        self.tree_compartidos.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.tree_compartidos.setStyleSheet("""
            QTreeWidget { border: none; background-color: #ffffff; alternate-background-color: #f6f8fa; font-size: 11px; outline: none; }
            QTreeWidget:focus { border: none; outline: none; }
            QTreeWidget::item { height: 32px; border-bottom: 1px solid #f0f0f0; color: #24292f; }
            QTreeWidget::item:selected { background-color: #eaf2ff; color: #24292f; }
            QHeaderView::section { background-color: #f6f8fa; padding: 6px; font-weight: bold; border: none; border-bottom: 2px solid #d0d7de; color: #57606a; font-size: 11px; }
        """)
        layout_c.addWidget(self.tree_compartidos)
        self.tabs.addTab(self.tab_compartidos, "Compartidos")
        
        flayout.addWidget(self.tabs)
        
        self.cargar_datos()
        
        self.tree.itemSelectionChanged.connect(self.actualizar_estado_boton_borrar)
        self.tree_compartidos.itemSelectionChanged.connect(self.actualizar_estado_boton_borrar)
        self.tabs.currentChanged.connect(lambda: self.actualizar_estado_boton_borrar())
        
        btn_layout = QHBoxLayout()
        
        self.btn_cargar = QPushButton("Cargar / Retomar")
        self.btn_cargar.setEnabled(False)
        self.btn_cargar.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cargar.setStyleSheet("""
            QPushButton { 
                background-color: transparent; 
                border: 2px solid #2da44e; 
                color: #2da44e; 
                border-radius: 0px; 
                font-weight: bold; 
                padding: 10px 20px; 
            }
            QPushButton:hover { 
                background-color: #2da44e; 
                color: white; 
            }
            QPushButton:disabled { 
                background-color: transparent; 
                border: 2px solid #d0d7de; 
                color: #8c959f; 
            }
        """)
        self.btn_cargar.clicked.connect(self.aceptar_seleccion)
        
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_cargar)
        
        flayout.addLayout(btn_layout);        main_layout.addWidget(self.frame)
        self.actualizar_estado_boton_borrar() # Asegurar estado inicial
        if self.rol == "Tesista":
            self.tabs.setCurrentIndex(0)
        self.resize(950, 650)

    def eventFilter(self, obj, event):
        from PyQt6.QtCore import QEvent, QPoint
        from PyQt6.QtWidgets import QToolTip, QPushButton
        if event.type() == QEvent.Type.ToolTip:
            if isinstance(obj, QPushButton) and obj.toolTip():
                # Obtener la posición global de la esquina inferior del botón
                global_pos = obj.mapToGlobal(obj.rect().bottomLeft())
                # Mostrar el tooltip desplazado a la izquierda para garantizar visualización
                custom_x = global_pos.x() - 60
                custom_y = global_pos.y() - 10
                QToolTip.showText(QPoint(custom_x, custom_y), obj.toolTip(), obj)
                return True
        return super().eventFilter(obj, event)

    def abrir_compartir(self):
        tab_index = self.tabs.currentIndex()
        if tab_index != 0: return
        
        selected_items = self.tree.selectedItems()
        selected_reports = [item for item in selected_items if item.data(0, Qt.ItemDataRole.UserRole) and item.data(0, Qt.ItemDataRole.UserRole).get("type") == "reporte"]
        
        if not selected_reports: return
        
        # Obtener todos los IDs de reportes seleccionados
        id_reps = [item.data(0, Qt.ItemDataRole.UserRole)["id"] for item in selected_reports]
        diag = DialogoCompartirReporte(self.id_usuario, id_reps, self)
        diag.exec()
        self.cargar_datos()

    def validar_reporte(self, id_reporte):
        """Inicia el flujo de revisión: cierra el historial y carga el reporte para inspección."""
        # La validación real (DB) ocurre en VentanaBaseAnalisis.confirmar_validacion()
        # Aquí solo cerramos el diálogo pasando el flag ir_a_metricas=True
        self.seleccion = {"type": "reporte", "id_reporte": id_reporte, "ir_a_metricas": True}
        self.accept()

    def cargar_para_corregir(self, id_reporte):
        """Permite al tesista cargar su trabajo que requiere corrección directamente."""
        self.seleccion = {"type": "reporte", "id_reporte": id_reporte}
        self.accept()

    def descargar_reporte_id(self, id_reporte):
        from bd.database import conectar
        import json
        import os
        conn = conectar(); cur = conn.cursor()
        try:
            cur.execute("""
                SELECT A.datos_persistentes, I.ruta_archivo 
                FROM Analisis A 
                JOIN Imagen I ON A.id_imagen = I.id_imagen 
                WHERE A.id_reporte = ?
            """, (id_reporte,))
            rows = cur.fetchall()
            all_metrics = []
            for r in rows:
                datos_json, ruta_img = r
                if datos_json:
                    d = json.loads(datos_json)
                    m = d.get("metricas_acumuladas", [])
                    nombre_img = os.path.basename(ruta_img) if ruta_img else "Imagen Sin Nombre"
                    if isinstance(m, list):
                        for item in m:
                            if isinstance(item, dict) and "nombre_imagen" not in item:
                                item["nombre_imagen"] = nombre_img
                        all_metrics.extend(m)
                    elif isinstance(m, dict):
                        if "nombre_imagen" not in m:
                            m["nombre_imagen"] = nombre_img
                        all_metrics.append(m)
            
            if not all_metrics:
                from vistas.utilidades import DialogoNotificacion
                DialogoNotificacion("Aviso", "Este reporte no tiene métricas completadas para descargar.", "info", self).exec()
                return
                
            # Usar la lógica global de descarga
            original_metrics = self.parent().metricas_reporte
            self.parent().metricas_reporte = all_metrics
            self.parent().descargar_reporte()
            self.parent().metricas_reporte = original_metrics
            
        except Exception as e:
            from vistas.utilidades import DialogoNotificacion
            DialogoNotificacion("Error", f"Error al descargar: {e}", "error", self).exec()
        finally:
            conn.close()

    def actualizar_estado_boton_borrar(self):
        tab_index = self.tabs.currentIndex()
        if self.rol == "Tesista":
            tab_index = 1
        if tab_index == 0:
            selected_items = self.tree.selectedItems()
            selected_reports = [item for item in selected_items if item.data(0, Qt.ItemDataRole.UserRole) and item.data(0, Qt.ItemDataRole.UserRole).get("type") == "reporte"]
            num_seleccionados = len(selected_reports)
            self.btn_borrar_icon.setEnabled(num_seleccionados > 0)
            self.btn_compartir_icon.setEnabled(self.rol == "Investigador" and num_seleccionados > 0)
            # Deshabilitar Cargar si el reporte tiene validación pendiente (botón Validar activo)
            if num_seleccionados == 1:
                data = selected_reports[0].data(0, Qt.ItemDataRole.UserRole)
                estado_compartido = data.get("estado_compartido", "")
                is_completed = data.get("reporte_completo", False)
                tiene_comentarios = data.get("tiene_comentarios", False)
                if self.rol == "Tesista":
                    if is_completed:
                        if estado_compartido == "Pendiente" and not tiene_comentarios:
                            self.btn_cargar.setEnabled(False)
                        else:
                            self.btn_cargar.setEnabled(True)
                    else:
                        self.btn_cargar.setEnabled(True)
                else:
                    self.btn_cargar.setEnabled(not debe_bloquear_carga(data, self.id_usuario))
            else:
                self.btn_cargar.setEnabled(False)
        else:
            selected_items = self.tree_compartidos.selectedItems()
            selected_reports = [item for item in selected_items if item.data(0, Qt.ItemDataRole.UserRole) and item.data(0, Qt.ItemDataRole.UserRole).get("type") == "reporte"]
            num_seleccionados = len(selected_reports)
            self.btn_borrar_icon.setEnabled(False) # No se pueden borrar reportes compartidos
            self.btn_compartir_icon.setEnabled(False)
            # Deshabilitar Cargar si el reporte seleccionado está pendiente de validación (Modificado)
            if num_seleccionados == 1:
                data = selected_reports[0].data(0, Qt.ItemDataRole.UserRole)
                estado_compartido = data.get("estado_compartido", "")
                is_completed = data.get("reporte_completo", False)
                tiene_comentarios = data.get("tiene_comentarios", False)
                if self.rol == "Tesista":
                    if is_completed:
                        if estado_compartido == "Pendiente" and not tiene_comentarios:
                            self.btn_cargar.setEnabled(False)
                        else:
                            self.btn_cargar.setEnabled(True)
                    else:
                        self.btn_cargar.setEnabled(True)
                else:
                    self.btn_cargar.setEnabled(not debe_bloquear_carga(data, self.id_usuario))
            else:
                self.btn_cargar.setEnabled(False)

    def showEvent(self, event):
        super().showEvent(event)
        if self.parent():
            p_geom = self.parent().geometry()
            self.move(p_geom.x() + (p_geom.width() - self.width()) // 2, p_geom.y() + (p_geom.height() - self.height()) // 2)

    def mousePressEvent(self, event):
        from PyQt6.QtCore import Qt
        if event.button() == Qt.MouseButton.LeftButton:
            if self.windowHandle():
                self.windowHandle().startSystemMove()
            else:
                self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            event.accept()
        else:
            super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        from PyQt6.QtCore import Qt
        if not self.windowHandle() and hasattr(self, "_drag_pos") and event.buttons() == Qt.MouseButton.LeftButton:
            self.move(event.globalPosition().toPoint() - self._drag_pos)
            event.accept()
        else:
            super().mouseMoveEvent(event)

    def cargar_datos(self):
        self.tree.clear()
        self.tree_compartidos.clear()
        
        from bd.database import conectar
        conn = conectar(); cur = conn.cursor()
        try:
            # 1. Cargar Mis Reportes
            cur.execute("SELECT id_reporte, nombre_reporte, fecha_creacion, estado FROM Reporte WHERE id_usuario = ? ORDER BY fecha_creacion DESC", (self.id_usuario,))
            reportes = cur.fetchall()
            for rep in reportes:
                id_rep, nombre, fecha, estado = rep
                
                # Sincronizar estado general si todos los análisis internos están completados
                cur.execute("SELECT paso_actual FROM Analisis WHERE id_reporte = ?", (id_rep,))
                analisis_pasos = cur.fetchall()
                total_analisis = len(analisis_pasos)
                completados = sum(1 for an in analisis_pasos if an[0] >= 5)
                reporte_completo = (total_analisis > 0 and completados == total_analisis)
                
                if analisis_pasos:
                    todos_completados = reporte_completo
                    if todos_completados and estado == 'En progreso':
                        cur.execute("UPDATE Reporte SET estado = 'Completado' WHERE id_reporte = ?", (id_rep,))
                        conn.commit()
                        estado = 'Completado'
                    elif not todos_completados and estado == 'Completado':
                        cur.execute("UPDATE Reporte SET estado = 'En progreso' WHERE id_reporte = ?", (id_rep,))
                        conn.commit()
                        estado = 'En progreso'
                
                # Verificar si está compartido y si fue modificado (seleccionando también comentarios)
                cur.execute("""
                    SELECT RC.estado, U.nombre_usuario, RC.comentarios 
                    FROM ReporteCompartido RC 
                    JOIN Usuario U ON RC.id_destinatario = U.id_usuario
                    WHERE RC.id_reporte = ?
                """, (id_rep,))
                share_info = cur.fetchone()
                
                estado_texto = estado
                es_modificado = False
                es_pendiente = False
                tiene_comentarios = False
                if share_info:
                    sh_estado, sh_user, sh_comentarios = share_info
                    tiene_comentarios = bool(sh_comentarios and sh_comentarios.strip())
                    if sh_estado == 'Modificado':
                        estado_texto = f"Modificado (por {sh_user})"
                        es_modificado = True
                    elif sh_estado == 'Pendiente':
                        estado_texto = f"Pendiente de trabajo (por {sh_user})"
                        es_pendiente = True
                    elif sh_estado == 'Validado':
                        estado_texto = f"Validado ({sh_user})"
                
                if self.rol == "Tesista":
                    if share_info and tiene_comentarios:
                        estado_texto = "Pendiente de trabajo"
                    else:
                        estado_texto = "Completado" if reporte_completo else "Incompleto"
                
                # Obtener total de detecciones para este reporte
                cur.execute("""
                    SELECT SUM(A.cantidad_microglias) FROM Analisis A
                    WHERE A.id_reporte = ?
                """, (id_rep,))
                total_det = cur.fetchone()[0] or 0

                rep_item = QTreeWidgetItem(self.tree, [nombre, str(fecha), estado_texto, "", "", ""])
                font_bold = QFont()
                font_bold.setBold(True)
                rep_item.setFont(0, font_bold)
                rep_item.setFont(1, font_bold)
                rep_item.setFont(2, font_bold)
                sh_estado_val = share_info[0] if share_info else ""
                rep_item.setData(0, Qt.ItemDataRole.UserRole, {
                    "type": "reporte", "id": id_rep,
                    "estado_compartido": sh_estado_val,
                    "id_prop": self.id_usuario,  # En Mis Reportes siempre soy el propietario
                    "reporte_completo": reporte_completo,
                    "tiene_comentarios": tiene_comentarios
                })
                
                cur.execute("""
                    SELECT A.id_analisis, I.ruta_archivo, A.fecha_analisis, A.paso_actual, A.cantidad_microglias, A.datos_persistentes
                    FROM Analisis A JOIN Imagen I ON A.id_imagen = I.id_imagen
                    WHERE A.id_reporte = ? ORDER BY A.fecha_analisis ASC
                """, (id_rep,))
                análisis = cur.fetchall()
                for an in análisis:
                    id_an, ruta, f_an, paso, cant, dp = an
                    if not cant or cant == 0:
                        # Fallback 1: Contar registros en Microglia
                        cur.execute("SELECT COUNT(*) FROM Microglia WHERE id_analisis = ?", (id_an,))
                        cant = cur.fetchone()[0] or 0
                        if cant == 0 and dp:
                            # Fallback 2: Contar boxes en datos_persistentes JSON
                            try:
                                import json
                                datos = json.loads(dp)
                                cant = len(datos.get("boxes", []))
                            except:
                                pass
                    st = "Completado" if paso >= 5 else f"En proceso ({paso}/4)"
                    an_item = QTreeWidgetItem(rep_item, [os.path.basename(ruta), str(f_an), st, str(cant), "", ""])
                    an_item.setData(0, Qt.ItemDataRole.UserRole, {"type": "analisis", "id": id_an, "id_reporte": id_rep})
                    an_item.setDisabled(True)
                rep_item.setExpanded(True)
                
                # reporte_completo, total_analisis, completados ya calculados arriba
                
                # Botón de Validar o Icono de Espera (si aplica)
                if self.rol == "Investigador" and share_info:
                    if es_modificado:
                        btn_valid = QPushButton()
                        btn_valid.setIcon(QIcon("assets/buttons/validar.png"))
                        btn_valid.setIconSize(QSize(18, 18))
                        btn_valid.setFixedSize(30, 30)
                        btn_valid.setCursor(Qt.CursorShape.PointingHandCursor)
                        btn_valid.setStyleSheet("""
                            QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
                            QPushButton:hover { background-color: #eaf2ff; border-radius: 15px; }
                            QPushButton:disabled { opacity: 0.1; }
                        """)
                        btn_valid.setToolTip("Validar reporte")
                        btn_valid.installEventFilter(self)
                        btn_valid.clicked.connect(lambda checked, r_id=id_rep: self.validar_reporte(r_id))
                        
                        container_val = QWidget()
                        layout_v = QHBoxLayout(container_val)
                        layout_v.addWidget(btn_valid)
                        layout_v.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        layout_v.setContentsMargins(0, 0, 0, 0)
                        self.tree.setItemWidget(rep_item, 4, container_val)
                    elif es_pendiente:
                        lbl_esperar = QLabel()
                        lbl_esperar.setStyleSheet("background-color: transparent; border: none;")
                        dpr = self.devicePixelRatioF()
                        pix_size = int(18 * dpr)
                        pixmap = QPixmap("assets/buttons/esperar.png").scaled(
                            pix_size, pix_size,
                            Qt.AspectRatioMode.KeepAspectRatio,
                            Qt.TransformationMode.SmoothTransformation
                        )
                        pixmap.setDevicePixelRatio(dpr)
                        lbl_esperar.setPixmap(pixmap)
                        lbl_esperar.setToolTip("Esperando a que el colaborador realice cambios")
                        lbl_esperar.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        
                        container_esp = QWidget()
                        container_esp.setStyleSheet("background-color: transparent; border: none;")
                        layout_esp = QHBoxLayout(container_esp)
                        layout_esp.addWidget(lbl_esperar)
                        layout_esp.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        layout_esp.setContentsMargins(0, 0, 0, 0)
                        self.tree.setItemWidget(rep_item, 4, container_esp)
                elif self.rol == "Tesista":
                    if share_info and tiene_comentarios:
                        btn_msg = QPushButton()
                        btn_msg.setIcon(QIcon("assets/buttons/msg.png"))
                        btn_msg.setIconSize(QSize(18, 18))
                        btn_msg.setFixedSize(30, 30)
                        btn_msg.setCursor(Qt.CursorShape.PointingHandCursor)
                        btn_msg.setStyleSheet("""
                            QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
                            QPushButton:hover { background-color: #eaf2ff; border-radius: 15px; }
                            QPushButton:disabled { opacity: 0.1; }
                        """)
                        btn_msg.setToolTip("Cargar trabajo para corregir observaciones")
                        btn_msg.installEventFilter(self)
                        btn_msg.clicked.connect(lambda checked, r_id=id_rep: self.cargar_para_corregir(r_id))
                        
                        container_msg = QWidget()
                        container_msg.setStyleSheet("background-color: transparent; border: none;")
                        layout_m = QHBoxLayout(container_msg)
                        layout_m.addWidget(btn_msg)
                        layout_m.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        layout_m.setContentsMargins(0, 0, 0, 0)
                        self.tree.setItemWidget(rep_item, 4, container_msg)
                    elif not reporte_completo:
                        btn_pend = QPushButton()
                        btn_pend.setIcon(QIcon("assets/buttons/pendiente.png"))
                        btn_pend.setIconSize(QSize(18, 18))
                        btn_pend.setFixedSize(30, 30)
                        btn_pend.setCursor(Qt.CursorShape.PointingHandCursor)
                        btn_pend.setStyleSheet("""
                            QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
                            QPushButton:hover { background-color: #eaf2ff; border-radius: 15px; }
                        """)
                        btn_pend.setToolTip("Análisis pendiente (imágenes incompletas)")
                        btn_pend.installEventFilter(self)
                        btn_pend.clicked.connect(lambda checked, r_id=id_rep: self.cargar_para_corregir(r_id))
                        
                        container_pend = QWidget()
                        container_pend.setStyleSheet("background-color: transparent; border: none;")
                        layout_p = QHBoxLayout(container_pend)
                        layout_p.addWidget(btn_pend)
                        layout_p.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        layout_p.setContentsMargins(0, 0, 0, 0)
                        self.tree.setItemWidget(rep_item, 4, container_pend)
                
                # Botón de Descargar
                btn_dl = QPushButton()
                btn_dl.setIcon(QIcon("assets/buttons/download.png"))
                btn_dl.setIconSize(QSize(18, 18))
                btn_dl.setFixedSize(30, 30)
                btn_dl.setCursor(Qt.CursorShape.PointingHandCursor)
                btn_dl.setStyleSheet("""
                    QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
                    QPushButton:hover { background-color: #eaf2ff; border-radius: 15px; }
                    QPushButton:disabled { opacity: 0.1; }
                """)
                if reporte_completo:
                    btn_dl.setEnabled(True)
                    btn_dl.setToolTip("Descargar reporte")
                else:
                    btn_dl.setEnabled(False)
                    btn_dl.setToolTip(f"Incompleto ({completados}/{total_analisis})")
                btn_dl.installEventFilter(self)
                btn_dl.clicked.connect(lambda checked, r_id=id_rep: self.descargar_reporte_id(r_id))
                
                container_dl = QWidget()
                layout_d = QHBoxLayout(container_dl)
                layout_d.addWidget(btn_dl)
                layout_d.setAlignment(Qt.AlignmentFlag.AlignCenter)
                layout_d.setContentsMargins(0, 0, 0, 0)
                self.tree.setItemWidget(rep_item, 5, container_dl)
                
            # 2. Cargar Reportes Compartidos (Enviados y Recibidos) (seleccionando también comentarios)
            cur.execute("""
                SELECT 
                    RC.id_reporte, 
                    R.nombre_reporte, 
                    RC.fecha_compartido, 
                    RC.estado, 
                    U_prop.nombre_usuario AS propietario_name,
                    U_dest.nombre_usuario AS destinatario_name,
                    RC.id_propietario,
                    RC.id_destinatario,
                    RC.comentarios
                FROM ReporteCompartido RC
                JOIN Reporte R ON RC.id_reporte = R.id_reporte
                JOIN Usuario U_prop ON RC.id_propietario = U_prop.id_usuario
                JOIN Usuario U_dest ON RC.id_destinatario = U_dest.id_usuario
                WHERE RC.id_destinatario = ? OR RC.id_propietario = ?
                ORDER BY RC.fecha_compartido DESC
            """, (self.id_usuario, self.id_usuario))
            compartidos = cur.fetchall()
            for comp in compartidos:
                id_rep, nombre, fecha, estado_compartido, propietario_name, destinatario_name, id_prop, id_dest, comentarios_compartido = comp
                
                # Verificar si TODAS las imágenes del reporte compartido están completadas
                cur.execute("SELECT COUNT(*) FROM Analisis WHERE id_reporte = ?", (id_rep,))
                total_analisis = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM Analisis WHERE id_reporte = ? AND paso_actual >= 5", (id_rep,))
                completados = cur.fetchone()[0]
                reporte_completo = (total_analisis > 0 and completados == total_analisis)

                if id_prop == self.id_usuario:
                    rol_text = f"Para: {destinatario_name}"
                else:
                    rol_text = f"De: {propietario_name}"

                if self.rol == "Tesista":
                    if estado_compartido == 'Pendiente' and comentarios_compartido and comentarios_compartido.strip():
                        estado_texto = "Pendiente de trabajo"
                    else:
                        estado_texto = "Completado" if reporte_completo else "Incompleto"
                else:
                    if id_prop == self.id_usuario:
                        # Reporte que YO compartí con alguien más
                        if estado_compartido == 'Pendiente':
                            estado_texto = "Pendiente de trabajo"
                        elif estado_compartido == 'Modificado':
                            estado_texto = f"Modificado (por {destinatario_name})"
                        else:
                            estado_texto = "Validado (por ti)"
                    else:
                        # Reporte que compartieron CONMIGO
                        if estado_compartido == 'Pendiente':
                            estado_texto = "Pendiente de trabajo"
                        elif estado_compartido == 'Modificado':
                            estado_texto = "Modificado por ti"
                        else:
                            estado_texto = "Validado por Investigador"
                
                # Obtener total de detecciones para este reporte
                cur.execute("""
                    SELECT SUM(A.cantidad_microglias) FROM Analisis A
                    WHERE A.id_reporte = ?
                """, (id_rep,))
                total_det = cur.fetchone()[0] or 0

                rep_item = QTreeWidgetItem(self.tree_compartidos, [nombre, rol_text, str(fecha), estado_texto, "", "", ""])
                font_bold = QFont()
                font_bold.setBold(True)
                rep_item.setFont(0, font_bold)
                rep_item.setFont(1, font_bold)
                rep_item.setFont(2, font_bold)
                rep_item.setFont(3, font_bold)
                rep_item.setData(0, Qt.ItemDataRole.UserRole, {"type": "reporte", "id": id_rep})
                
                cur.execute("""
                    SELECT A.id_analisis, I.ruta_archivo, A.fecha_analisis, A.paso_actual, A.cantidad_microglias, A.datos_persistentes
                    FROM Analisis A JOIN Imagen I ON A.id_imagen = I.id_imagen
                    WHERE A.id_reporte = ? ORDER BY A.fecha_analisis ASC
                """, (id_rep,))
                análisis = cur.fetchall()
                for an in análisis:
                    id_an, ruta, f_an, paso, cant, dp = an
                    if not cant or cant == 0:
                        # Fallback 1: Contar registros en Microglia
                        cur.execute("SELECT COUNT(*) FROM Microglia WHERE id_analisis = ?", (id_an,))
                        cant = cur.fetchone()[0] or 0
                        if cant == 0 and dp:
                            # Fallback 2: Contar boxes en datos_persistentes JSON
                            try:
                                import json
                                datos = json.loads(dp)
                                cant = len(datos.get("boxes", []))
                            except:
                                pass
                    st = "Completado" if paso >= 5 else f"En proceso ({paso}/4)"
                    an_item = QTreeWidgetItem(rep_item, [os.path.basename(ruta), "", str(f_an), st, str(cant), "", ""])
                    an_item.setData(0, Qt.ItemDataRole.UserRole, {"type": "analisis", "id": id_an, "id_reporte": id_rep})
                    an_item.setDisabled(True)
                rep_item.setExpanded(True)
                
                # Guardar estado en item para uso en actualizar_estado_boton_borrar
                tiene_comentarios_comp = bool(comentarios_compartido and comentarios_compartido.strip())
                rep_item.setData(0, Qt.ItemDataRole.UserRole, {
                    "type": "reporte", "id": id_rep,
                    "estado_compartido": estado_compartido,
                    "id_prop": id_prop, "id_dest": id_dest,
                    "reporte_completo": reporte_completo,
                    "tiene_comentarios": tiene_comentarios_comp
                })

                # Botón de Validar o Icono de Espera / Pendiente
                if self.rol == "Tesista":
                    if estado_compartido == 'Pendiente' and comentarios_compartido and comentarios_compartido.strip():
                        btn_msg = QPushButton()
                        btn_msg.setIcon(QIcon("assets/buttons/msg.png"))
                        btn_msg.setIconSize(QSize(18, 18))
                        btn_msg.setFixedSize(30, 30)
                        btn_msg.setCursor(Qt.CursorShape.PointingHandCursor)
                        btn_msg.setStyleSheet("""
                            QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
                            QPushButton:hover { background-color: #eaf2ff; border-radius: 15px; }
                            QPushButton:disabled { opacity: 0.1; }
                        """)
                        btn_msg.setToolTip("Cargar trabajo para corregir observaciones")
                        btn_msg.installEventFilter(self)
                        btn_msg.clicked.connect(lambda checked, r_id=id_rep: self.cargar_para_corregir(r_id))
                        
                        container_msg = QWidget()
                        container_msg.setStyleSheet("background-color: transparent; border: none;")
                        layout_m = QHBoxLayout(container_msg)
                        layout_m.addWidget(btn_msg)
                        layout_m.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        layout_m.setContentsMargins(0, 0, 0, 0)
                        self.tree_compartidos.setItemWidget(rep_item, 5, container_msg)
                    elif not reporte_completo:
                        btn_pend = QPushButton()
                        btn_pend.setIcon(QIcon("assets/buttons/pendiente.png"))
                        btn_pend.setIconSize(QSize(18, 18))
                        btn_pend.setFixedSize(30, 30)
                        btn_pend.setCursor(Qt.CursorShape.PointingHandCursor)
                        btn_pend.setStyleSheet("""
                            QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
                            QPushButton:hover { background-color: #eaf2ff; border-radius: 15px; }
                        """)
                        btn_pend.setToolTip("Análisis pendiente")
                        btn_pend.installEventFilter(self)
                        btn_pend.clicked.connect(lambda checked, r_id=id_rep: self.cargar_para_corregir(r_id))
                        
                        container_pend = QWidget()
                        container_pend.setStyleSheet("background-color: transparent; border: none;")
                        layout_p = QHBoxLayout(container_pend)
                        layout_p.addWidget(btn_pend)
                        layout_p.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        layout_p.setContentsMargins(0, 0, 0, 0)
                        self.tree_compartidos.setItemWidget(rep_item, 5, container_pend)
                    elif estado_compartido == 'Pendiente':
                        lbl_esperar = QLabel()
                        lbl_esperar.setStyleSheet("background-color: transparent; border: none;")
                        dpr = self.devicePixelRatioF()
                        pix_size = int(18 * dpr)
                        pixmap = QPixmap("assets/buttons/esperar.png").scaled(
                            pix_size, pix_size,
                            Qt.AspectRatioMode.KeepAspectRatio,
                            Qt.TransformationMode.SmoothTransformation
                        )
                        pixmap.setDevicePixelRatio(dpr)
                        lbl_esperar.setPixmap(pixmap)
                        lbl_esperar.setToolTip("Esperando validación del investigador")
                        lbl_esperar.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        
                        container_esp = QWidget()
                        container_esp.setStyleSheet("background-color: transparent; border: none;")
                        layout_esp = QHBoxLayout(container_esp)
                        layout_esp.addWidget(lbl_esperar)
                        layout_esp.setAlignment(Qt.AlignmentFlag.AlignCenter)
                        layout_esp.setContentsMargins(0, 0, 0, 0)
                        self.tree_compartidos.setItemWidget(rep_item, 5, container_esp)
                else:
                    if id_prop == self.id_usuario:
                        if estado_compartido == 'Modificado':
                            construir_boton_validar(
                                self.tree_compartidos, rep_item, id_rep,
                                reporte_completo, total_analisis, completados,
                                self.validar_reporte
                            )
                        elif estado_compartido == 'Pendiente':
                            lbl_esperar = QLabel()
                            lbl_esperar.setStyleSheet("background-color: transparent; border: none;")
                            dpr = self.devicePixelRatioF()
                            pix_size = int(18 * dpr)
                            pixmap = QPixmap("assets/buttons/esperar.png").scaled(
                                pix_size, pix_size,
                                Qt.AspectRatioMode.KeepAspectRatio,
                                Qt.TransformationMode.SmoothTransformation
                            )
                            pixmap.setDevicePixelRatio(dpr)
                            lbl_esperar.setPixmap(pixmap)
                            lbl_esperar.setToolTip("Esperando a que el colaborador realice cambios")
                            lbl_esperar.setAlignment(Qt.AlignmentFlag.AlignCenter)
                            
                            container_esp = QWidget()
                            container_esp.setStyleSheet("background-color: transparent; border: none;")
                            layout_esp = QHBoxLayout(container_esp)
                            layout_esp.addWidget(lbl_esperar)
                            layout_esp.setAlignment(Qt.AlignmentFlag.AlignCenter)
                            layout_esp.setContentsMargins(0, 0, 0, 0)
                            self.tree_compartidos.setItemWidget(rep_item, 5, container_esp)
                    elif id_dest == self.id_usuario:
                        if estado_compartido == 'Pendiente':
                            btn_msg = QPushButton()
                            btn_msg.setIcon(QIcon("assets/buttons/msg.png"))
                            btn_msg.setIconSize(QSize(18, 18))
                            btn_msg.setFixedSize(30, 30)
                            btn_msg.setCursor(Qt.CursorShape.PointingHandCursor)
                            btn_msg.setStyleSheet("""
                                QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
                                QPushButton:hover { background-color: #eaf2ff; border-radius: 15px; }
                                QPushButton:disabled { opacity: 0.1; }
                            """)
                            btn_msg.setToolTip("Cargar trabajo para corregir observaciones")
                            btn_msg.installEventFilter(self)
                            btn_msg.clicked.connect(lambda checked, r_id=id_rep: self.cargar_para_corregir(r_id))
                            
                            container_msg = QWidget()
                            container_msg.setStyleSheet("background-color: transparent; border: none;")
                            layout_m = QHBoxLayout(container_msg)
                            layout_m.addWidget(btn_msg)
                            layout_m.setAlignment(Qt.AlignmentFlag.AlignCenter)
                            layout_m.setContentsMargins(0, 0, 0, 0)
                            self.tree_compartidos.setItemWidget(rep_item, 5, container_msg)
                
                # Botón de Descargar
                if id_prop == self.id_usuario:
                    if estado_compartido == 'Validado' and reporte_completo:
                        dl_hab, dl_tip = True, "Descargar reporte"
                    elif estado_compartido == 'Modificado':
                        dl_hab, dl_tip = False, "Valida el reporte primero para poder descargar"
                    else:
                        dl_hab = reporte_completo
                        dl_tip = "Descargar reporte" if reporte_completo else f"Reporte incompleto ({completados}/{total_analisis} imágenes listas)"
                else:
                    dl_hab = reporte_completo
                    dl_tip = "Descargar reporte" if reporte_completo else f"Reporte incompleto ({completados}/{total_analisis} imágenes listas)"
                construir_boton_descargar(
                    self.tree_compartidos, rep_item, 6, id_rep,
                    dl_hab, dl_tip, self.descargar_reporte_id
                )
                
        except Exception as e: 
            logging.error(f"Error historial: {e}")
        finally: 
            conn.close()

    def aceptar_seleccion(self):
        tab_index = self.tabs.currentIndex()
        if self.rol == "Tesista":
            tab_index = 1
        if tab_index == 0:
            selected_items = self.tree.selectedItems()
        else:
            selected_items = self.tree_compartidos.selectedItems()
            
        selected_reports = [item for item in selected_items if item.data(0, Qt.ItemDataRole.UserRole) and item.data(0, Qt.ItemDataRole.UserRole).get("type") == "reporte"]
        
        if not selected_reports: return
        
        item_seleccionado = selected_reports[0]
        data = item_seleccionado.data(0, Qt.ItemDataRole.UserRole)

        # Bloquear carga si el propietario intenta cargar un reporte que aún debe validar
        if tab_index == 1:
            if debe_bloquear_carga(data, self.id_usuario):
                from vistas.utilidades import DialogoNotificacion
                DialogoNotificacion(
                    "Validación Requerida",
                    "El tesista ya completó el trabajo. Debes validar el reporte antes de poder cargarlo o descargarlo.",
                    "warning", self
                ).exec()
                return

        self.seleccion = {"type": "reporte", "id_reporte": data["id"], "estado": item_seleccionado.text(2)}
        self.accept()

    def borrar_reportes_seleccionados(self):
        # Recopilar todos los reportes seleccionados de la pestaña de Mis Reportes únicamente
        selected_items = self.tree.selectedItems()
        items_a_borrar = [item for item in selected_items if item.data(0, Qt.ItemDataRole.UserRole) and item.data(0, Qt.ItemDataRole.UserRole).get("type") == "reporte"]
        
        if not items_a_borrar:
            from vistas.utilidades import DialogoNotificacion
            DialogoNotificacion("Atención", "Selecciona al menos un reporte de tu lista personal para borrar.", "warning", self).exec()
            return
            
        from vistas.utilidades import DialogoConfirmacion
        msg = f"¿Estás seguro de borrar {len(items_a_borrar)} reporte(s) seleccionado(s) y todos sus datos asociados?"
        if not DialogoConfirmacion("Borrar Selección", msg).exec(): return
        
        from bd.database import conectar
        conn = conectar(); cur = conn.cursor()
        try:
            for item in items_a_borrar:
                data = item.data(0, Qt.ItemDataRole.UserRole)
                id_rep = data["id"]
                cur.execute("DELETE FROM ReporteCompartido WHERE id_reporte = ?", (id_rep,))
                cur.execute("DELETE FROM Microglia WHERE id_analisis IN (SELECT id_analisis FROM Analisis WHERE id_reporte = ?)", (id_rep,))
                cur.execute("DELETE FROM Analisis WHERE id_reporte = ?", (id_rep,))
                cur.execute("DELETE FROM Reporte WHERE id_reporte = ?", (id_rep,))
                self.tree.takeTopLevelItem(self.tree.indexOfTopLevelItem(item))
            conn.commit()
        except Exception as e: 
            logging.error(f"Error borrado masivo: {e}")
        finally: 
            conn.close()
            self.actualizar_estado_boton_borrar()


class VentanaBaseAnalisis(ValidacionReporteMixin, QMainWindow):
    def mostrar_notificacion(self, titulo, mensaje, tipo="info"):
        from vistas.utilidades import DialogoNotificacion
        DialogoNotificacion(titulo, mensaje, tipo, self).exec()

    def __init__(self, id_usuario, rol, nombre_usuario=""):
        super().__init__()
        self.id_usuario = id_usuario; self.rol = rol; self.nombre_usuario = nombre_usuario
        self.ruta_imagen_actual = None
        self.pixmaps_globales = {"Original": None, "Filtrada": None, "Esqueleto": None}
        self.crops_en_memoria = {}
        self.crops_filtrados_temp = {}
        self.metadatos_imagen = {"campo": "", "tiempo": ""}
        self.metricas_reporte = []
        self.reporte_finalizado_actual = False
        self.id_reporte_actual = None
        self.id_analisis_actual = None
        self.paso_actual = 0
        self.metricas_extraidas_ciclo_actual = False
        self._init_validacion_estado()  # Inicializa reporte_validado_cargado = False
        self.setWindowTitle(f"Prototipo Microglías - Panel ({self.rol})")
        
        from vistas.utilidades import set_app_icon
        set_app_icon(self)
        
        screen_geom = QApplication.primaryScreen().geometry()
        self.resize(int(screen_geom.width() * 0.8), int(screen_geom.height() * 0.8))
        self.setMinimumSize(1050, 700)
        self.setStyleSheet("""
            QToolTip {
                background-color: #24292f;
                color: #ffffff;
                border: 1px solid #24292f;
                border-radius: 4px;
                padding: 3px 6px;
                font-size: 10px;
            }
        """)
        self.inicializar_ui()

    def inicializar_ui(self):
        widget_central = QWidget(); layout_principal = QHBoxLayout()
        self.menu_lateral = QVBoxLayout()
        
        # Layout superior del menú lateral para el botón de historial (≡) - Absolute Top
        layout_historial_top = QHBoxLayout()
        layout_historial_top.setContentsMargins(0, 0, 0, 0)
        self.btn_historial = QPushButton()
        self.btn_historial.setIcon(QIcon("assets/buttons/historial.png"))
        self.btn_historial.setIconSize(QSize(22, 22))
        self.btn_historial.setFixedSize(35, 35)
        self.btn_historial.setToolTip("Historial")
        self.btn_historial.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_historial.setStyleSheet("""
            QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
            QPushButton:hover { background-color: #eaf2ff; border-radius: 17px; }
        """)
        
        # Información del usuario al lado derecho
        layout_user_info = QVBoxLayout()
        lbl_rango = QLabel(self.rol)
        lbl_rango.setStyleSheet("color: #007bff; font-weight: bold; font-size: 13px; margin-left: 2px; padding: 0;")
        layout_user_info.addWidget(lbl_rango)

        if self.rol not in ["Invitado", "Guest"]:
            lbl_nombre = QLabel(self.nombre_usuario)
            lbl_nombre.setStyleSheet("color: #666; font-size: 11px; margin-left: 2px; padding: 0;")
            layout_user_info.addWidget(lbl_nombre)
        
        layout_user_info.setSpacing(0)
        
        layout_historial_top.addWidget(self.btn_historial)
        layout_historial_top.addLayout(layout_user_info)
        layout_historial_top.addStretch()
        self.menu_lateral.addLayout(layout_historial_top)
        self.menu_lateral.addSpacing(10)

        self.btn_cargar = QPushButton("Cargar Imagen"); label_caract = QLabel("Caracterización:"); label_caract.setStyleSheet("font-weight: bold; margin-top: 15px;")
        self.btn_conteo = QPushButton("Detectar Microglías"); self.btn_filtrar = QPushButton("Filtrar Imagen"); self.btn_ramas = QPushButton("Mostrar Ramas"); label_reporte = QLabel("Reportes:"); label_reporte.setStyleSheet("font-weight: bold; margin-top: 15px;")
        
        self.btn_obtener_metricas = QPushButton("Obtener métricas")
        self.btn_agregar_imagen_reporte = QPushButton("Agregar imagen")
        self.btn_descargar_reporte = QPushButton("Descargar reporte")
        self.btn_finalizar_reporte = QPushButton("Finalizar reporte")
        
        self.btn_corregir_filtrado = QPushButton("Corregir Filtrado")

        estilo_btn_menu = "QPushButton { background-color: transparent; text-align: left; padding: 8px 10px; font-weight: normal; color: #333333; border: 1px solid transparent; outline: none; font-size: 11px;} QPushButton:hover { background-color: #eaf2ff; border-radius: 5px; } QPushButton:disabled { color: #aaaaaa; }"
        lista_botones = [
            self.btn_cargar, self.btn_conteo, self.btn_filtrar, self.btn_ramas, 
            self.btn_obtener_metricas, 
            self.btn_descargar_reporte, self.btn_finalizar_reporte, self.btn_corregir_filtrado
        ]
        for btn in lista_botones: 
            btn.setStyleSheet(estilo_btn_menu)
            
        self.btn_corregir_filtrado.hide()
        self.btn_corregir_filtrado.setStyleSheet(estilo_btn_menu + "QPushButton { color: #0969da; font-weight: bold; }")

        # Panel de validación con checkboxes (definido en ValidacionReporteMixin)
        self._crear_panel_validacion(self.menu_lateral)


        # Conectar el botón de historial
        self.btn_historial.clicked.connect(self.abrir_historial)

        
        self.frame_filtros = QFrame()
        self.frame_filtros.hide()
        layout_filtros = QVBoxLayout(self.frame_filtros)
        layout_filtros.setContentsMargins(0, 10, 0, 10)
        lbl_f_titulo = QLabel("Ajuste de Filtros Globales"); lbl_f_titulo.setStyleSheet("font-weight: bold; color: #003366;")

        layout_filtros.addWidget(lbl_f_titulo)
        lbl_clahe = QLabel("Contraste (CLAHE):")
        self.sld_clahe = QSlider(Qt.Orientation.Horizontal); self.sld_clahe.setRange(0, 10); self.sld_clahe.setValue(2)
        layout_filtros.addWidget(lbl_clahe); layout_filtros.addWidget(self.sld_clahe)
        lbl_gauss = QLabel("Suavizado Gaussiano:")
        self.sld_gauss = QSlider(Qt.Orientation.Horizontal); self.sld_gauss.setRange(1, 15); self.sld_gauss.setSingleStep(2); self.sld_gauss.setValue(5)
        layout_filtros.addWidget(lbl_gauss); layout_filtros.addWidget(self.sld_gauss)
        lbl_otsu = QLabel("Umbral Binarización:")
        self.sld_otsu = QSlider(Qt.Orientation.Horizontal); self.sld_otsu.setRange(-50, 50); self.sld_otsu.setValue(0)
        layout_filtros.addWidget(lbl_otsu); layout_filtros.addWidget(self.sld_otsu)
        lbl_ruido = QLabel("Eliminar Ruido Máx (px):")
        self.sld_ruido = QSlider(Qt.Orientation.Horizontal); self.sld_ruido.setRange(0, 200); self.sld_ruido.setValue(50)
        layout_filtros.addWidget(lbl_ruido); layout_filtros.addWidget(self.sld_ruido)
        estilo_slider = "QSlider::groove:horizontal { border: 1px solid #d0d7de; height: 4px; background: #f6f8fa; margin: 2px 0; border-radius: 2px; } QSlider::handle:horizontal { background: #ffffff; border: 1px solid #3a61a0; width: 12px; height: 12px; margin: -5px 0; border-radius: 6px; } QSlider::handle:horizontal:hover { background: #eaf2ff; }"
        self.sld_clahe.setStyleSheet(estilo_slider); self.sld_gauss.setStyleSheet(estilo_slider); self.sld_otsu.setStyleSheet(estilo_slider); self.sld_ruido.setStyleSheet(estilo_slider)


        btn_f_layout = QHBoxLayout()
        btn_aceptar_filtro = QPushButton("Aceptar")
        btn_aceptar_filtro.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_aceptar_filtro.setStyleSheet("""
            QPushButton { 
                background-color: transparent; 
                border: 1.5px solid #2da44e; 
                color: #2da44e; 
                font-weight: bold; 
                padding: 6px; 
                font-size: 11px; 
                border-radius: 5px; 
            }
            QPushButton:hover { 
                background-color: #2da44e; 
                color: white; 
            }
        """)
        btn_cancelar_filtro = QPushButton("Cancelar")
        btn_cancelar_filtro.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_cancelar_filtro.setStyleSheet("""
            QPushButton { 
                background-color: transparent; 
                border: 1.5px solid #cf222e; 
                color: #cf222e; 
                font-weight: bold; 
                padding: 6px; 
                font-size: 11px; 
                border-radius: 5px; 
            }
            QPushButton:hover { 
                background-color: #cf222e; 
                color: white; 
            }
        """)
        btn_f_layout.addWidget(btn_cancelar_filtro)
        btn_f_layout.addWidget(btn_aceptar_filtro)
        layout_filtros.addLayout(btn_f_layout)
        self.menu_lateral.addWidget(self.frame_filtros)
        
        self.sld_clahe.valueChanged.connect(self.previsualizar_filtrado)
        self.sld_gauss.valueChanged.connect(self.previsualizar_filtrado)
        self.sld_otsu.valueChanged.connect(self.previsualizar_filtrado)
        self.sld_ruido.valueChanged.connect(self.previsualizar_filtrado)
        btn_aceptar_filtro.clicked.connect(self.confirmar_filtrado)
        btn_cancelar_filtro.clicked.connect(self.cancelar_filtrado)


        self.menu_lateral.addStretch()
        self.btn_cerrar_sesion = QPushButton("Cerrar Sesión"); self.btn_cerrar_sesion.setStyleSheet("QPushButton { background-color: transparent; border: 2px solid #cc0000; color: #cc0000; font-weight: bold; border-radius: 8px; padding: 10px; margin-top: 20px; } QPushButton:hover { background-color: #cc0000; color: white; }"); self.menu_lateral.addWidget(self.btn_cerrar_sesion)
        frame_menu = QFrame(); frame_menu.setObjectName("menu_lateral"); frame_menu.setFixedWidth(200); frame_menu.setLayout(self.menu_lateral)
        
        area_imagen = QVBoxLayout()
        controles_superiores = QHBoxLayout()
        controles_superiores.setContentsMargins(15, 5, 15, 5)
        controles_superiores.setSpacing(10)
        self.combo_vista = QComboBox(); self.combo_vista.addItem("Original"); self.combo_vista.setMinimumWidth(140); self.combo_vista.setStyleSheet("""
            QComboBox { background-color: white; border: 1px solid #d0d7de; border-radius: 6px; padding: 4px 10px; font-size: 11px; font-weight: bold; color: #24292f; min-width: 140px; }
            QComboBox:hover { background-color: #f6f8fa; }
            QComboBox::drop-down { 
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 20px; 
                border: none;
            }
            QComboBox::down-arrow { 
                image: url(assets/buttons/abajo.png);
                width: 12px; 
                height: 12px;
            }
            QComboBox QAbstractItemView { background-color: white; border: 1px solid #d0d7de; selection-background-color: #eaf2ff; selection-color: #0969da; outline: none; }
        """); self.combo_vista.setEnabled(False); self.combo_vista.currentTextChanged.connect(self.cambiar_vista_global)
        
        # Botones de navegación global
        self.btn_ant_global = QPushButton()
        self.btn_ant_global.setIcon(QIcon("assets/buttons/izq.png"))
        self.btn_ant_global.setIconSize(QSize(12, 12))
        self.btn_ant_global.setFixedSize(22, 22)
        self.btn_ant_global.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_ant_global.setStyleSheet("""
            QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
            QPushButton:hover { background-color: #eaf2ff; border-radius: 11px; }
            QPushButton:disabled { opacity: 0.3; }
        """)
        self.btn_ant_global.clicked.connect(self.anterior_vista_global)
        self.btn_ant_global.setEnabled(False)
        
        self.btn_sig_global = QPushButton()
        self.btn_sig_global.setIcon(QIcon("assets/buttons/der.png"))
        self.btn_sig_global.setIconSize(QSize(12, 12))
        self.btn_sig_global.setFixedSize(22, 22)
        self.btn_sig_global.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_sig_global.setStyleSheet("""
            QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; }
            QPushButton:hover { background-color: #eaf2ff; border-radius: 11px; }
            QPushButton:disabled { opacity: 0.3; }
        """)
        self.btn_sig_global.clicked.connect(self.siguiente_vista_global)
        self.btn_sig_global.setEnabled(False)
        
        # Botón de Observaciones (sólo para tesista, se muestra si hay comentario)
        self.btn_observaciones = QPushButton("  Observaciones")
        self.btn_observaciones.setIcon(QIcon("assets/buttons/msg.png"))
        self.btn_observaciones.setIconSize(QSize(14, 14))
        self.btn_observaciones.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_observaciones.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: 1.5px solid #0969da;
                color: #0969da;
                font-weight: bold;
                border-radius: 6px;
                padding: 4px 10px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #0969da;
                color: white;
            }
        """)
        self.btn_observaciones.hide()

        controles_superiores.addWidget(self.btn_ant_global)
        controles_superiores.addWidget(self.combo_vista)
        controles_superiores.addWidget(self.btn_sig_global)
        controles_superiores.addWidget(self.btn_observaciones)
        controles_superiores.addStretch()
        
        self.lbl_info_conteo = QLabel("Microglías detectadas: 0"); self.lbl_info_conteo.setStyleSheet("font-size: 11px; font-weight: bold; color: #3a61a0; background-color: white; border: 1px solid #d0d7de; border-radius: 6px; padding: 4px 10px;"); self.lbl_info_conteo.setAlignment(Qt.AlignmentFlag.AlignCenter); controles_superiores.addWidget(self.lbl_info_conteo); controles_superiores.addSpacing(15)
        
        estilo_herramienta = "QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; padding: 2px; } QPushButton:hover { background-color: #eaf2ff; border-radius: 17px; } QPushButton:checked { background-color: #cce5ff; border: 1px solid #007bff; border-radius: 17px; } QPushButton:disabled { opacity: 0.5; }"
        
        self.btn_herramienta_caja = SafeToolTipButton()
        self.btn_herramienta_caja.setFixedSize(35, 35)
        self.btn_herramienta_caja.setIcon(QIcon("assets/buttons/seleccionar.png"))
        self.btn_herramienta_caja.setIconSize(QSize(20, 20))
        self.btn_herramienta_caja.setCustomToolTip("Crear seleccion")
        self.btn_herramienta_caja.setStyleSheet(estilo_herramienta)
        self.btn_herramienta_caja.setCheckable(True)
        self.btn_herramienta_caja.setEnabled(False)
        self.btn_herramienta_caja.hide() 
        
        self.btn_herramienta_eliminar = SafeToolTipButton()
        self.btn_herramienta_eliminar.setFixedSize(35, 35)
        self.btn_herramienta_eliminar.setIcon(QIcon("assets/buttons/borrar.png"))
        self.btn_herramienta_eliminar.setIconSize(QSize(20, 20))
        self.btn_herramienta_eliminar.setCustomToolTip("Eliminar seleccion")
        self.btn_herramienta_eliminar.setStyleSheet(estilo_herramienta)
        self.btn_herramienta_eliminar.setCheckable(True)
        self.btn_herramienta_eliminar.setEnabled(False)
        self.btn_herramienta_eliminar.hide()
        
        self.btn_herramienta_caja.clicked.connect(self.toggle_herramienta_caja)
        self.btn_herramienta_eliminar.clicked.connect(self.toggle_herramienta_eliminar)

        controles_superiores.addWidget(self.btn_herramienta_caja); controles_superiores.addWidget(self.btn_herramienta_eliminar); controles_superiores.addSpacing(15)
        
        lbl_minus = QLabel("-"); lbl_minus.setStyleSheet("font-size: 24px; font-weight: bold; color: #555;"); lbl_plus = QLabel("+"); lbl_plus.setStyleSheet("font-size: 20px; font-weight: bold; color: #555;")
        self.sld_nivel_zoom = QSlider(Qt.Orientation.Horizontal); self.sld_nivel_zoom.setRange(50, 400); self.sld_nivel_zoom.setValue(100); self.sld_nivel_zoom.setFixedWidth(120); self.sld_nivel_zoom.setEnabled(False)
        self.sld_nivel_zoom.setStyleSheet("""
            QSlider::groove:horizontal { border: 1px solid #d0d7de; height: 4px; background: #f6f8fa; margin: 2px 0; border-radius: 2px; } 
            QSlider::handle:horizontal { background: #ffffff; border: 1px solid #3a61a0; width: 10px; height: 10px; margin: -4px 0; border-radius: 5px; }
            QSlider::handle:horizontal:hover { background: #eaf2ff; }
        """)

        
        self.btn_zoom_reset = QPushButton("↺")
        self.btn_zoom_reset.setFixedSize(35, 35)
        self.btn_zoom_reset.setToolTip("Restablecer Zoom")
        self.btn_zoom_reset.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_zoom_reset.setStyleSheet("""
            QPushButton { background-color: transparent; border: 1px solid transparent; outline: none; font-size: 18px; font-weight: bold; color: #3a61a0; padding: 0; text-align: center; }
            QPushButton:hover { background-color: #eaf2ff; border-radius: 17px; }
            QPushButton:disabled { color: #aaaaaa; }
        """)
        self.btn_zoom_reset.setEnabled(False)
        
        self.btn_bloquear_zoom = SafeToolTipButton()
        self.btn_bloquear_zoom.setFixedSize(35, 35)
        self.btn_bloquear_zoom.setIcon(QIcon("assets/buttons/desbloqueado.png"))
        self.btn_bloquear_zoom.setIconSize(QSize(20, 20))
        self.btn_bloquear_zoom.setCustomToolTip("Bloquear zoom")
        self.btn_bloquear_zoom.setStyleSheet(estilo_herramienta)
        self.btn_bloquear_zoom.setCheckable(True)
        self.btn_bloquear_zoom.setEnabled(False)
        self.btn_bloquear_zoom.toggled.connect(self.toggle_bloqueo_zoom)
        
        controles_superiores.addWidget(lbl_minus); controles_superiores.addWidget(self.sld_nivel_zoom); controles_superiores.addWidget(lbl_plus); controles_superiores.addSpacing(5); controles_superiores.addWidget(self.btn_zoom_reset); controles_superiores.addWidget(self.btn_bloquear_zoom)
        
        self.stacked_visor = QStackedWidget()
        self.visor_imagen = InteractiveImageViewer(); self.visor_imagen.setText("Sube una imagen .tiff para empezar el análisis..."); self.visor_imagen.setStyleSheet("border: 2px dashed #aaa; background-color: #f0f0f0; font-size: 18px; color: #666;")
        self.visor_imagen.conteo_actualizado.connect(self.conteo_modificado_auto_save); self.visor_imagen.nueva_caja_dibujada.connect(self.agregar_microglia_manual); self.visor_imagen.nivel_zoom_cambiado.connect(self.sld_nivel_zoom.setValue)
        self.sld_nivel_zoom.valueChanged.connect(self.visor_imagen.set_zoom); self.btn_zoom_reset.clicked.connect(self.reset_zoom)
        
        # Panel para tabla de métricas (tipo Excel con Dashboard de KPIs)
        self.panel_metricas_tabla = QFrame()
        self.panel_metricas_tabla.setStyleSheet("background-color: white; border: 1px solid #d0d7de; border-radius: 6px;")
        layout_panel_met = QVBoxLayout(self.panel_metricas_tabla)
        layout_panel_met.setContentsMargins(20, 20, 20, 20)
        layout_panel_met.setSpacing(15)
        
        # Cabecera
        self.lbl_titulo_tabla = QLabel("<b>Métricas Morfológicas de Microglías (Reporte de Sesión)</b>")
        self.lbl_titulo_tabla.setStyleSheet("font-size: 15px; color: #3a61a0; border: none; font-weight: bold;")
        self.lbl_titulo_tabla.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout_panel_met.addWidget(self.lbl_titulo_tabla)
        
        self.lbl_subtitulo_imagen = QLabel()
        self.lbl_subtitulo_imagen.setStyleSheet("font-size: 12px; color: #57606a; border: none; font-weight: normal;")
        self.lbl_subtitulo_imagen.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout_panel_met.addWidget(self.lbl_subtitulo_imagen)
        
        # Contenedor Visual de Procedimiento con Miniaturas
        self.widget_proceso = QWidget()
        layout_proceso = QHBoxLayout(self.widget_proceso)
        layout_proceso.setContentsMargins(0, 5, 0, 5)
        layout_proceso.setSpacing(15)
        layout_proceso.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Estilo de etiqueta de miniatura
        thumb_style = "border: 1px solid #d0d7de; border-radius: 4px; background-color: #f6f8fa; color: #8c95a0; font-size: 11px;"
        
        # Bloque 1: Original
        w_p1 = QWidget()
        lay_p1 = QVBoxLayout(w_p1)
        lay_p1.setContentsMargins(0, 0, 0, 0)
        lay_p1.setSpacing(4)
        lay_p1.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_txt_p1 = QLabel("1. Original / Conteo")
        lbl_txt_p1.setStyleSheet("font-size: 11px; color: #57606a; font-weight: bold; border: none;")
        lbl_txt_p1.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_thumb_original = QLabel()
        self.lbl_thumb_original.setFixedSize(220, 140)
        self.lbl_thumb_original.setStyleSheet(thumb_style)
        self.lbl_thumb_original.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay_p1.addWidget(lbl_txt_p1)
        lay_p1.addWidget(self.lbl_thumb_original)
        
        # Flecha 1
        lbl_flecha1 = QLabel("➔")
        lbl_flecha1.setStyleSheet("font-size: 20px; color: #0969da; border: none; font-weight: bold;")
        lbl_flecha1.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Bloque 2: Filtrado
        w_p2 = QWidget()
        lay_p2 = QVBoxLayout(w_p2)
        lay_p2.setContentsMargins(0, 0, 0, 0)
        lay_p2.setSpacing(4)
        lay_p2.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_txt_p2 = QLabel("2. Filtrado")
        lbl_txt_p2.setStyleSheet("font-size: 11px; color: #57606a; font-weight: bold; border: none;")
        lbl_txt_p2.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_thumb_filtrada = QLabel()
        self.lbl_thumb_filtrada.setFixedSize(220, 140)
        self.lbl_thumb_filtrada.setStyleSheet(thumb_style)
        self.lbl_thumb_filtrada.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay_p2.addWidget(lbl_txt_p2)
        lay_p2.addWidget(self.lbl_thumb_filtrada)
        
        # Flecha 2
        lbl_flecha2 = QLabel("➔")
        lbl_flecha2.setStyleSheet("font-size: 20px; color: #0969da; border: none; font-weight: bold;")
        lbl_flecha2.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Bloque 3: Esqueleto
        w_p3 = QWidget()
        lay_p3 = QVBoxLayout(w_p3)
        lay_p3.setContentsMargins(0, 0, 0, 0)
        lay_p3.setSpacing(4)
        lay_p3.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_txt_p3 = QLabel("3. Esqueleto / Ramas")
        lbl_txt_p3.setStyleSheet("font-size: 11px; color: #57606a; font-weight: bold; border: none;")
        lbl_txt_p3.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_thumb_esqueleto = QLabel()
        self.lbl_thumb_esqueleto.setFixedSize(220, 140)
        self.lbl_thumb_esqueleto.setStyleSheet(thumb_style)
        self.lbl_thumb_esqueleto.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay_p3.addWidget(lbl_txt_p3)
        lay_p3.addWidget(self.lbl_thumb_esqueleto)
        
        layout_proceso.addWidget(w_p1)
        layout_proceso.addWidget(lbl_flecha1)
        layout_proceso.addWidget(w_p2)
        layout_proceso.addWidget(lbl_flecha2)
        layout_proceso.addWidget(w_p3)
        layout_panel_met.addWidget(self.widget_proceso)
        
        # Tarjetas de resumen (KPI Cards)
        self.kpi_layout = QHBoxLayout()
        self.kpi_layout.setSpacing(15)
        
        # Tarjeta 1: Total Microglías
        self.card_total = QFrame()
        self.card_total.setStyleSheet("background-color: #f6f8fa; border: 1px solid #d0d7de; border-radius: 6px;")
        self.card_total.setFixedHeight(75)
        layout_total = QVBoxLayout(self.card_total)
        layout_total.setSpacing(4)
        layout_total.setContentsMargins(12, 10, 12, 10)
        lbl_title_total = QLabel("Total Microglías")
        lbl_title_total.setStyleSheet("font-size: 11px; color: #57606a; font-weight: bold; border: none;")
        self.val_total = QLabel("0")
        self.val_total.setStyleSheet("font-size: 20px; color: #3a61a0; font-weight: bold; border: none;")
        layout_total.addWidget(lbl_title_total)
        layout_total.addWidget(self.val_total)
        
        # Tarjeta 2: Promedio Puntos de Bifurcación
        self.card_junc = QFrame()
        self.card_junc.setStyleSheet("background-color: #f6f8fa; border: 1px solid #d0d7de; border-radius: 6px;")
        self.card_junc.setFixedHeight(75)
        layout_junc = QVBoxLayout(self.card_junc)
        layout_junc.setSpacing(4)
        layout_junc.setContentsMargins(12, 10, 12, 10)
        lbl_title_junc = QLabel("Prom. Bifurcaciones")
        lbl_title_junc.setStyleSheet("font-size: 11px; color: #57606a; font-weight: bold; border: none;")
        self.val_junc = QLabel("0.0")
        self.val_junc.setStyleSheet("font-size: 20px; color: #2da44e; font-weight: bold; border: none;")
        layout_junc.addWidget(lbl_title_junc)
        layout_junc.addWidget(self.val_junc)

        # Tarjeta 3: Promedio Puntos Terminales
        self.card_ends = QFrame()
        self.card_ends.setStyleSheet("background-color: #f6f8fa; border: 1px solid #d0d7de; border-radius: 6px;")
        self.card_ends.setFixedHeight(75)
        layout_ends = QVBoxLayout(self.card_ends)
        layout_ends.setSpacing(4)
        layout_ends.setContentsMargins(12, 10, 12, 10)
        lbl_title_ends = QLabel("Prom. Ptos Terminales")
        lbl_title_ends.setStyleSheet("font-size: 11px; color: #57606a; font-weight: bold; border: none;")
        self.val_ends = QLabel("0.0")
        self.val_ends.setStyleSheet("font-size: 20px; color: #cf222e; font-weight: bold; border: none;")
        layout_ends.addWidget(lbl_title_ends)
        layout_ends.addWidget(self.val_ends)

        # Tarjeta 4: Promedio Longitud de Ramas
        self.card_length = QFrame()
        self.card_length.setStyleSheet("background-color: #f6f8fa; border: 1px solid #d0d7de; border-radius: 6px;")
        self.card_length.setFixedHeight(75)
        layout_length = QVBoxLayout(self.card_length)
        layout_length.setSpacing(4)
        layout_length.setContentsMargins(12, 10, 12, 10)
        lbl_title_len = QLabel("Prom. Longitud Rama")
        lbl_title_len.setStyleSheet("font-size: 11px; color: #57606a; font-weight: bold; border: none;")
        self.val_len = QLabel("0.0 px")
        self.val_len.setStyleSheet("font-size: 20px; color: #0969da; font-weight: bold; border: none;")
        layout_length.addWidget(lbl_title_len)
        layout_length.addWidget(self.val_len)
        
        self.kpi_layout.addWidget(self.card_total)
        self.kpi_layout.addWidget(self.card_junc)
        self.kpi_layout.addWidget(self.card_ends)
        self.kpi_layout.addWidget(self.card_length)
        # Título de la Sección de Métricas (debajo de las miniaturas)
        lbl_titulo_seccion = QLabel("<b>Métricas Morfológicas de la Sesión</b>")
        lbl_titulo_seccion.setStyleSheet("font-size: 13px; color: #24292f; border: none; font-weight: bold; margin-top: 10px; margin-bottom: 5px;")
        lbl_titulo_seccion.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout_panel_met.addWidget(lbl_titulo_seccion)
        
        layout_panel_met.addLayout(self.kpi_layout)
        
        # Tabla estilo Excel
        self.tabla_metricas = QTableWidget()
        self.tabla_metricas.setStyleSheet("""
            QTableWidget { border: 1px solid #e1e4e8; background-color: #ffffff; gridline-color: #e1e4e8; font-size: 11px; border-radius: 4px; }
            QTableWidget::item { padding: 8px; color: #24292f; border-bottom: 1px solid #f0f0f0; }
            QTableWidget::item:selected { background-color: #eaf2ff; color: #0969da; }
            QHeaderView::section { background-color: #f6f8fa; padding: 8px; font-weight: bold; border: 1px solid #e1e4e8; color: #57606a; font-size: 11px; }
            
            QScrollBar:vertical {
                border: none;
                background: #f6f8fa;
                width: 10px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #b6d4fe;
                min-height: 20px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical:hover {
                background: #0969da;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                border: none;
                background: none;
                height: 0px;
            }
            
            QScrollBar:horizontal {
                border: none;
                background: #f6f8fa;
                height: 10px;
                margin: 0px;
            }
            QScrollBar::handle:horizontal {
                background: #b6d4fe;
                min-width: 20px;
                border-radius: 5px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #0969da;
            }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                border: none;
                background: none;
                width: 0px;
            }
        """)
        self.tabla_metricas.setAlternatingRowColors(True)
        self.tabla_metricas.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout_panel_met.addWidget(self.tabla_metricas)
        
        self.stacked_visor.addWidget(self.visor_imagen)
        self.stacked_visor.addWidget(self.panel_metricas_tabla)
        
        area_imagen.addLayout(controles_superiores); area_imagen.addWidget(self.stacked_visor, stretch=1)
        layout_principal.addWidget(frame_menu); layout_principal.addLayout(area_imagen, stretch=1); widget_central.setLayout(layout_principal); self.setCentralWidget(widget_central)
        
        self.btn_cargar.clicked.connect(self.cargar_imagen); self.btn_cerrar_sesion.clicked.connect(self.cerrar_sesion); self.btn_conteo.clicked.connect(self.execute_microglia_counting); self.btn_filtrar.clicked.connect(self.ejecutar_filtrado); self.btn_ramas.clicked.connect(self.mostrar_ramas_morfologia); self.btn_corregir_filtrado.clicked.connect(self.corregir_filtrado)
        self.btn_observaciones.clicked.connect(self.mostrar_observaciones_popup)
        
        self.btn_obtener_metricas.clicked.connect(self.obtener_metricas)
        self.btn_agregar_imagen_reporte.clicked.connect(self.agregar_imagen_reporte)
        self.btn_descargar_reporte.clicked.connect(self.descargar_reporte)
        self.btn_finalizar_reporte.clicked.connect(self.finalizar_reporte)
        # confirmar_validacion y abrir_historial se conectan desde ValidacionReporteMixin

        self.actualizar_estado_flujo(0)
        
        if self.rol in ["Invitado", "Guest"]: self.btn_historial.hide()
        
        # Verificar al inicio si es Tesista y tiene reportes devueltos con comentarios
        if self.rol == "Tesista":
            QTimer.singleShot(1000, self.verificar_reportes_devueltos_al_inicio)

    def toggle_herramienta_caja(self, checked):
        if checked: self.btn_herramienta_eliminar.setChecked(False); self.visor_imagen.current_tool = "draw"; self.visor_imagen.hovered_index = -1; self.visor_imagen.draw_current_state()
        else: self.visor_imagen.current_tool = "pointer"

    def toggle_herramienta_eliminar(self, checked):
        if checked: self.btn_herramienta_caja.setChecked(False); self.visor_imagen.current_tool = "delete"
        else: self.visor_imagen.current_tool = "pointer"

    def reset_zoom(self): self.sld_nivel_zoom.setValue(100); self.visor_imagen.pan_x = 0; self.visor_imagen.pan_y = 0; self.visor_imagen.update()

    def toggle_bloqueo_zoom(self, checked):
        if checked:
            self.btn_bloquear_zoom.setIcon(QIcon("assets/buttons/bloqueado.png"))
            self.btn_bloquear_zoom.setCustomToolTip("Desbloquear zoom")
        else:
            self.btn_bloquear_zoom.setIcon(QIcon("assets/buttons/desbloqueado.png"))
            self.btn_bloquear_zoom.setCustomToolTip("Bloquear zoom")
            
        self.sld_nivel_zoom.setEnabled(not checked); self.btn_zoom_reset.setEnabled(not checked); self.visor_imagen.lock_zoom(checked)

    def conteo_modificado_auto_save(self, conteo):
        self.actualizar_etiqueta_conteo(conteo)
        self.save_current_progress(mostrar_notif=False)

    def actualizar_estado_flujo(self, paso):
        self.paso_actual = paso
        # Reset report buttons by default
        self.btn_obtener_metricas.setEnabled(False)
        self.btn_agregar_imagen_reporte.setEnabled(False)
        self.btn_descargar_reporte.setEnabled(False)
        self.btn_finalizar_reporte.setEnabled(False)

        # Activar/Desactivar controles de zoom a partir del conteo (paso 1 o mayor)
        zoom_activado = (paso >= 1)
        self.sld_nivel_zoom.setEnabled(zoom_activado)
        self.btn_zoom_reset.setEnabled(zoom_activado)
        self.btn_bloquear_zoom.setEnabled(zoom_activado)

        # Habilitación estricta de botones según la etapa actual de análisis (se desactiva la etapa anterior)
        es_validacion = getattr(self, 'reporte_validado_cargado', False)
        if es_validacion:
            # Modo revisión: todos los botones de acción deshabilitados
            self.btn_cargar.setEnabled(False)
            self.btn_conteo.setEnabled(False)
            self.btn_filtrar.setEnabled(False)
            self.btn_ramas.setEnabled(False)
            self.btn_obtener_metricas.setEnabled(False)
            self.btn_agregar_imagen_reporte.setEnabled(False)
            self.btn_descargar_reporte.setEnabled(False)
            self.btn_finalizar_reporte.setEnabled(False)
            # Habilitar la navegación de vistas para revisar libremente
            self.combo_vista.setEnabled(True)
            self.btn_herramienta_caja.hide()
            self.btn_herramienta_eliminar.hide()
            self.btn_corregir_filtrado.hide()
        else:
            if self.rol == "Tesista":
                if self.id_reporte_actual is None:
                    self.btn_cargar.setEnabled(False)
                else:
                    # Si el rol es Tesista y ya hay imágenes, sólo permitir si está validado por el investigador
                    from bd.database import conectar
                    conn = conectar(); cur = conn.cursor()
                    try:
                        cur.execute("SELECT COUNT(*) FROM Analisis WHERE id_reporte = ?", (self.id_reporte_actual,))
                        total_imagenes = cur.fetchone()[0]
                        cur.execute("SELECT estado FROM ReporteCompartido WHERE id_reporte = ?", (self.id_reporte_actual,))
                        comp_res = cur.fetchone()
                        reporte_validado = (comp_res is not None and comp_res[0] == 'Validado')
                    except Exception as e:
                        logging.error(f"Error checking validation status for cargar button: {e}")
                        total_imagenes = 0
                        reporte_validado = False
                    finally:
                        conn.close()
                    
                    if total_imagenes == 0:
                        self.btn_cargar.setEnabled(paso == 0 or paso == 5)
                    else:
                        self.btn_cargar.setEnabled(reporte_validado and (paso == 0 or paso == 5))
            else:
                self.btn_cargar.setEnabled(paso == 0 or paso == 5)
            self.btn_conteo.setEnabled(paso == 1)
            self.btn_filtrar.setEnabled(paso == 2)
            self.btn_ramas.setEnabled(paso == 3)

        if paso == 0:
            if not es_validacion: self.combo_vista.setEnabled(False)
            self.btn_herramienta_caja.hide(); self.btn_herramienta_eliminar.hide()
        elif paso == 1:
            if not es_validacion: self.combo_vista.setEnabled(False)
            self.btn_herramienta_caja.hide(); self.btn_herramienta_eliminar.hide()
        elif paso == 2:
            self.combo_vista.setEnabled(True)
            self.btn_herramienta_caja.show(); self.btn_herramienta_eliminar.show()
            self.btn_herramienta_caja.setEnabled(True); self.btn_herramienta_eliminar.setEnabled(True)
        elif paso == 3:
            self.combo_vista.setEnabled(True)
            self.btn_herramienta_caja.hide(); self.btn_herramienta_caja.setChecked(False)
            self.btn_herramienta_eliminar.hide(); self.btn_herramienta_eliminar.setChecked(False)
            self.visor_imagen.current_tool = "pointer"
        elif paso == 4:
            self.btn_corregir_filtrado.show()
            self.btn_herramienta_caja.hide(); self.btn_herramienta_eliminar.hide()
            self.btn_obtener_metricas.setEnabled(True)
        elif paso == 5:
            self.btn_corregir_filtrado.hide()
            self.btn_herramienta_caja.hide(); self.btn_herramienta_eliminar.hide()
            self.btn_obtener_metricas.setEnabled(False)
            self.aplicar_ui_paso5_validacion()  # Delegado a ValidacionReporteMixin
            if not es_validacion: self.combo_vista.setEnabled(False)
            
        if paso not in [4]: self.btn_corregir_filtrado.hide()
        
        self.actualizar_elementos_combo_vista()
        self.actualizar_botones_navegacion()
        
        # Guardado automático de progreso al cambiar de fase
        if paso > 0:
            self.save_current_progress(mostrar_notif=False)

    def verificar_reportes_devueltos_al_inicio(self):
        """Verifica al inicio si el tesista tiene reportes devueltos (Pendientes) con comentarios."""
        from bd.database import conectar
        conn = conectar(); cur = conn.cursor()
        try:
            cur.execute("""
                SELECT R.nombre_reporte, RC.comentarios 
                FROM ReporteCompartido RC
                JOIN Reporte R ON RC.id_reporte = R.id_reporte
                WHERE RC.id_propietario = ? AND RC.estado = 'Pendiente' AND RC.comentarios IS NOT NULL AND RC.comentarios != ''
            """, (self.id_usuario,))
            rows = cur.fetchall()
            if rows:
                from vistas.utilidades import DialogoNotificacion
                msg_final = "El investigador ha devuelto los siguientes reportes para su corrección:\n\n"
                for nombre_rep, coment in rows:
                    msg_final += f"• Reporte: {nombre_rep}\n  Observaciones: {coment}\n\n"
                msg_final += "Por favor, abre tu Historial para cargarlos y realizar las modificaciones correspondientes."
                DialogoNotificacion(
                    "Correcciones Requeridas",
                    msg_final,
                    "warning",
                    self
                ).exec()
        except Exception as e:
            logging.error(f"Error al verificar reportes devueltos al inicio: {e}")
        finally:
            conn.close()

    def mostrar_observaciones_popup(self):
        """Muestra las observaciones del investigador en un diálogo emergente."""
        comentarios = getattr(self, "comentarios_correccion", "")
        if comentarios:
            from vistas.utilidades import DialogoNotificacion
            DialogoNotificacion(
                "Observaciones del Investigador",
                comentarios,
                "warning",
                self
            ).exec()

    def save_current_progress(self, mostrar_notif=True):
        if getattr(self, "cargando_reporte", False):
            return
        if not self.ruta_imagen_actual: return
        
        from bd.database import conectar
        import json
        conn = conectar(); cur = conn.cursor()
        
        try:
            # 1. Asegurar Reporte (Grupo)
            if not self.id_reporte_actual:
                cur.execute("INSERT INTO Reporte (id_usuario, nombre_reporte) VALUES (?, ?)", 
                           (self.id_usuario, f"Reporte {datetime.now().strftime('%Y-%m-%d %H:%M')}"))
                self.id_reporte_actual = cur.lastrowid

            # 2. Asegurar Imagen
            cur.execute("SELECT id_imagen FROM Imagen WHERE ruta_archivo = ?", (self.ruta_imagen_actual,))
            res = cur.fetchone()
            if not res:
                cur.execute("INSERT INTO Imagen (id_usuario, ruta_archivo, formato, campo, tiempo_muestra) VALUES (?,?,?,?,?)",
                           (self.id_usuario, self.ruta_imagen_actual, 'TIFF', self.metadatos_imagen.get("campo"), self.metadatos_imagen.get("tiempo")))
                id_img = cur.lastrowid
            else: id_img = res[0]

            # 3. Guardar Análisis con Estado Persistente
            if not self.id_analisis_actual:
                cur.execute("INSERT INTO Analisis (id_reporte, id_imagen, cantidad_microglias, paso_actual) VALUES (?,?,?,?)",
                           (self.id_reporte_actual, id_img, len(self.visor_imagen.boxes), self.paso_actual))
                self.id_analisis_actual = cur.lastrowid
            else:
                # Solo actualizar cantidad_microglias si hay boxes reales, para no sobreescribir un conteo previo con 0
                if self.visor_imagen.boxes:
                    cur.execute("UPDATE Analisis SET cantidad_microglias = ?, paso_actual = ? WHERE id_analisis = ?",
                               (len(self.visor_imagen.boxes), self.paso_actual, self.id_analisis_actual))
                else:
                    # Sin boxes: preservar cantidad_microglias existente, solo actualizar paso
                    cur.execute("UPDATE Analisis SET paso_actual = ? WHERE id_analisis = ?",
                               (self.paso_actual, self.id_analisis_actual))

            # 4. Sincronizar Microglias (Detecciones individuales) — solo si hay boxes activos
            if self.visor_imagen.boxes:
                # Recuperar métricas si existían para no borrarlas al guardar
                cur.execute("SELECT crop_path, puntos_finales, uniones_triples, uniones_cuadruples, longitud_promedio_ramas, longitud_maxima_rama, ruta_mas_larga, lineas, puntos_union, voxeles_union, voxeles_losa FROM Microglia WHERE id_analisis = ?", (self.id_analisis_actual,))
                metricas_existentes = {row[0]: row[1:] for row in cur.fetchall()}
                
                cur.execute("DELETE FROM Microglia WHERE id_analisis = ?", (self.id_analisis_actual,))
                for box in self.visor_imagen.boxes:
                    offsets = box.get('offsets', {})
                    f_clahe = offsets.get('clahe', 0)
                    f_gauss = offsets.get('gauss', 0)
                    f_otsu = offsets.get('otsu', 0)
                    f_ruido = offsets.get('ruido', 0)
                    areas_elim = json.dumps(box.get('removal_areas', []))
                    c_path = box.get('crop_path', '')
                    
                    if c_path in metricas_existentes:
                        m = metricas_existentes[c_path]
                        cur.execute("""
                            INSERT INTO Microglia (id_analisis, bbox_x, bbox_y, bbox_w, bbox_h, crop_path, 
                                        filtro_clahe, filtro_gauss, filtro_otsu, filtro_ruido, filtro_cierre, areas_eliminadas,
                                        puntos_finales, uniones_triples, uniones_cuadruples, longitud_promedio_ramas, longitud_maxima_rama, ruta_mas_larga, lineas, puntos_union, voxeles_union, voxeles_losa) 
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """, (self.id_analisis_actual, box['x'], box['y'], box['w'], box['h'], c_path, 
                               f_clahe, f_gauss, f_otsu, f_ruido, 0, areas_elim,
                              m[0], m[1], m[2], m[3], m[4], m[5], m[6], m[7], m[8], m[9]))
                    else:
                        cur.execute("INSERT INTO Microglia (id_analisis, bbox_x, bbox_y, bbox_w, bbox_h, crop_path, filtro_clahe, filtro_gauss, filtro_otsu, filtro_ruido, filtro_cierre, areas_eliminadas) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                                   (self.id_analisis_actual, box['x'], box['y'], box['w'], box['h'], c_path, f_clahe, f_gauss, f_otsu, f_ruido, 0, areas_elim))
            
            # 5. Si es un reporte compartido y el usuario actual es el destinatario, marcarlo como Modificado y limpiar comentarios
            cur.execute("SELECT id_usuario FROM Reporte WHERE id_reporte = ?", (self.id_reporte_actual,))
            res_owner = cur.fetchone()
            if res_owner and res_owner[0] != self.id_usuario:
                cur.execute("UPDATE ReporteCompartido SET estado = 'Modificado', comentarios = NULL WHERE id_reporte = ? AND id_destinatario = ?",
                             (self.id_reporte_actual, self.id_usuario))
            
            conn.commit()
            if mostrar_notif:
                self.mostrar_notificacion("Éxito", "Progreso guardado correctamente.", "info")
        except Exception as e:
            self.mostrar_notificacion("Error", f"No se pudo guardar: {e}", "error")
        finally: conn.close()

        # abrir_historial y confirmar_validacion viven en ValidacionReporteMixin
        # (procesamiento/validacion_reporte.py)


    def cargar_reporte_especifico(self, seleccion):
        self.cargando_reporte = True
        id_reporte = seleccion["id_reporte"]
        from bd.database import conectar
        import json
        conn = conectar(); cur = conn.cursor()
        try:
            # 1. Obtener el último análisis de este reporte
            cur.execute("""
                SELECT A.id_analisis, I.ruta_archivo, I.campo, I.tiempo_muestra, A.paso_actual
                FROM Analisis A JOIN Imagen I ON A.id_imagen = I.id_imagen
                WHERE A.id_reporte = ? ORDER BY A.id_analisis DESC LIMIT 1
            """, (id_reporte,))
            res = cur.fetchone()
            
            self.id_reporte_actual = id_reporte
            
            # 1.5. Obtener comentarios del reporte compartido si existen y el usuario es el propietario (estudiante/tesista)
            tiene_correcciones = False
            self.comentarios_correccion = ""
            if hasattr(self, "btn_observaciones"):
                self.btn_observaciones.hide()

            cur.execute("SELECT comentarios, id_propietario FROM ReporteCompartido WHERE id_reporte = ?", (id_reporte,))
            comp_res = cur.fetchone()
            if comp_res:
                comentarios_investigador, prop_id = comp_res
                if self.rol == "Tesista" and comentarios_investigador and comentarios_investigador.strip():
                    tiene_correcciones = True
                    self.comentarios_correccion = comentarios_investigador
                    if hasattr(self, "btn_observaciones"):
                        self.btn_observaciones.show()
                    from vistas.utilidades import DialogoNotificacion
                    # Mostrar comentarios al tesista con un ligero retardo
                    QTimer.singleShot(600, lambda c=comentarios_investigador: DialogoNotificacion(
                        "Observaciones del Investigador",
                        f"El investigador ha rechazado el reporte con las siguientes observaciones:\n\n{c}",
                        "warning", self
                    ).exec())
            
            if not res:
                # Reporte vacío? (No debería pasar)
                self.id_analisis_actual = None; self.metricas_reporte = []
                self.actualizar_estado_flujo(0)
                return

            id_an, ruta, campo, tiempo, paso = res
            
            # 2. Decidir si retomar o empezar nueva imagen
            if paso < 5:
                # RETOMAR ANÁLISIS INCOMPLETO
                from red.config import es_cliente
                if es_cliente():
                    from red.cliente import asegurar_archivo_local
                    ruta = asegurar_archivo_local(ruta)
                self.ruta_imagen_actual = ruta
                self.metadatos_imagen = {"campo": campo, "tiempo": tiempo}
                self.id_analisis_actual = id_an
                self.paso_actual = paso

                cur.execute("""
                    SELECT bbox_x, bbox_y, bbox_w, bbox_h, crop_path, filtro_clahe, filtro_gauss, filtro_otsu, filtro_ruido, areas_eliminadas
                    FROM Microglia WHERE id_analisis = ?
                """, (id_an,))
                boxes_db = cur.fetchall()
                boxes = []
                for (bx, by, bw, bh, cp, fc, fg, fo, fr, ae) in boxes_db:
                    b = {
                        "x": bx, "y": by, "w": bw, "h": bh,
                        "crop_path": cp.replace("\\", "/") if cp else "",
                        "offsets": {"clahe": fc or 0, "gauss": fg or 0, "otsu": fo or 0, "ruido": fr or 0},
                        "removal_areas": json.loads(ae) if ae else []
                    }
                    if es_cliente() and b["crop_path"]:
                        from red.cliente import asegurar_archivo_local
                        asegurar_archivo_local(b["crop_path"])
                    boxes.append(b)
                self.metricas_reporte = []

                # Cargar imagen
                from PyQt6.QtGui import QImage, QPixmap
                import cv2; import numpy as np
                try:
                    with open(ruta, "rb") as f:
                        file_bytes = np.frombuffer(f.read(), dtype=np.uint8)
                    cv_img = cv2.imdecode(file_bytes, cv2.IMREAD_UNCHANGED)
                except Exception:
                    cv_img = None
                if cv_img is not None:
                    if cv_img.dtype == np.uint16: cv_img = ((cv_img-cv_img.min())/(cv_img.max()-cv_img.min())*255).astype(np.uint8)
                    if len(cv_img.shape) == 2: h, w = cv_img.shape; qimg = QImage(cv_img.data, w, h, w, QImage.Format.Format_Grayscale8)
                    else: cv_img = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB); h, w, ch = cv_img.shape; qimg = QImage(cv_img.data, w, h, ch * w, QImage.Format.Format_RGB888)
                    pixmap = QPixmap.fromImage(qimg)
                else: pixmap = QPixmap(ruta)
                
                self.pixmaps_globales = {"Original": pixmap, "Filtrada": None, "Esqueleto": None}
                self.visor_imagen.set_image_and_boxes(pixmap, boxes)
                
                # Reconstruir la lista del combo_vista y las imágenes globales según el paso alcanzado
                self.combo_vista.blockSignals(True)
                # Limpiar opciones extra del combo (mantener solo Original)
                while self.combo_vista.count() > 1:
                    self.combo_vista.removeItem(1)
                
                if paso >= 3:
                    pix_f = self.construir_imagen_global("filtradas")
                    if pix_f:
                        self.pixmaps_globales["Filtrada"] = pix_f
                        self.combo_vista.addItem("Filtrada")
                
                if paso >= 4:
                    pix_e = self.construir_imagen_global("esqueletos")
                    if pix_e:
                        self.pixmaps_globales["Esqueleto"] = pix_e
                        self.combo_vista.addItem("Esqueleto")
                
                # Seleccionar la vista correspondiente al paso donde se quedó el análisis
                if paso >= 4:
                    modo_default = "Esqueleto"
                elif paso == 3:
                    modo_default = "Filtrada"
                else:
                    modo_default = "Original"

                self.combo_vista.setCurrentText(modo_default)
                self.combo_vista.blockSignals(False)
                if not getattr(self, "reporte_validado_cargado", False):
                    self.combo_vista.setEnabled(paso >= 2)
                else:
                    self.combo_vista.setEnabled(True)
                
                # Sincronizar el visor con el modo cargado
                self.visor_imagen.set_view_mode(modo_default, self.pixmaps_globales.get(modo_default))
                self.actualizar_estado_flujo(paso)
                self.actualizar_botones_navegacion()
                if not tiene_correcciones:
                    self.mostrar_notificacion("Éxito", f"Continuando análisis: {os.path.basename(ruta)}", "info")
            else:
                # ÚLTIMA IMAGEN COMPLETADA -> CARGAR MÉTRICAS Y PEDIR NUEVA IMAGEN
                self.ruta_imagen_actual = ruta
                self.metadatos_imagen = {"campo": campo, "tiempo": tiempo}
                self.id_analisis_actual = id_an
                self.paso_actual = 5
                
                # Cargar e inicializar la imagen original
                from PyQt6.QtGui import QImage, QPixmap
                import cv2; import numpy as np
                try:
                    with open(ruta, "rb") as f:
                        file_bytes = np.frombuffer(f.read(), dtype=np.uint8)
                    cv_img = cv2.imdecode(file_bytes, cv2.IMREAD_UNCHANGED)
                except Exception:
                    cv_img = None
                if cv_img is not None:
                    if cv_img.dtype == np.uint16: cv_img = ((cv_img-cv_img.min())/(cv_img.max()-cv_img.min())*255).astype(np.uint8)
                    if len(cv_img.shape) == 2: h, w = cv_img.shape; qimg = QImage(cv_img.data, w, h, w, QImage.Format.Format_Grayscale8)
                    else: cv_img = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB); h, w, ch = cv_img.shape; qimg = QImage(cv_img.data, w, h, ch * w, QImage.Format.Format_RGB888)
                    pixmap = QPixmap.fromImage(qimg)
                else: pixmap = QPixmap(ruta)
                
                self.pixmaps_globales = {"Original": pixmap, "Filtrada": None, "Esqueleto": None}
                
                cur.execute("""
                    SELECT bbox_x, bbox_y, bbox_w, bbox_h, crop_path, filtro_clahe, filtro_gauss, filtro_otsu, filtro_ruido, areas_eliminadas
                    FROM Microglia WHERE id_analisis = ?
                """, (id_an,))
                boxes_db = cur.fetchall()
                if boxes_db:
                    boxes = []
                    for (bx, by, bw, bh, cp, fc, fg, fo, fr, ae) in boxes_db:
                        b = {
                            "x": bx, "y": by, "w": bw, "h": bh,
                            "crop_path": cp.replace("\\", "/") if cp else "",
                            "offsets": {"clahe": fc or 0, "gauss": fg or 0, "otsu": fo or 0, "ruido": fr or 0},
                            "removal_areas": json.loads(ae) if ae else []
                        }
                        boxes.append(b)
                    self.visor_imagen.set_image_and_boxes(pixmap, boxes)
                else:
                    self.visor_imagen.set_image_and_boxes(pixmap, [])
                self.metricas_reporte = []
                
                # Reconstruir las imágenes procesadas Filtrada y Esqueleto para las miniaturas
                pix_f = self.construir_imagen_global("filtradas")
                if pix_f:
                    self.pixmaps_globales["Filtrada"] = pix_f
                
                pix_e = self.construir_imagen_global("esqueletos")
                if pix_e:
                    self.pixmaps_globales["Esqueleto"] = pix_e
                
                self.actualizar_estado_flujo(5)
                self.combo_vista.blockSignals(True)
                self.combo_vista.setCurrentText("Métricas")
                self.combo_vista.blockSignals(False)
                # Mostrar de inmediato la pantalla de Métricas directamente
                self.mostrar_tabla_metricas()
                self.actualizar_botones_navegacion()
                if not getattr(self, "reporte_validado_cargado", False):
                    if not tiene_correcciones:
                        self.mostrar_notificacion("Reporte Cargado", "Última imagen completada. Por favor, añade una nueva imagen para continuar el reporte.", "info")

        except Exception as e: self.mostrar_notificacion("Error", f"Fallo al cargar reporte: {e}", "error")
        finally:
            self.cargando_reporte = False
            conn.close()


    def cerrar_reporte_actual(self):
        """Cierra el reporte actual y restablece la interfaz al estado inicial."""
        self.id_reporte_actual = None
        self.id_analisis_actual = None
        self.ruta_imagen_actual = None
        self.pixmaps_globales = {"Original": None, "Filtrada": None, "Esqueleto": None}
        self.crops_en_memoria = {}
        self.crops_filtrados_temp = {}
        self.metadatos_imagen = {"campo": "", "tiempo": ""}
        self.metricas_reporte = []
        self.reporte_finalizado_actual = False
        self.paso_actual = 0
        self.metricas_extraidas_ciclo_actual = False
        self.reporte_validado_cargado = False
        
        # Restablecer combo de vistas
        self.combo_vista.blockSignals(True)
        while self.combo_vista.count() > 1:
            self.combo_vista.removeItem(1)
        self.combo_vista.setCurrentIndex(0)
        self.combo_vista.setEnabled(False)
        self.combo_vista.blockSignals(False)
        
        self.comentarios_correccion = ""
        if hasattr(self, "btn_observaciones"):
            self.btn_observaciones.hide()
            
        # Ocultar botones de corregir y herramientas
        self.btn_corregir_filtrado.hide()
        self.btn_herramienta_caja.hide()
        self.btn_herramienta_eliminar.hide()
        
        # Limpiar visor de imágenes
        self.visor_imagen.original_pixmap = None
        self.visor_imagen.boxes = []
        self.visor_imagen._current_pixmap = None
        self.visor_imagen.setText("Sube una imagen .tiff para empezar el análisis...")
        self.visor_imagen.setStyleSheet("border: 2px dashed #aaa; background-color: #f0f0f0; font-size: 18px; color: #666;")
        self.visor_imagen.update()
        self.actualizar_etiqueta_conteo(0)
        
        # Limpiar tabla de métricas y restablecer visor activo
        if hasattr(self, "tabla_metricas"):
            self.tabla_metricas.setRowCount(0)
        if hasattr(self, "stacked_visor"):
            self.stacked_visor.setCurrentIndex(0)
        
        # Restablecer estado de los botones laterales del flujo
        self.actualizar_estado_flujo(0)


    def actualizar_etiqueta_conteo(self, conteo): self.lbl_info_conteo.setText(f"Microglías detectadas: {conteo}")

    def cargar_imagen(self):
        dialogo_intermedio = DialogoCargarImagen(self)
        if dialogo_intermedio.exec():
            ruta_archivo = dialogo_intermedio.ruta_seleccionada
            campo = dialogo_intermedio.campo_val
            tiempo = dialogo_intermedio.tiempo_val
            
            # Check for duplicates in current session report
            for item in self.metricas_reporte:
                if item["campo"] == campo and item["tiempo"] == tiempo:
                    from vistas.utilidades import DialogoConfirmacion
                    diag = DialogoConfirmacion("Advertencia de Duplicado", f"Ya existen métricas guardadas para el Campo '{campo}' y Tiempo '{tiempo}'.\n\n¿Deseas continuar de todos modos?")
                    if not diag.exec():
                        return # User cancelled
                    break
                    
            # --- LIMPIEZA SEGURA Y GARANTIZADA DE ESTADO PREVIO ---
            if self.reporte_finalizado_actual:
                self.metricas_reporte.clear()
                self.reporte_finalizado_actual = False
                
            self.stacked_visor.setCurrentIndex(0)
            self.visor_imagen.set_image_and_boxes(None, [])
            self.ruta_imagen_actual = None
            self.id_analisis_actual = None  # ¡Garantiza que se creará un nuevo análisis y no se sobrescribirá el anterior!
            self.pixmaps_globales = {"Original": None, "Filtrada": None, "Esqueleto": None}
            self.combo_vista.blockSignals(True)
            self.combo_vista.clear()
            self.combo_vista.blockSignals(False)
            self.metricas_extraidas_ciclo_actual = False
            self.paso_actual = 0  # Pasamos temporalmente al paso 0 para la carga
            self.id_analisis_actual = None  # CRUCIAL: Resetear el ID para que guarde un nuevo registro en BD
            self.crops_en_memoria.clear()
            self.crops_filtrados_temp.clear()
            # -----------------------------------------------------
            
            self.metadatos_imagen["campo"] = campo; self.metadatos_imagen["tiempo"] = tiempo
            if ruta_archivo:
                self.ruta_imagen_actual = ruta_archivo; pixmap = QPixmap(ruta_archivo)
                if pixmap.isNull():
                    try:
                        import cv2; import numpy as np
                        try:
                            with open(ruta_archivo, "rb") as f:
                                file_bytes = np.frombuffer(f.read(), dtype=np.uint8)
                            cv_img = cv2.imdecode(file_bytes, cv2.IMREAD_UNCHANGED)
                        except Exception:
                            cv_img = None
                        if cv_img is not None:
                            if cv_img.dtype == np.uint16: cv_img = ((cv_img - cv_img.min()) / (cv_img.max() - cv_img.min()) * 255).astype(np.uint8)
                            if len(cv_img.shape) == 2: h, w = cv_img.shape; qimg = QImage(cv_img.data, w, h, w, QImage.Format.Format_Grayscale8)
                            else: cv_img = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB); h, w, ch = cv_img.shape; qimg = QImage(cv_img.data, w, h, ch * w, QImage.Format.Format_RGB888)
                            pixmap = QPixmap.fromImage(qimg)
                    except Exception as e: logging.error(f"Error al cargar imagen: {e}")
                if not pixmap.isNull():
                    self.pixmaps_globales["Original"] = pixmap
                    self.pixmaps_globales["Filtrada"] = None
                    self.pixmaps_globales["Esqueleto"] = None
                    self.btn_herramienta_caja.setChecked(False)
                    self.btn_herramienta_eliminar.setChecked(False)
                    self.visor_imagen.current_tool = "pointer"
                    self.btn_bloquear_zoom.setChecked(False)
                    self.reset_zoom()
                    self.visor_imagen.set_image_and_boxes(pixmap, [])
                    self.actualizar_estado_flujo(1)
                    self.combo_vista.blockSignals(True)
                    self.combo_vista.clear()
                    self.combo_vista.addItem("Original")
                    self.combo_vista.setCurrentText("Original")
                    self.combo_vista.blockSignals(False)
                    self.actualizar_botones_navegacion()
                    self.visor_imagen.view_mode = "Original"
                    self.mostrar_notificacion("Imagen cargada", "Imagen lista para el análisis.", "info")
                else: self.mostrar_notificacion("Error", "El archivo está corrupto o no es válido.", "error")

    def cambiar_vista_global(self, texto_vista):
        if not texto_vista:
            return
        if texto_vista == "Métricas":
            self.mostrar_tabla_metricas()
            self.sld_nivel_zoom.setEnabled(False)
            self.btn_zoom_reset.setEnabled(False)
            self.btn_bloquear_zoom.setEnabled(False)
            self.actualizar_botones_navegacion()
            return
            
        self.stacked_visor.setCurrentIndex(0)
        
        # Restaurar controles de zoom para las vistas de imágenes según el paso actual
        zoom_activado = (self.paso_actual >= 1)
        self.sld_nivel_zoom.setEnabled(zoom_activado)
        self.btn_zoom_reset.setEnabled(zoom_activado)
        self.btn_bloquear_zoom.setEnabled(zoom_activado)
        
        pixmap_guardado = self.pixmaps_globales.get(texto_vista)
        if pixmap_guardado:
            self.visor_imagen.set_view_mode(texto_vista, pixmap_guardado)
            self.actualizar_botones_navegacion()
        else:
            self.mostrar_notificacion("Aviso", f"Aún no has generado el paso: {texto_vista}.", "warning")
            self.combo_vista.blockSignals(True)
            self.combo_vista.setCurrentText(self.visor_imagen.view_mode)
            self.combo_vista.blockSignals(False)

    def obtener_metricas_reporte_desde_bd(self):
        from bd.database import conectar
        import os
        conn = conectar()
        cur = conn.cursor()
        
        cur.execute("""
            SELECT A.id_analisis, I.ruta_archivo, I.campo, I.tiempo_muestra
            FROM Analisis A
            JOIN Imagen I ON A.id_imagen = I.id_imagen
            WHERE A.id_reporte = ? AND A.paso_actual >= 5
            ORDER BY A.id_analisis ASC
        """, (self.id_reporte_actual,))
        analisis_list = cur.fetchall()
        
        resultado = []
        for id_an, ruta, campo, tiempo in analisis_list:
            cur.execute("""
                SELECT lineas, puntos_union, puntos_finales, voxeles_union, voxeles_losa,
                       longitud_promedio_ramas, uniones_triples, uniones_cuadruples,
                       longitud_maxima_rama, ruta_mas_larga
                FROM Microglia
                WHERE id_analisis = ?
                ORDER BY id_microglia ASC
            """, (id_an,))
            micros = cur.fetchall()
            
            metricas_imagen = []
            for m in micros:
                metricas_imagen.append({
                    "lines": m[0] or 0,
                    "junction points": m[1] or 0,
                    "end points": m[2] or 0,
                    "junction voxels": m[3] or 0,
                    "slab voxels": m[4] or 0,
                    "average branch length": m[5] or 0.0,
                    "triple points": m[6] or 0,
                    "quadruple points": m[7] or 0,
                    "maximum branch length": m[8] or 0.0,
                    "longest shortest path": m[9] or 0.0
                })
                
            resultado.append({
                "campo": campo or "Sin Campo",
                "tiempo": tiempo or "Sin Tiempo",
                "nombre_imagen": os.path.basename(ruta) if ruta else "Imagen Sin Nombre",
                "metricas": metricas_imagen
            })
            
        conn.close()
        return resultado

    def mostrar_tabla_metricas(self):
        self.stacked_visor.setCurrentIndex(1)
        
        # 1. Actualizar títulos de forma dinámica según el estado del reporte
        nombre_archivo = os.path.basename(self.ruta_imagen_actual) if self.ruta_imagen_actual else "Sin Imagen"
        if self.paso_actual == 5 or getattr(self, 'reporte_finalizado_actual', False):
            self.lbl_titulo_tabla.setText("<b>Último Procedimiento Realizado (Contexto del Análisis Anterior)</b>")
            self.lbl_subtitulo_imagen.setText(f"Imagen analizada anteriormente: <span style='color: #0969da; font-weight: bold;'>{nombre_archivo}</span>")
        else:
            self.lbl_titulo_tabla.setText("<b>Métricas Morfológicas de Microglías (Reporte de Sesión)</b>")
            self.lbl_subtitulo_imagen.setText(f"Imagen en análisis activo: <span style='color: #2da44e; font-weight: bold;'>{nombre_archivo}</span>")
            
        # Actualizar miniaturas de imágenes del procedimiento con suavizado avanzado
        if self.pixmaps_globales.get("Original") is not None:
            thumb = self.pixmaps_globales["Original"].scaled(220, 140, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            self.lbl_thumb_original.setPixmap(thumb)
        else:
            self.lbl_thumb_original.setText("Sin Imagen")

        if self.pixmaps_globales.get("Filtrada") is not None:
            thumb = self.pixmaps_globales["Filtrada"].scaled(220, 140, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            self.lbl_thumb_filtrada.setPixmap(thumb)
        else:
            self.lbl_thumb_filtrada.setText("Sin Imagen")

        if self.pixmaps_globales.get("Esqueleto") is not None:
            thumb = self.pixmaps_globales["Esqueleto"].scaled(220, 140, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            self.lbl_thumb_esqueleto.setPixmap(thumb)
        else:
            self.lbl_thumb_esqueleto.setText("Sin Imagen")
        
        # 2. Limpiar y poblar la tabla
        self.tabla_metricas.clearContents()
        
        # Columnas a mostrar
        columnas = [
            "Campo", "Tiempo", "No. Microglía", "Lines", "Junction Points", "End Points", 
            "Junction Voxels", "Slab Voxels", "Avg. Branch Length", 
            "Triple Points", "Quadruple Points", "Max Branch Length", "Longest Shortest Path"
        ]
        self.tabla_metricas.setColumnCount(len(columnas))
        self.tabla_metricas.setHorizontalHeaderLabels(columnas)
        
        # Recopilar todas las filas de la sesión actual leyendo de la BD
        filas = []
        metricas_bd = self.obtener_metricas_reporte_desde_bd()
        total_microglias = 0
        sum_junction_points = 0
        sum_end_points = 0
        sum_branch_length = 0.0
        count_microglias_with_branch_len = 0
        
        for img_data in metricas_bd:
            campo = str(img_data.get("campo", "Sin Campo"))
            tiempo = str(img_data.get("tiempo", "Sin Tiempo"))
            
            for idx, met in enumerate(img_data.get("metricas", []), start=1):
                # Extraer valores numéricos
                lines = met.get("lines", 0)
                junc_pts = met.get("junction points", 0)
                end_pts = met.get("end points", 0)
                junc_vx = met.get("junction voxels", 0)
                slab_vx = met.get("slab voxels", 0)
                avg_branch_len = met.get("average branch length", 0.0)
                triple_pts = met.get("triple points", 0)
                quad_pts = met.get("quadruple points", 0)
                max_branch_len = met.get("maximum branch length", 0.0)
                longest_path = met.get("longest shortest path", 0.0)
                
                # Para resúmenes
                total_microglias += 1
                sum_junction_points += junc_pts
                sum_end_points += end_pts
                if isinstance(avg_branch_len, (int, float)):
                    sum_branch_length += avg_branch_len
                    count_microglias_with_branch_len += 1
                
                # Dar formato
                f_avg_len = f"{avg_branch_len:.2f}" if isinstance(avg_branch_len, (int, float)) else str(avg_branch_len)
                f_max_len = f"{max_branch_len:.2f}" if isinstance(max_branch_len, (int, float)) else str(max_branch_len)
                f_longest = f"{longest_path:.2f}" if isinstance(longest_path, (int, float)) else str(longest_path)
                
                filas.append([
                    campo, tiempo, str(idx), str(lines), str(junc_pts), str(end_pts),
                    str(junc_vx), str(slab_vx), f_avg_len, str(triple_pts), str(quad_pts),
                    f_max_len, f_longest
                ])
                
        self.tabla_metricas.setRowCount(len(filas))
        
        for row_idx, row_data in enumerate(filas):
            for col_idx, val in enumerate(row_data):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.tabla_metricas.setItem(row_idx, col_idx, item)
                
        # Auto-ajustar columnas
        self.tabla_metricas.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        for c in range(len(columnas)):
            self.tabla_metricas.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeMode.ResizeToContents)
            
        # Actualizar tarjetas de resumen (KPIs)
        avg_junc = sum_junction_points / total_microglias if total_microglias > 0 else 0.0
        avg_end = sum_end_points / total_microglias if total_microglias > 0 else 0.0
        avg_len = sum_branch_length / count_microglias_with_branch_len if count_microglias_with_branch_len > 0 else 0.0
        
        self.val_total.setText(str(total_microglias))
        self.val_junc.setText(f"{avg_junc:.1f}")
        self.val_ends.setText(f"{avg_end:.1f}")
        self.val_len.setText(f"{avg_len:.2f} px")

    def anterior_vista_global(self):
        idx = self.combo_vista.currentIndex()
        if idx > 0:
            self.combo_vista.setCurrentIndex(idx - 1)

    def siguiente_vista_global(self):
        # Si estamos en la vista de esqueleto en el paso 4, avanzar ">" ejecuta la extracción de métricas
        if not getattr(self, "reporte_validado_cargado", False):
            if self.paso_actual == 4 and self.combo_vista.currentText() == "Esqueleto":
                self.obtener_metricas()
                return
            
        idx = self.combo_vista.currentIndex()
        if idx < self.combo_vista.count() - 1:
            self.combo_vista.setCurrentIndex(idx + 1)

    def actualizar_botones_navegacion(self):
        # Actualizar comentarios de proceso
        if hasattr(self, "_actualizar_tooltip_comentarios"):
            self._actualizar_tooltip_comentarios()

        # Si es modo revisión/validación, habilitar navegación libre entre los pasos cargados
        if getattr(self, "reporte_validado_cargado", False):
            if hasattr(self, "actualizar_nav_en_validacion") and self.actualizar_nav_en_validacion():
                return

        # Si ya se sacaron las métricas (Paso 5)...
        if self.paso_actual >= 5:
            # Caso normal: bloquear navegación
            self.btn_sig_global.show()
            self.btn_sig_global.setEnabled(False)
            self.btn_ant_global.show()
            self.btn_ant_global.setEnabled(False)
            return

        self.btn_sig_global.show()
        self.btn_ant_global.show()
        
        texto_vista = self.combo_vista.currentText()
        
        if texto_vista == "Original":
            # Original va a tener activados > y eliminado <
            self.btn_ant_global.setEnabled(False)
            self.btn_sig_global.setEnabled(self.combo_vista.count() > 1)
            
        elif texto_vista == "Filtrada":
            # Filtrado < y > siempre y cuando ya este el esqueletizado
            self.btn_ant_global.setEnabled(True)
            tiene_esqueleto = (self.combo_vista.findText("Esqueleto") >= 0 or self.paso_actual >= 4)
            self.btn_sig_global.setEnabled(tiene_esqueleto)
            
        elif texto_vista == "Esqueleto":
            # Esqueletizado tendra < activado y > desactivado o mejor eliminalo
            self.btn_ant_global.setEnabled(True)
            self.btn_sig_global.setEnabled(False)
            
        else:
            idx = self.combo_vista.currentIndex()
            self.btn_ant_global.setEnabled(idx > 0)
            self.btn_sig_global.setEnabled(idx < self.combo_vista.count() - 1)

    def actualizar_elementos_combo_vista(self):
        self.combo_vista.blockSignals(True)
        vista_seleccionada = self.combo_vista.currentText()
        self.combo_vista.clear()
        
        # El orden estricto de opciones es: Original -> Filtrada -> Esqueleto -> Métricas
        self.combo_vista.addItem("Original")
        
        if self.pixmaps_globales.get("Filtrada") is not None:
            self.combo_vista.addItem("Filtrada")
            
        if self.pixmaps_globales.get("Esqueleto") is not None:
            self.combo_vista.addItem("Esqueleto")
            
        if self.paso_actual >= 5:
            self.combo_vista.addItem("Métricas")
            
        items = [self.combo_vista.itemText(i) for i in range(self.combo_vista.count())]
        if vista_seleccionada in items:
            self.combo_vista.setCurrentText(vista_seleccionada)
        else:
            self.combo_vista.setCurrentIndex(self.combo_vista.count() - 1)
            
        self.combo_vista.blockSignals(False)
        
        # Sincronizar explícitamente el visor de imágenes o panel con la selección actual si hay un análisis en curso
        if self.paso_actual > 0:
            self.cambiar_vista_global(self.combo_vista.currentText())

    def cerrar_sesion(self):
        from vistas.login import VentanaLogin
        self.ventana_login = VentanaLogin(); self.ventana_login.setObjectName("ventana_login"); self.ventana_login.show(); self.close()

    def abrir_config_red(self):
        from red.config_gui import DialogoConfigRed
        dialogo = DialogoConfigRed(self)
        dialogo.exec()

    def execute_microglia_counting(self):
        if not self.ruta_imagen_actual: self.mostrar_notificacion("Advertencia", "Por favor, carga una imagen primero.", "warning"); return
        
        # Si ya se ejecutó y estamos en un paso avanzado, dar opción de sólo volver a editar o reprocesar con IA
        if self.paso_actual >= 2:
            from PyQt6.QtWidgets import QMessageBox
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Icon.Question)
            msg.setWindowTitle("Regresar a Detección")
            msg.setText("¿Qué deseas hacer al regresar al paso de Detección (Conteo)?")
            
            btn_edit = msg.addButton("Editar detecciones manualmente", QMessageBox.ButtonRole.ActionRole)
            btn_yolo = msg.addButton("Volver a procesar con IA (YOLO)", QMessageBox.ButtonRole.ActionRole)
            btn_cancel = msg.addButton("Cancelar", QMessageBox.ButtonRole.RejectRole)
            
            msg.exec()
            clicked = msg.clickedButton()
            if clicked == btn_edit:
                self.actualizar_estado_flujo(2)
                self.combo_vista.blockSignals(True)
                self.combo_vista.setCurrentText("Original")
                self.combo_vista.blockSignals(False)
                self.cambiar_vista_global("Original")
                return
            elif clicked == btn_cancel or clicked is None:
                return
                
        dialogo = DialogoCarga("Aplicando conteo...\nPor favor, espera.", self); dialogo.show()
        from PyQt6.QtWidgets import QApplication; QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor); QApplication.processEvents()
        try:
            output_dir = os.path.join(os.getcwd(), "analisis_resultados")
            from red.config import es_cliente
            if es_cliente():
                from red.cliente import cliente_ejecutar_conteo_ia
                crops_folder, count, boxes_data = cliente_ejecutar_conteo_ia(self.ruta_imagen_actual, base_output_folder=output_dir)
            else:
                from procesamiento.deteccion import ejecutar_conteo_ia
                crops_folder, count, boxes_data = ejecutar_conteo_ia(self.ruta_imagen_actual, base_output_folder=output_dir)
            self.visor_imagen.set_image_and_boxes(self.pixmaps_globales["Original"], boxes_data)

            dialogo.close(); QApplication.restoreOverrideCursor(); self.actualizar_estado_flujo(2); self.mostrar_notificacion("2. Detección", f"Se detectaron {count} posibles microglías.\n\nUsa las herramientas superiores si necesitas agregar o eliminar selecciones.", "info")
        except Exception as e: dialogo.close(); QApplication.restoreOverrideCursor(); self.mostrar_notificacion("Error", str(e), "error")

    def agregar_microglia_manual(self, x, y, w, h):
        if not self.ruta_imagen_actual: return
        area_nueva = w * h
        for box in self.visor_imagen.boxes:
            bx, by, bw, bh = box["x"], box["y"], box["w"], box["h"]; ix1 = max(x, bx); iy1 = max(y, by); ix2 = min(x + w, bx + bw); iy2 = min(y + h, by + bh); iw_inter = max(0, ix2 - ix1); ih_inter = max(0, iy2 - iy1); inter_area = iw_inter * ih_inter
            if inter_area > 0:
                ioa_nueva = inter_area / area_nueva; ioa_existente = inter_area / (bw * bh)
                if ioa_nueva >= 0.40 or ioa_existente >= 0.40: self.mostrar_notificacion("Acción bloqueada", "No puedes crear una microglía que esté contenida dentro de otra existente o que se superponga fuertemente.", "warning"); return
        base_name = Path(self.ruta_imagen_actual).stem; crops_folder = os.path.join(os.getcwd(), "analisis_resultados", base_name, "crops"); os.makedirs(crops_folder, exist_ok=True)
        orig_pixmap = self.pixmaps_globales["Original"]
        if not orig_pixmap: return
        from procesamiento.deteccion import recortar_y_guardar_manual
        ruta_guardado, nombre_archivo = recortar_y_guardar_manual(orig_pixmap, x, y, w, h, crops_folder)
        
        from red.config import es_cliente
        if es_cliente():
            from red.cliente import cliente_subir_archivo
            cliente_subir_archivo(ruta_guardado)

        nueva_caja = {"x": x, "y": y, "w": w, "h": h, "crop_path": ruta_guardado, "offsets": {"clahe": 0, "gauss": 0, "otsu": 0, "ruido": 0}, "removal_areas": []}

        self.visor_imagen.boxes.append(nueva_caja); self.visor_imagen.draw_current_state(); self.actualizar_etiqueta_conteo(len(self.visor_imagen.boxes))
        self.save_current_progress(mostrar_notif=False)


    def construir_imagen_global(self, carpeta_origen):
        import cv2; import numpy as np
        orig_pixmap = self.pixmaps_globales["Original"]
        if not orig_pixmap: return None
        orig_w = orig_pixmap.width(); orig_h = orig_pixmap.height()
        lienzo = np.zeros((orig_h, orig_w), dtype=np.uint8)
        base_name = Path(self.ruta_imagen_actual).stem
        base_dir = os.path.join(os.getcwd(), "analisis_resultados", base_name)
        
        for box in self.visor_imagen.boxes:
            x, y, w, h = int(box["x"]), int(box["y"]), int(box["w"]), int(box["h"])
            nombre_archivo = os.path.basename(box["crop_path"])
            ruta_recorte = os.path.join(base_dir, carpeta_origen, nombre_archivo)
            
            from red.config import es_cliente
            if es_cliente():
                from red.cliente import asegurar_archivo_local
                asegurar_archivo_local(ruta_recorte)
            
            # Salvaguarda: si no existe en la carpeta relativa actual, intentar por ruta absoluta del crop
            if not os.path.exists(ruta_recorte) and "crop_path" in box:
                ruta_recorte = box["crop_path"].replace("\\", "/").replace("/crops/", f"/{carpeta_origen}/")
                if es_cliente():
                    asegurar_archivo_local(ruta_recorte)
                
            if os.path.exists(ruta_recorte):
                try:
                    with open(ruta_recorte, "rb") as f:
                        file_bytes = bytearray(f.read())
                    img_array = np.asarray(file_bytes, dtype=np.uint8)
                    recorte = cv2.imdecode(img_array, cv2.IMREAD_GRAYSCALE)
                    if recorte is not None:
                        rh, rw = recorte.shape
                        y_fin = min(y + rh, orig_h)
                        x_fin = min(x + rw, orig_w)
                        h_real = y_fin - y
                        w_real = x_fin - x
                        lienzo[y:y_fin, x:x_fin] = recorte[:h_real, :w_real]
                except Exception:
                    pass
        qimg = QImage(lienzo.data, orig_w, orig_h, orig_w, QImage.Format.Format_Grayscale8).copy()
        return QPixmap.fromImage(qimg)

    def construir_imagen_global_memoria(self):
        import numpy as np
        orig_pixmap = self.pixmaps_globales["Original"]; orig_w = orig_pixmap.width(); orig_h = orig_pixmap.height(); lienzo = np.zeros((orig_h, orig_w), dtype=np.uint8)
        for box in self.visor_imagen.boxes:
            x, y, w, h = int(box["x"]), int(box["y"]), int(box["w"]), int(box["h"])
            nombre_archivo = os.path.basename(box["crop_path"])
            recorte = self.crops_filtrados_temp.get(nombre_archivo)
            if recorte is not None:
                rh, rw = recorte.shape; y_fin = min(y + rh, orig_h); x_fin = min(x + rw, orig_w); h_real = y_fin - y; w_real = x_fin - x
                lienzo[y:y_fin, x:x_fin] = recorte[:h_real, :w_real]
        qimg = QImage(lienzo.data, orig_w, orig_h, orig_w, QImage.Format.Format_Grayscale8).copy()
        return QPixmap.fromImage(qimg)

    def ejecutar_filtrado(self):
        if not self.ruta_imagen_actual or not self.visor_imagen.boxes: self.mostrar_notificacion("Advertencia", "Aplica el conteo primero.", "warning"); return
        
        # Guardar paso previo para poder regresar al cancelar
        self.paso_previo = self.paso_actual
        
        import cv2; import numpy as np
        self.crops_en_memoria.clear()
        self.crops_filtrados_temp.clear()
        
        from PyQt6.QtWidgets import QApplication
        dialogo = DialogoCarga("Cargando imágenes a memoria...\nPor favor, espera.", self); dialogo.show(); QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor); QApplication.processEvents()
        
        for box in self.visor_imagen.boxes:
            crop_path = box["crop_path"]
            if os.path.exists(crop_path):
                with open(crop_path, "rb") as f: file_bytes = bytearray(f.read())
                img_array = np.asarray(file_bytes, dtype=np.uint8)
                img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
                if img is not None:
                    nombre = os.path.basename(crop_path)
                    self.crops_en_memoria[nombre] = img
                    
        dialogo.close(); QApplication.restoreOverrideCursor()
        
        self.frame_filtros.show()
        for btn in [self.btn_cargar, self.btn_conteo, self.btn_filtrar, self.btn_ramas, self.btn_obtener_metricas, self.btn_agregar_imagen_reporte, self.btn_descargar_reporte, self.btn_finalizar_reporte]: btn.setEnabled(False)
        
        # Ocultar herramientas manuales al iniciar el filtrado
        self.btn_herramienta_caja.setChecked(False)
        self.btn_herramienta_eliminar.setChecked(False)
        self.btn_herramienta_caja.hide()
        self.btn_herramienta_eliminar.hide()
        self.visor_imagen.current_tool = "pointer"
        
        self.combo_vista.setEnabled(True)
        items_combo = [self.combo_vista.itemText(i) for i in range(self.combo_vista.count())]
        if "Previsualización" not in items_combo:
            self.combo_vista.addItem("Previsualización")
        
        self.combo_vista.blockSignals(True)
        self.combo_vista.setCurrentText("Previsualización")
        self.combo_vista.blockSignals(False)
        self.actualizar_botones_navegacion()
        self.visor_imagen.view_mode = "Previsualización"
        
        self.previsualizar_filtrado()

    def previsualizar_filtrado(self, *args):
        if not self.crops_en_memoria: return
        g_clahe_clip = self.sld_clahe.value()
        g_k_val = self.sld_gauss.value()
        g_otsu_offset = self.sld_otsu.value()
        g_ruido = self.sld_ruido.value()
        
        # Mapa de nombre a info (offsets y areas) para eficiencia
        mapa_info = {}
        for box in self.visor_imagen.boxes:
            nombre = os.path.basename(box["crop_path"])
            mapa_info[nombre] = {
                "offsets": box.get("offsets", {"clahe":0, "gauss":0, "otsu":0, "ruido":0}),
                "removal_areas": box.get("removal_areas", [])
            }

        from procesamiento.filtrado import procesar_crop_individual
        for nombre, img in self.crops_en_memoria.items():
            info = mapa_info.get(nombre, {"offsets": {"clahe":0, "gauss":0, "otsu":0, "ruido":0}, "removal_areas": []})
            offsets = info["offsets"]
            bin_img = procesar_crop_individual(img, g_clahe_clip, g_k_val, g_otsu_offset, g_ruido, offsets, info["removal_areas"])
            self.crops_filtrados_temp[nombre] = bin_img

            
        pixmap_filtrada = self.construir_imagen_global_memoria()
        self.pixmaps_globales["Previsualización"] = pixmap_filtrada
        self.visor_imagen.set_view_mode("Previsualización", pixmap_filtrada)


    def confirmar_filtrado(self):
        base_name = Path(self.ruta_imagen_actual).stem; filtradas_dir = os.path.join(os.getcwd(), "analisis_resultados", base_name, "filtradas"); os.makedirs(filtradas_dir, exist_ok=True); import cv2; count = 0
        from PyQt6.QtWidgets import QApplication
        try:
            dialogo = DialogoCarga("Guardando filtros aplicados...\nPor favor, espera.", self); dialogo.show(); QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor); QApplication.processEvents()
            for nombre, bin_img in self.crops_filtrados_temp.items():
                out_path = os.path.join(filtradas_dir, nombre)
                is_success, im_buf_arr = cv2.imencode(".png", bin_img)
                if is_success: 
                    im_buf_arr.tofile(out_path)
                    count += 1
                    from red.config import es_cliente
                    if es_cliente():
                        from red.cliente import cliente_subir_archivo
                        cliente_subir_archivo(out_path)
            dialogo.close(); QApplication.restoreOverrideCursor()
            
            if count > 0:
                self.pixmaps_globales["Filtrada"] = self.pixmaps_globales["Previsualización"]
                self.combo_vista.blockSignals(True)
                items_combo = [self.combo_vista.itemText(i) for i in range(self.combo_vista.count())]
                if "Filtrada" not in items_combo: self.combo_vista.addItem("Filtrada")
                idx = self.combo_vista.findText("Previsualización")
                if idx >= 0: self.combo_vista.removeItem(idx)
                self.combo_vista.setCurrentText("Filtrada")
                self.combo_vista.blockSignals(False)
                self.visor_imagen.view_mode = "Filtrada"
                
                self.frame_filtros.hide()
                self.actualizar_estado_flujo(3)
                self.mostrar_notificacion("3. Filtrado", f"Se aplicaron los filtros a {count} microglías.", "info")
            else: self.mostrar_notificacion("Error", "No se guardó ninguna imagen.", "error")
        except Exception as error: dialogo.close(); QApplication.restoreOverrideCursor(); self.mostrar_notificacion("Error", f"Falló el guardado: {str(error)}", "error")

    def cancelar_filtrado(self):
        self.crops_en_memoria.clear()
        self.crops_filtrados_temp.clear()
        self.frame_filtros.hide()
        
        self.combo_vista.blockSignals(True)
        idx = self.combo_vista.findText("Previsualización")
        if idx >= 0: self.combo_vista.removeItem(idx)
        
        # Volver a la vista previa del paso restaurado
        paso_restaurar = getattr(self, "paso_previo", 2)
        vista_restaurar = "Original"
        if paso_restaurar >= 3:
            vista_restaurar = "Filtrada"
        if paso_restaurar >= 4:
            vista_restaurar = "Esqueleto"
            
        self.combo_vista.setCurrentText(vista_restaurar)
        self.combo_vista.blockSignals(False)
        self.cambiar_vista_global(vista_restaurar)
        
        self.actualizar_estado_flujo(paso_restaurar)

    def mostrar_ramas_morfologia(self):
        if not self.ruta_imagen_actual or not self.visor_imagen.boxes: self.mostrar_notificacion("Advertencia", "Aplica el conteo y filtrado primero.", "warning"); return
        base_name = Path(self.ruta_imagen_actual).stem; filtradas_dir = os.path.join(os.getcwd(), "analisis_resultados", base_name, "filtradas"); esqueletos_dir = os.path.join(os.getcwd(), "analisis_resultados", base_name, "esqueletos"); os.makedirs(esqueletos_dir, exist_ok=True)
        from PyQt6.QtWidgets import QApplication; count = 0
        import cv2
        from red.config import es_cliente
        try:
            dialogo = DialogoCarga("Generando esqueletos topológicos...\nPor favor, espera.", self); dialogo.show(); QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor); QApplication.processEvents()
            for box in self.visor_imagen.boxes:
                crop_path = box["crop_path"]; nombre = os.path.basename(crop_path); fil_path = os.path.join(filtradas_dir, nombre)
                
                if es_cliente():
                    from red.cliente import asegurar_archivo_local, cliente_generar_esqueleto_de_archivo
                    asegurar_archivo_local(fil_path)
                    out_path = os.path.join(esqueletos_dir, nombre)
                    res_path = cliente_generar_esqueleto_de_archivo(fil_path, esqueletos_dir, nombre)
                    if res_path and os.path.exists(res_path):
                        count += 1
                else:
                    from procesamiento.esqueletizado import generar_esqueleto_de_archivo
                    if os.path.exists(fil_path):
                        skeleton_img = generar_esqueleto_de_archivo(fil_path)
                        if skeleton_img is not None:
                            out_path = os.path.join(esqueletos_dir, nombre)
                            is_success, im_buf_arr = cv2.imencode(".png", skeleton_img)
                            if is_success: im_buf_arr.tofile(out_path); count += 1
            dialogo.close(); QApplication.restoreOverrideCursor()
            if count > 0: pixmap_esqueleto = self.construir_imagen_global("esqueletos"); self.pixmaps_globales["Esqueleto"] = pixmap_esqueleto; self.actualizar_estado_flujo(4); self.combo_vista.setCurrentText("Esqueleto"); self.mostrar_notificacion("4. Esqueleto (Ramas)", f"Se generaron {count} esqueletos topológicos.\n\nYa puedes avanzar a obtener las métricas finales.", "info")
            else: self.mostrar_notificacion("Advertencia", "No se generaron esqueletos. Verifica la carpeta de filtrado.", "warning")
        except Exception as error: dialogo.close(); QApplication.restoreOverrideCursor(); self.mostrar_notificacion("Error de Procesamiento", f"Falló el cálculo:\n{str(error)}", "error")

    def corregir_filtrado(self):
        from vistas.utilidades import DialogoConfirmacion
        diag = DialogoConfirmacion("Corregir Filtrado", "¿Estás seguro de que deseas eliminar el filtrado y esqueletizado actual para volver a ajustar los parámetros?")
        if not diag.exec(): return
        
        self.pixmaps_globales["Filtrada"] = None
        self.pixmaps_globales["Esqueleto"] = None
        
        self.combo_vista.blockSignals(True)
        for label in ["Esqueleto", "Filtrada"]:
            idx = self.combo_vista.findText(label)
            if idx >= 0: self.combo_vista.removeItem(idx)
        self.combo_vista.setCurrentText("Original")
        self.combo_vista.blockSignals(False)
        self.cambiar_vista_global("Original")
        
        if self.ruta_imagen_actual:
            try:
                import shutil
                base_name = Path(self.ruta_imagen_actual).stem
                base_dir = os.path.join(os.getcwd(), "analisis_resultados", base_name)
                for sub in ["filtradas", "esqueletos"]:
                    folder = os.path.join(base_dir, sub)
                    if os.path.exists(folder): shutil.rmtree(folder); os.makedirs(folder, exist_ok=True)
            except Exception as e: logging.error(f"Error al limpiar carpetas: {e}")
            
        # Resetear sliders globales y offsets individuales
        self.sld_clahe.setValue(2)
        self.sld_gauss.setValue(5)
        self.sld_otsu.setValue(0)
        self.sld_ruido.setValue(50)
        for box in self.visor_imagen.boxes:
            box["offsets"] = {"clahe": 0, "gauss": 0, "otsu": 0, "ruido": 0}
            box["removal_areas"] = []
            box["esqueleto_modificado"] = False
            
        self.actualizar_estado_flujo(2)


        self.mostrar_notificacion("Corregir Filtrado", "Se han eliminado las fases anteriores. Puedes ajustar los filtros nuevamente.", "info")


    def obtener_metricas(self):
        if not self.ruta_imagen_actual or not self.visor_imagen.boxes:
            self.mostrar_notificacion("Advertencia", "No hay datos para extraer métricas.", "warning")
            return
            
        base_name = Path(self.ruta_imagen_actual).stem
        esqueletos_dir = os.path.join(os.getcwd(), "analisis_resultados", base_name, "esqueletos")
        
        if not os.path.exists(esqueletos_dir):
            self.mostrar_notificacion("Advertencia", "No se encontraron esqueletos generados.", "warning")
            return
            
        from procesamiento.metricas import extraer_metricas_esqueleto
        from PyQt6.QtWidgets import QApplication
        
        dialogo = DialogoCarga("Extrayendo métricas morfológicas...\nPor favor, espera.", self)
        dialogo.show()
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        QApplication.processEvents()
        
        metricas_imagen = []
        from red.config import es_cliente
        from bd.database import conectar
        conn = conectar()
        cur = conn.cursor()
        
        for box in self.visor_imagen.boxes:
            nombre = os.path.basename(box["crop_path"])
            out_path = os.path.join(esqueletos_dir, nombre)
            
            if es_cliente():
                from red.cliente import asegurar_archivo_local
                asegurar_archivo_local(out_path)
                
            if os.path.exists(out_path):
                try:
                    if es_cliente():
                        from red.cliente import cliente_extraer_metricas_esqueleto
                        met = cliente_extraer_metricas_esqueleto(out_path)
                    else:
                        met = extraer_metricas_esqueleto(out_path)
                    metricas_imagen.append(met)
                    
                    try:
                        cur.execute("""
                            UPDATE Microglia 
                            SET puntos_finales = ?, 
                                uniones_triples = ?, 
                                uniones_cuadruples = ?, 
                                longitud_promedio_ramas = ?, 
                                longitud_maxima_rama = ?, 
                                ruta_mas_larga = ?,
                                lineas = ?,
                                puntos_union = ?,
                                voxeles_union = ?,
                                voxeles_losa = ?
                            WHERE id_analisis = ? AND crop_path = ?
                        """, (
                            met.get("end points", 0),
                            met.get("triple points", 0),
                            met.get("quadruple points", 0),
                            met.get("average branch length", 0.0),
                            met.get("maximum branch length", 0.0),
                            met.get("longest shortest path", 0.0),
                            met.get("lines", 0),
                            met.get("junction points", 0),
                            met.get("junction voxels", 0),
                            met.get("slab voxels", 0),
                            self.id_analisis_actual,
                            box.get("crop_path", "")
                        ))
                    except Exception as db_err:
                        logging.error(f"Error actualizando DB con métricas: {db_err}")
                        
                except Exception as e:
                    logging.error(f"Error extrayendo métricas de {nombre}: {e}")
                    
        try:
            conn.commit()
        except Exception as e:
            logging.error(f"Error en commit de BD: {e}")
        finally:
            conn.close()
            
        dialogo.close()
        QApplication.restoreOverrideCursor()
        
        if not metricas_imagen:
            self.mostrar_notificacion("Error", "No se pudieron extraer métricas de ninguna microglía.", "error")
            return
            
        self.metricas_reporte = [] # Ya no se usa self.metricas_reporte en memoria
        
        # Limpiar marcas de esqueleto modificado para quitar el resaltado amarillo
        for box in self.visor_imagen.boxes:
            box["esqueleto_modificado"] = False
        self.visor_imagen.draw_current_state()
        
        self.metricas_extraidas_ciclo_actual = True
        self.actualizar_estado_flujo(5)
        
        self.combo_vista.blockSignals(True)
        self.combo_vista.setCurrentText("Métricas")
        self.combo_vista.blockSignals(False)
        
        self.mostrar_notificacion("5. Métricas", "Métricas extraídas y análisis completado exitosamente.", "info")
        self.mostrar_tabla_metricas()

    def agregar_imagen_reporte(self):
        if self.reporte_finalizado_actual:
            self.metricas_reporte.clear()
            self.reporte_finalizado_actual = False
            
        self.stacked_visor.setCurrentIndex(0)
        self.visor_imagen.set_image_and_boxes(None, [])
        self.ruta_imagen_actual = None
        self.id_analisis_actual = None
        self.pixmaps_globales = {"Original": None, "Filtrada": None, "Esqueleto": None}
        self.combo_vista.blockSignals(True)
        self.combo_vista.clear()
        self.combo_vista.blockSignals(False)
        
        self.metricas_extraidas_ciclo_actual = False
        self.actualizar_estado_flujo(0)
        self.mostrar_notificacion("Info", "Sesión lista para cargar otra imagen y agregar al reporte.", "info")

    def descargar_reporte(self):
        metricas_bd = self.obtener_metricas_reporte_desde_bd()
        if not metricas_bd:
            self.mostrar_notificacion("Advertencia", "No hay métricas acumuladas para descargar.", "warning")
            return
            
        # Depuración: imprimir estructura de datos
        print(f"DEBUG: metricas_bd len: {len(metricas_bd)}")
            
        from datetime import datetime; from PyQt6.QtWidgets import QFileDialog; from pathlib import Path
        fecha_str = datetime.now().strftime("%Y%m%d_%H%M"); default_name = f"Reporte_{fecha_str}.xlsx"
        
        from PyQt6.QtCore import QStandardPaths
        import os
        docs_dir = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.DocumentsLocation)
        base_dir = docs_dir if docs_dir and os.path.exists(docs_dir) else os.path.expanduser("~")
        default_path = os.path.join(base_dir, default_name)
        
        filepath, filter_selected = QFileDialog.getSaveFileName(self, "Guardar Reporte", default_path, "Excel Files (*.xlsx);;PDF Files (*.pdf);;Both Formats (*.xlsx *.pdf)")
        if not filepath: return
            
        try:
            reporte_por_tiempo = {}
            for img_data in metricas_bd:
                t = str(img_data.get("tiempo", "X HORA")).upper()
                if t not in reporte_por_tiempo: reporte_por_tiempo[t] = []
                reporte_por_tiempo[t].append(img_data)
                
            total_global_microglias = sum(len(img_data.get("metricas", [])) for img_data in metricas_bd)

            columnas_labels = ["No.", "Lines", "Junction Points", "End Points", "Junction Voxels", "Slab Voxels", "Avg. Branch Length", "Triple points", "Quadruple points", "Max Branch Length", "Longest Shortest path"]
            metric_keys = ["lines", "junction points", "end points", "junction voxels", "slab voxels", "average branch length", "triple points", "quadruple points", "maximum branch length", "longest shortest path"]

            save_xlsx = "Excel" in filter_selected or "Both" in filter_selected or filepath.endswith(".xlsx")
            save_pdf = "PDF" in filter_selected or "Both" in filter_selected or filepath.endswith(".pdf")

            if save_xlsx:
                xlsx_path = filepath if filepath.endswith(".xlsx") else str(Path(filepath).with_suffix(".xlsx"))
                import openpyxl; from openpyxl.styles import PatternFill, Font, Alignment; from openpyxl.utils import get_column_letter
                wb = openpyxl.Workbook(); wb.remove(wb.active)
                
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                imagen_bg_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                header_bg_fill = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
                light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                bold_font = Font(bold=True); header_black_font = Font(color="000000", bold=True); center_alignment = Alignment(horizontal="center", vertical="center")
                anchos_fijos = [9.3, 18.0, 22.6, 18.6, 22.6, 16.6, 26.6, 20.0, 26.6, 26.6, 32.0]
                
                for tiempo, lista_campos in reporte_por_tiempo.items():
                    ws = wb.create_sheet(title=tiempo[:31])
                    for col_idx, width in enumerate(anchos_fijos, start=1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = width
                    
                    row_idx = 1
                    for img_data in lista_campos:
                        # Fila 1: Campo
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(columnas_labels))
                        for c in range(1, len(columnas_labels) + 1):
                            ws.cell(row=row_idx, column=c).fill = yellow_fill
                        
                        campo_val = img_data.get("campo", "")
                        if isinstance(campo_val, dict): campo_val = str(campo_val)
                        cell_title = ws.cell(row=row_idx, column=1, value=f"Campo: {str(campo_val)}")
                        cell_title.font = bold_font; cell_title.alignment = center_alignment; row_idx += 1
                        
                        # Fila 2: Nombre de Imagen (Nueva)
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(columnas_labels))
                        for c in range(1, len(columnas_labels) + 1):
                            ws.cell(row=row_idx, column=c).fill = imagen_bg_fill
                        
                        nombre_img = img_data.get("nombre_imagen", "Desconocido")
                        cell_img = ws.cell(row=row_idx, column=1, value=f"Imagen: {nombre_img}")
                        cell_img.font = bold_font; cell_img.alignment = center_alignment; row_idx += 1
                        
                        for col_idx, label in enumerate(columnas_labels, start=1):
                            cell_h = ws.cell(row=row_idx, column=col_idx, value=label)
                            cell_h.font = header_black_font; cell_h.fill = header_bg_fill; cell_h.alignment = center_alignment
                        row_idx += 1
                        
                        for i, met in enumerate(img_data["metricas"], start=1):
                            cell_num = ws.cell(row=row_idx, column=1, value=i); cell_num.alignment = center_alignment
                            for col_idx, key in enumerate(metric_keys, start=2):
                                val = met.get(key, "")
                                if isinstance(val, (dict, list)): val = str(val)
                                cell_m = ws.cell(row=row_idx, column=col_idx, value=val); cell_m.alignment = center_alignment
                            if i % 2 != 0:
                                for c in range(1, len(columnas_labels) + 1):
                                    ws.cell(row=row_idx, column=c).fill = light_gray_fill
                            row_idx += 1
                        
                        # Totales individuales por cada Imagen/Campo en Excel (sin total acumulado global)
                        row_idx += 1
                        ws.cell(row=row_idx, column=1, value=f"Total de Microglías en {campo_val}:").font = bold_font
                        ws.cell(row=row_idx, column=2, value=len(img_data.get("metricas", []))).font = bold_font
                        row_idx += 1
                wb.save(xlsx_path)

            if save_pdf:
                pdf_path = filepath if filepath.endswith(".pdf") else str(Path(filepath).with_suffix(".pdf"))
                try:
                    from fpdf import FPDF
                    class PDFReport(FPDF):
                        def header(self):
                            self.set_font("Arial", "B", 14)
                            self.cell(0, 10, "Reporte de Métricas Morfológicas - Microglías", 0, 1, "C")
                            self.ln(5)

                    pdf = PDFReport(orientation="L", unit="mm", format="A4")
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf_widths = [12, 22, 26, 22, 26, 20, 31, 23, 31, 31, 36] 

                    for tiempo, lista_campos in reporte_por_tiempo.items():
                        pdf.add_page()
                        pdf.set_font("Arial", "B", 12)
                        pdf.cell(0, 10, f"TIEMPO: {tiempo}", 0, 1, "L")
                        
                        for img_data in lista_campos:
                            # Fila 1: Campo
                            pdf.set_fill_color(255, 255, 0); pdf.set_font("Arial", "B", 10)
                            pdf.cell(sum(pdf_widths), 8, f"Campo: {img_data['campo']}", 1, 1, "C", True)
                            
                            # Fila 2: Nombre de Imagen (Nueva)
                            pdf.set_fill_color(220, 230, 241); pdf.set_font("Arial", "B", 9)
                            nombre_img = img_data.get("nombre_imagen", "Desconocido")
                            pdf.cell(sum(pdf_widths), 8, f"Imagen: {nombre_img}", 1, 1, "C", True)
                            
                            pdf.set_fill_color(149, 179, 215); pdf.set_text_color(0, 0, 0); pdf.set_font("Arial", "B", 7)
                            for i, label in enumerate(columnas_labels):
                                pdf.cell(pdf_widths[i], 8, label, 1, 0, "C", True)
                            pdf.ln()
                            
                            pdf.set_font("Arial", "", 9); pdf.set_text_color(0, 0, 0)
                            for idx, met in enumerate(img_data["metricas"], start=1):
                                if idx % 2 != 0: pdf.set_fill_color(242, 242, 242)
                                else: pdf.set_fill_color(255, 255, 255)
                                
                                pdf.cell(pdf_widths[0], 7, str(idx), 1, 0, "C", True)
                                values = []
                                for k in metric_keys:
                                    v = met.get(k, "")
                                    values.append(str(v) if not isinstance(v, (int, float)) else v)
                                    
                                for i_v, val in enumerate(values):
                                    pdf.cell(pdf_widths[i_v+1], 7, str(val), 1, 0, "C", True)
                                pdf.ln()
                            # Totales individuales por cada Imagen
                            pdf.ln(2)
                            pdf.set_font("Arial", "B", 9)
                            pdf.cell(0, 8, f"Total de Microglias en {img_data['campo']}: {len(img_data.get('metricas', []))}", 0, 1, "L")
                            pdf.ln(5)
                    pdf.output(pdf_path)
                except Exception as e:
                    import logging
                    logging.error(f"Error generando PDF: {e}")

            self.mostrar_notificacion("Éxito", f"Reporte guardado en: {os.path.basename(filepath)}", "info")
            self.btn_finalizar_reporte.setEnabled(True)
        except Exception as error:
            self.mostrar_notificacion("Error", f"Falló la exportación: {str(error)}", "error")

    def finalizar_reporte(self):
        # Validar que todas las imágenes del reporte estén terminadas (Paso >= 5)
        if self.id_reporte_actual:
            from bd.database import conectar
            conn = conectar(); cur = conn.cursor()
            try:
                cur.execute("SELECT COUNT(*) FROM Analisis WHERE id_reporte = ? AND paso_actual < 5", (self.id_reporte_actual,))
                incompletas = cur.fetchone()[0] or 0
                if incompletas > 0:
                    self.mostrar_notificacion("Acción Bloqueada", "No puedes finalizar el reporte porque tienes imágenes con análisis inconclusos. Por favor, completa todas las imágenes antes de finalizar.", "warning")
                    return
            except Exception as e:
                print(f"Error al validar análisis incompletos: {e}")
            finally:
                conn.close()

        from vistas.utilidades import DialogoConfirmacion
        if self.rol == "Invitado":
            msg = "¿Finalizar reporte? En modo Invitado las métricas no se guardarán."
        else:
            msg = "¿Finalizar reporte? Podrás consultarlo en tu historial."
            
        if not DialogoConfirmacion("Finalizar Reporte", msg).exec(): return
        
        self.cerrar_reporte_actual()
        
        if self.rol == "Invitado":
            msg_final = "Reporte finalizado localmente. Pantalla restablecida."
        else:
            msg_final = "Reporte finalizado y guardado con éxito."
            
        self.mostrar_notificacion("Reporte Finalizado", msg_final, "info")
